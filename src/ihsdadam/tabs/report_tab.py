"""Report Generation tab -- build HTML crash prediction reports from Data Compiler output."""

import os
import webbrowser
from typing import Dict, List, Optional, Tuple

import openpyxl
from PySide6.QtCore import Qt, Signal, QMimeData
from PySide6.QtGui import QDrag
from PySide6.QtWidgets import (
    QAbstractItemView,
    QFileDialog,
    QFrame,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QListWidget,
    QListWidgetItem,
    QMessageBox,
    QPushButton,
    QRadioButton,
    QScrollArea,
    QTableWidget,
    QTableWidgetItem,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from .. import theme
from ..report_engine import Report, generic_bar, severity_bar


# ── Excel column indices (0-based, matching Data Compiler headers) ───────────

# Highway sheet (15 cols)
_HWY_EVAL = 0
_HWY_SEGMENT = 1
_HWY_YEAR = 3
_HWY_TYPE = 6
_HWY_LENGTH = 7
_HWY_K = 8
_HWY_A = 9
_HWY_B = 10
_HWY_C = 11
_HWY_PDO = 12

# Intersection / RampTerminal / SiteSet sheets (13 cols)
_INT_TYPE = 2
_INT_TITLE = 3
_INT_YEAR = 4
_INT_K = 7
_INT_A = 8
_INT_B = 9
_INT_C = 10
_INT_O = 11


# ── Data reader functions ────────────────────────────────────────────────────


def _cell(row: tuple, idx: int):
    """Safely access a tuple index -- returns None if out of range."""
    return row[idx] if idx < len(row) else None


def _fv(val) -> float:
    """Safely convert cell value to float."""
    if val is None:
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def read_highway_rows(wb: openpyxl.Workbook) -> List[dict]:
    """Read Highway sheet into list of dicts.  FI = K+A+B+C (calculated)."""
    if "Highway" not in wb.sheetnames:
        return []
    ws = wb["Highway"]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or _cell(row, 0) is None:
            continue
        k = _fv(_cell(row, _HWY_K))
        a = _fv(_cell(row, _HWY_A))
        b = _fv(_cell(row, _HWY_B))
        c = _fv(_cell(row, _HWY_C))
        pdo = _fv(_cell(row, _HWY_PDO))
        rows.append({
            "eval_name": _cell(row, _HWY_EVAL),
            "segment": _cell(row, _HWY_SEGMENT),
            "year": _cell(row, _HWY_YEAR),
            "type": str(_cell(row, _HWY_TYPE) or "").strip(),
            "length": _fv(_cell(row, _HWY_LENGTH)),
            "K": k, "A": a, "B": b, "C": c,
            "PDO": pdo,
            "FI": k + a + b + c,
        })
    return rows


def read_int_rows(wb: openpyxl.Workbook, sheet: str = "Intersection") -> List[dict]:
    """Read Intersection or RampTerminal sheet.  FI = K+A+B+C (calculated)."""
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or _cell(row, 0) is None:
            continue
        k = _fv(_cell(row, _INT_K))
        a = _fv(_cell(row, _INT_A))
        b = _fv(_cell(row, _INT_B))
        c = _fv(_cell(row, _INT_C))
        o = _fv(_cell(row, _INT_O))
        rows.append({
            "type": str(_cell(row, _INT_TYPE) or "").strip(),
            "title": _cell(row, _INT_TITLE),
            "year": _cell(row, _INT_YEAR),
            "K": k, "A": a, "B": b, "C": c,
            "O": o,
            "FI": k + a + b + c,
        })
    return rows


def scan_highway_types(excel_path: str) -> List[str]:
    """Read the Highway sheet and return sorted unique type codes found."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    types = set()
    if "Highway" in wb.sheetnames:
        ws = wb["Highway"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or _cell(row, 0) is None:
                continue
            t = str(_cell(row, _HWY_TYPE) or "").strip().upper()
            if t:
                types.add(t)
    wb.close()
    return sorted(types)


def group_highway_by_type(
    rows: List[dict], groups: Dict[str, List[str]]
) -> Tuple[Dict[str, dict], Dict[str, dict]]:
    """Group highway rows by functional class type.

    Returns (grouped_kabco, grouped_fi) where each is
    {group_name: {K, A, B, C, PD, FI, PDO, L, Total}}.
    """
    kabco = {}
    fi = {}
    for group_name, type_codes in groups.items():
        codes = {c.strip().upper() for c in type_codes}
        matching = [r for r in rows if r["type"].upper() in codes]
        kabco[group_name] = {
            "name": group_name,
            "K": sum(r["K"] for r in matching),
            "A": sum(r["A"] for r in matching),
            "B": sum(r["B"] for r in matching),
            "C": sum(r["C"] for r in matching),
            "PD": sum(r["PDO"] for r in matching),
            "L": sum(r["length"] for r in matching),
        }
        kabco[group_name]["Total"] = (
            kabco[group_name]["K"]
            + kabco[group_name]["A"]
            + kabco[group_name]["B"]
            + kabco[group_name]["C"]
            + kabco[group_name]["PD"]
        )
        fi[group_name] = {
            "name": group_name,
            "FI": sum(r["FI"] for r in matching),
            "PDO": sum(r["PDO"] for r in matching),
            "L": sum(r["length"] for r in matching),
        }
        fi[group_name]["Total"] = fi[group_name]["FI"] + fi[group_name]["PDO"]
    return kabco, fi


def summarize_int_rows_kabco(rows: List[dict]) -> dict:
    """Summarize intersection/ramp rows for KABCO mode."""
    return {
        "K": sum(r["K"] for r in rows),
        "A": sum(r["A"] for r in rows),
        "B": sum(r["B"] for r in rows),
        "C": sum(r["C"] for r in rows),
        "PD": sum(r["O"] for r in rows),
        "Total": sum(r["K"] + r["A"] + r["B"] + r["C"] + r["O"] for r in rows),
    }


def summarize_int_rows_fi(rows: List[dict]) -> dict:
    """Summarize intersection/ramp rows for FI/PDO mode."""
    return {
        "FI": sum(r["FI"] for r in rows),
        "PDO": sum(r["O"] for r in rows),
        "Total": sum(r["FI"] + r["O"] for r in rows),
    }


def scan_filter_items(excel_path: str) -> Dict[str, List[str]]:
    """Scan an Excel file and return unique identifiers per data category."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    items: Dict[str, set] = {
        "Highway Alignments": set(),
        "Intersections": set(),
        "Ramp Terminals": set(),
        "Site Set Intersections": set(),
        "Site Set Ramp Terminals": set(),
    }
    if "Highway" in wb.sheetnames:
        for row in wb["Highway"].iter_rows(min_row=2, values_only=True):
            if row and _cell(row, 0) is not None:
                name = str(_cell(row, _HWY_SEGMENT) or "").strip()
                if name:
                    items["Highway Alignments"].add(name)
    for sheet, cat in [
        ("Intersection", "Intersections"),
        ("RampTerminal", "Ramp Terminals"),
        ("SiteSet_Int", "Site Set Intersections"),
        ("SiteSet_Ramp", "Site Set Ramp Terminals"),
    ]:
        if sheet in wb.sheetnames:
            for row in wb[sheet].iter_rows(min_row=2, values_only=True):
                if row and _cell(row, 0) is not None:
                    name = str(_cell(row, _INT_TITLE) or "").strip()
                    if name:
                        items[cat].add(name)
    wb.close()
    return {k: sorted(v) for k, v in items.items()}


def load_project_data(excel_path: str) -> dict:
    """Load all sheets from a Data Compiler Excel file."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    data = {
        "highway_rows": read_highway_rows(wb),
        "int_rows": read_int_rows(wb, "Intersection"),
        "ramp_rows": read_int_rows(wb, "RampTerminal"),
        "ss_int_rows": read_int_rows(wb, "SiteSet_Int"),
        "ss_ramp_rows": read_int_rows(wb, "SiteSet_Ramp"),
    }
    wb.close()
    return data


# ── Severity legend snippets ─────────────────────────────────────────────────

_KABCO_LEGEND = (
    '<span style="font-style:normal;display:inline-flex;gap:14px;font-size:0.52rem;font-weight:600;color:#6b7a90;">'
    '<span style="display:inline-flex;align-items:center;gap:3px;"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:#fd281b;"></span>K \u2013 Fatal</span>'
    '<span style="display:inline-flex;align-items:center;gap:3px;"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:#fcad32;"></span>A \u2013 Serious Injury</span>'
    '<span style="display:inline-flex;align-items:center;gap:3px;"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:#fbff47;"></span>B \u2013 Minor Injury</span>'
    '<span style="display:inline-flex;align-items:center;gap:3px;"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:#316bf9;"></span>C \u2013 Possible Injury</span>'
    '<span style="display:inline-flex;align-items:center;gap:3px;"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:#4dfe41;"></span>PD \u2013 Property Damage</span>'
    "</span>"
)

_FI_LEGEND = (
    '<span style="font-style:normal;display:inline-flex;gap:14px;font-size:0.52rem;font-weight:600;color:#6b7a90;">'
    '<span style="display:inline-flex;align-items:center;gap:3px;"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:#dc2626;"></span>Fatal / Injury</span>'
    '<span style="display:inline-flex;align-items:center;gap:3px;"><span style="display:inline-block;width:8px;height:8px;border-radius:2px;background:#16a34a;"></span>Property Damage Only</span>'
    "</span>"
)

_KABCO_COLS = [
    {"key": "name", "header": "Category"},
    {"key": "L", "header": "Length (mi)", "decimals": 1},
    {"key": "K", "header": "K", "decimals": 2},
    {"key": "A", "header": "A", "decimals": 2},
    {"key": "B", "header": "B", "decimals": 2},
    {"key": "C", "header": "C", "decimals": 2},
    {"key": "PD", "header": "PD", "decimals": 2},
    {"key": "Total", "header": "Total", "decimals": 2},
]

_KABCO_COLS_NO_LEN = [
    {"key": "name", "header": "Location"},
    {"key": "K", "header": "K", "decimals": 2},
    {"key": "A", "header": "A", "decimals": 2},
    {"key": "B", "header": "B", "decimals": 2},
    {"key": "C", "header": "C", "decimals": 2},
    {"key": "PD", "header": "PD", "decimals": 2},
    {"key": "Total", "header": "Total", "decimals": 2},
]

_FI_COLS = [
    {"key": "name", "header": "Category"},
    {"key": "L", "header": "Length (mi)", "decimals": 1},
    {"key": "FI", "header": "Fatal/Injury", "decimals": 2},
    {"key": "PDO", "header": "Prop. Damage", "decimals": 2},
    {"key": "Total", "header": "Total", "decimals": 2},
]

_FI_COLS_NO_LEN = [
    {"key": "name", "header": "Location"},
    {"key": "FI", "header": "Fatal/Injury", "decimals": 2},
    {"key": "PDO", "header": "Prop. Damage", "decimals": 2},
    {"key": "Total", "header": "Total", "decimals": 2},
]


# ── Drag-and-drop widgets for segment grouping ───────────────────────────────


class _DragListWidget(QListWidget):
    """QListWidget that provides text mime data when items are dragged."""

    def startDrag(self, supportedActions):
        items = self.selectedItems()
        if not items:
            return
        mime = QMimeData()
        mime.setText("\n".join(item.text() for item in items))
        drag = QDrag(self)
        drag.setMimeData(mime)
        result = drag.exec(Qt.DropAction.MoveAction)
        if result == Qt.DropAction.MoveAction:
            for item in items:
                row = self.row(item)
                if row >= 0:
                    self.takeItem(row)


class _DropTreeWidget(QTreeWidget):
    """QTreeWidget that accepts drops from the available-types QListWidget."""

    types_dropped = Signal(list)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)

    def dragEnterEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()
        else:
            super().dragEnterEvent(event)

    def dragMoveEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()
        else:
            super().dragMoveEvent(event)

    def dropEvent(self, event):
        if not event.mimeData().hasText():
            super().dropEvent(event)
            return

        codes = [c.strip() for c in event.mimeData().text().split("\n") if c.strip()]
        if not codes:
            return

        # Find target group
        target_item = self.itemAt(event.position().toPoint())
        if target_item is None:
            # No target -- if there are groups, use the first one
            if self.topLevelItemCount() == 0:
                return
            target_item = self.topLevelItem(0)

        group_item = target_item if target_item.parent() is None else target_item.parent()

        # Check for duplicates within this group
        existing = {group_item.child(j).text(0) for j in range(group_item.childCount())}
        for code in codes:
            if code not in existing:
                QTreeWidgetItem(group_item, [code])
                existing.add(code)

        self.expandItem(group_item)
        event.acceptProposedAction()
        self.types_dropped.emit(codes)


# ── Tab class ────────────────────────────────────────────────────────────────


class ReportTab(QWidget):
    """Report Generation tab."""

    status_message = Signal(str)
    progress_update = Signal(int, str)

    _DEFAULT_GROUPS = ["Mainline", "System Ramps", "Ramps", "Arterials"]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._project_path = ""
        self._last_output = ""
        self._all_discovered_types: List[str] = []
        self._setup_ui()
        self._add_default_groups()

    def set_project_path(self, path: str):
        self._project_path = path

    # ── UI ────────────────────────────────────────────────────────────────

    def _setup_ui(self):
        root = QHBoxLayout(self)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(10)

        # ── Left: config panel (scrollable) ──────────────────────────────
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.Shape.NoFrame)

        config_widget = QWidget()
        config = QVBoxLayout(config_widget)
        config.setSpacing(10)

        # -- Report mode --
        mode_box = QGroupBox("Report Mode")
        mode_layout = QVBoxLayout(mode_box)
        self._single_radio = QRadioButton("Single Project")
        self._single_radio.setChecked(True)
        self._multi_radio = QRadioButton("Multi-Project Comparison")
        self._single_radio.toggled.connect(self._on_mode_changed)
        mode_layout.addWidget(self._single_radio)
        mode_layout.addWidget(self._multi_radio)
        config.addWidget(mode_box)

        # -- Severity breakdown --
        sev_box = QGroupBox("Severity Breakdown")
        sev_layout = QVBoxLayout(sev_box)
        self._kabco_radio = QRadioButton("KABCO (K, A, B, C, PDO)")
        self._kabco_radio.setChecked(True)
        self._fi_radio = QRadioButton("FI vs PDO (Fatal+Injury vs Property Damage)")
        sev_layout.addWidget(self._kabco_radio)
        sev_layout.addWidget(self._fi_radio)
        config.addWidget(sev_box)

        # -- Project files --
        files_box = QGroupBox("Project Files")
        files_layout = QVBoxLayout(files_box)

        # Single file row
        self._single_file_widget = QWidget()
        sf_layout = QHBoxLayout(self._single_file_widget)
        sf_layout.setContentsMargins(0, 0, 0, 0)
        sf_layout.addWidget(QLabel("Excel File:"))
        self._single_path = QLineEdit()
        self._single_path.setPlaceholderText("Data Compiler .xlsx output...")
        sf_layout.addWidget(self._single_path, stretch=1)
        single_browse = QPushButton("Browse...")
        single_browse.clicked.connect(self._browse_single)
        sf_layout.addWidget(single_browse)
        files_layout.addWidget(self._single_file_widget)

        # Multi file list
        self._multi_file_widget = QWidget()
        mf_layout = QVBoxLayout(self._multi_file_widget)
        mf_layout.setContentsMargins(0, 0, 0, 0)

        mf_hint = QLabel(
            "Double-click the Label column to rename each alternative "
            "(e.g., 'No Build', 'Build Alt 1')."
        )
        mf_hint.setWordWrap(True)
        mf_hint.setStyleSheet(
            f"color: {theme.TEXT_SECONDARY}; font-size: 9pt; font-style: italic;"
        )
        mf_layout.addWidget(mf_hint)

        self._file_table = QTableWidget(0, 2)
        self._file_table.setHorizontalHeaderLabels(["Label", "File Path"])
        self._file_table.horizontalHeader().setSectionResizeMode(
            0, QHeaderView.ResizeMode.Interactive
        )
        self._file_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        self._file_table.setColumnWidth(0, 140)
        self._file_table.setMinimumHeight(120)
        mf_layout.addWidget(self._file_table)
        btn_row = QHBoxLayout()
        add_btn = QPushButton("Add Files...")
        add_btn.clicked.connect(self._add_multi_files)
        btn_row.addWidget(add_btn)
        remove_btn = QPushButton("Remove Selected")
        remove_btn.clicked.connect(self._remove_multi_file)
        btn_row.addWidget(remove_btn)
        btn_row.addStretch()
        mf_layout.addLayout(btn_row)
        files_layout.addWidget(self._multi_file_widget)
        self._multi_file_widget.setVisible(False)

        config.addWidget(files_box)

        # -- Segment grouping (drag-and-drop approach) --
        group_box = QGroupBox("Highway Segment Grouping")
        group_layout = QVBoxLayout(group_box)

        group_hint = QLabel(
            "Group your modeled segment types under a display name for the "
            "report. The group name is what appears in the output. The four "
            "default groups are suggestions and can be renamed or changed."
        )
        group_hint.setWordWrap(True)
        group_hint.setStyleSheet(
            f"color: {theme.TEXT_SECONDARY}; font-size: 9pt; font-style: italic;"
        )
        group_layout.addWidget(group_hint)

        self._types_status = QLabel(
            "Load project files to discover available segment types."
        )
        self._types_status.setWordWrap(True)
        self._types_status.setStyleSheet(
            f"color: {theme.TEXT_SECONDARY}; font-size: 9pt; font-style: italic;"
        )
        group_layout.addWidget(self._types_status)

        panels = QHBoxLayout()

        # Left: available (unassigned) types
        left_panel = QVBoxLayout()
        left_panel.addWidget(QLabel("Available Types:"))
        self._avail_types = _DragListWidget()
        self._avail_types.setSelectionMode(
            QListWidget.SelectionMode.ExtendedSelection
        )
        self._avail_types.setDragEnabled(True)
        self._avail_types.setMinimumHeight(140)
        left_panel.addWidget(self._avail_types)
        panels.addLayout(left_panel)

        # Center: transfer buttons (clear text labels)
        center_btns = QVBoxLayout()
        center_btns.addStretch()
        assign_btn = QPushButton("Add to Group  \u2192")
        assign_btn.setToolTip("Add selected types to the selected group")
        assign_btn.clicked.connect(self._assign_to_group)
        center_btns.addWidget(assign_btn)
        unassign_btn = QPushButton("\u2190  Remove")
        unassign_btn.setToolTip("Remove selected type from its group")
        unassign_btn.clicked.connect(self._unassign_from_group)
        center_btns.addWidget(unassign_btn)
        center_btns.addStretch()
        panels.addLayout(center_btns)

        # Right: groups tree
        right_panel = QVBoxLayout()
        right_panel.addWidget(QLabel("Groups:"))
        self._group_tree = _DropTreeWidget()
        self._group_tree.setHeaderHidden(True)
        self._group_tree.setMinimumHeight(140)
        self._group_tree.setAcceptDrops(True)
        self._group_tree.setDragDropMode(
            QAbstractItemView.DragDropMode.DropOnly
        )
        self._group_tree.types_dropped.connect(self._on_types_dropped)
        right_panel.addWidget(self._group_tree)

        grp_btn_row = QHBoxLayout()
        new_grp = QPushButton("New Group")
        new_grp.clicked.connect(self._new_group)
        grp_btn_row.addWidget(new_grp)
        del_grp = QPushButton("Delete Group")
        del_grp.clicked.connect(self._delete_group)
        grp_btn_row.addWidget(del_grp)
        up_btn = QPushButton("Move Up")
        up_btn.clicked.connect(self._move_group_up)
        grp_btn_row.addWidget(up_btn)
        down_btn = QPushButton("Move Down")
        down_btn.clicked.connect(self._move_group_down)
        grp_btn_row.addWidget(down_btn)
        grp_btn_row.addStretch()
        right_panel.addLayout(grp_btn_row)
        panels.addLayout(right_panel)

        group_layout.addLayout(panels)
        config.addWidget(group_box)

        # -- Data filter (advanced) --
        filter_box = QGroupBox("Data Filter (Advanced)")
        filter_layout = QVBoxLayout(filter_box)

        filter_warn = QLabel(
            "For advanced users only. Uncheck items to exclude them from "
            "the report. All items are included by default."
        )
        filter_warn.setWordWrap(True)
        filter_warn.setStyleSheet(
            f"color: {theme.WARNING_FG}; font-size: 8pt; font-style: italic;"
        )
        filter_layout.addWidget(filter_warn)

        self._filter_tree = QTreeWidget()
        self._filter_tree.setHeaderHidden(True)
        self._filter_tree.setMinimumHeight(120)
        self._filter_tree.itemChanged.connect(self._on_filter_item_changed)
        filter_layout.addWidget(self._filter_tree)

        self._filter_status = QLabel("Load project files to populate filter.")
        self._filter_status.setStyleSheet(
            f"color: {theme.TEXT_SECONDARY}; font-size: 9pt; font-style: italic;"
        )
        filter_layout.addWidget(self._filter_status)
        config.addWidget(filter_box)

        # -- Report fields --
        fields_box = QGroupBox("Report Fields")
        fields_layout = QVBoxLayout(fields_box)
        fields_layout.setSpacing(6)

        fields_layout.addWidget(QLabel("Project Title:"))
        self._title_input = QLineEdit()
        self._title_input.setPlaceholderText("e.g., Blatnik Bridge Replacement")
        fields_layout.addWidget(self._title_input)

        fields_layout.addWidget(QLabel("Subtitle / Modeling Scenario:"))
        self._subtitle_input = QLineEdit()
        self._subtitle_input.setPlaceholderText(
            "e.g., IHSDM Crash Prediction Summary"
        )
        fields_layout.addWidget(self._subtitle_input)

        fields_layout.addWidget(QLabel("Evaluation Years:"))
        self._years_input = QLineEdit()
        self._years_input.setPlaceholderText("e.g., 2032-2041")
        fields_layout.addWidget(self._years_input)

        fields_layout.addWidget(QLabel("Analyst Name:"))
        self._analyst_input = QLineEdit()
        self._analyst_input.setPlaceholderText(
            "For 'Printed and reviewed by' footer"
        )
        fields_layout.addWidget(self._analyst_input)

        config.addWidget(fields_box)

        # -- Logo --
        logo_box = QGroupBox("Logo (optional)")
        logo_layout = QHBoxLayout(logo_box)
        self._logo_input = QLineEdit()
        self._logo_input.setPlaceholderText("Default: HNTB logo (leave blank to use)")
        logo_layout.addWidget(self._logo_input, stretch=1)
        logo_browse = QPushButton("Browse...")
        logo_browse.clicked.connect(self._browse_logo)
        logo_layout.addWidget(logo_browse)
        config.addWidget(logo_box)

        config.addStretch()
        scroll.setWidget(config_widget)
        scroll.setMinimumWidth(420)
        root.addWidget(scroll, stretch=2)

        # ── Right: output panel ──────────────────────────────────────────
        right = QVBoxLayout()
        right.setSpacing(10)

        # -- Warning banner --
        warning_frame = QFrame()
        warning_frame.setStyleSheet(
            f"background-color: {theme.WARNING_BG};"
            f" border: 1px solid {theme.WARNING};"
            " border-radius: 6px;"
            " padding: 10px;"
        )
        warning_layout = QVBoxLayout(warning_frame)
        warning_layout.setContentsMargins(12, 10, 12, 10)
        warning_label = QLabel(
            "WARNING: For use with IHSDadaM Data Compiler output files ONLY. "
            "Do not modify output files in any way. Investigate data outputs "
            "at every step to ensure quality and accurate results. "
            "All summary documents must go through formal QA/QC review "
            "before distribution to a client."
        )
        warning_label.setWordWrap(True)
        warning_label.setStyleSheet(
            f"color: {theme.WARNING_FG};"
            " font-weight: 700;"
            " font-size: 9pt;"
            " background-color: transparent;"
            " border: none;"
            " padding: 0;"
        )
        warning_layout.addWidget(warning_label)
        right.addWidget(warning_frame)

        # -- Output controls --
        output_box = QGroupBox("Output")
        output_layout = QVBoxLayout(output_box)

        output_layout.addWidget(QLabel("Output HTML File:"))
        out_row = QHBoxLayout()
        self._output_input = QLineEdit("report_output.html")
        out_row.addWidget(self._output_input, stretch=1)
        out_browse = QPushButton("Browse...")
        out_browse.clicked.connect(self._browse_output)
        out_row.addWidget(out_browse)
        output_layout.addLayout(out_row)

        self._generate_btn = QPushButton("Generate Report")
        self._generate_btn.setProperty("accent", True)
        self._generate_btn.clicked.connect(self._generate_report)
        output_layout.addWidget(self._generate_btn)

        self._open_btn = QPushButton("Open in Browser")
        self._open_btn.setEnabled(False)
        self._open_btn.clicked.connect(self._open_in_browser)
        output_layout.addWidget(self._open_btn)

        self._status_label = QLabel("Configure report settings and click Generate.")
        self._status_label.setWordWrap(True)
        self._status_label.setStyleSheet(f"color: {theme.TEXT_SECONDARY};")
        output_layout.addWidget(self._status_label)

        output_layout.addStretch()
        right.addWidget(output_box, stretch=1)
        root.addLayout(right, stretch=1)

    # ── Mode switching ────────────────────────────────────────────────────

    def _on_mode_changed(self):
        is_single = self._single_radio.isChecked()
        self._single_file_widget.setVisible(is_single)
        self._multi_file_widget.setVisible(not is_single)

    # ── File browsing ─────────────────────────────────────────────────────

    def _browse_single(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Data Compiler Excel File", "",
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if path:
            self._single_path.setText(path)
            self._scan_loaded_files()

    def _add_multi_files(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Select Data Compiler Excel Files", "",
            "Excel Files (*.xlsx);;All Files (*)",
        )
        for path in paths:
            row = self._file_table.rowCount()
            self._file_table.insertRow(row)
            label = os.path.splitext(os.path.basename(path))[0]
            self._file_table.setItem(row, 0, QTableWidgetItem(label))
            path_item = QTableWidgetItem(path)
            path_item.setFlags(path_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            self._file_table.setItem(row, 1, path_item)
        if paths:
            self._scan_loaded_files()

    def _remove_multi_file(self):
        rows = sorted(
            {idx.row() for idx in self._file_table.selectedIndexes()},
            reverse=True,
        )
        for r in rows:
            self._file_table.removeRow(r)

    def _browse_output(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Select Output File", "report_output.html",
            "HTML Files (*.html);;All Files (*)",
        )
        if path:
            self._output_input.setText(path)

    def _browse_logo(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Select Logo Image", "",
            "Images (*.png *.jpg *.jpeg *.svg);;All Files (*)",
        )
        if path:
            self._logo_input.setText(path)

    # ── Segment grouping ─────────────────────────────────────────────────

    def _scan_loaded_files(self):
        """Auto-scan all loaded Excel files for highway type codes and filter items."""
        all_types = set()
        all_filter_items: Dict[str, set] = {}
        paths_to_scan = []

        if self._single_radio.isChecked():
            path = self._single_path.text().strip()
            if path and os.path.isfile(path):
                paths_to_scan.append(path)
        else:
            for i in range(self._file_table.rowCount()):
                item = self._file_table.item(i, 1)
                if item:
                    path = item.text().strip()
                    if path and os.path.isfile(path):
                        paths_to_scan.append(path)

        for path in paths_to_scan:
            try:
                all_types.update(scan_highway_types(path))
            except Exception:
                pass
            try:
                items = scan_filter_items(path)
                for cat, names in items.items():
                    if cat not in all_filter_items:
                        all_filter_items[cat] = set()
                    all_filter_items[cat].update(names)
            except Exception:
                pass

        self._all_discovered_types = sorted(all_types)
        self._refresh_available_types()

        n = len(self._all_discovered_types)
        if n:
            self._types_status.setText(
                f"Found {n} type(s): {', '.join(self._all_discovered_types)}"
            )
        else:
            self._types_status.setText("No highway types found in loaded files.")

        self._populate_filter_tree(all_filter_items)

    def _refresh_available_types(self):
        """Show only types not already assigned to a group."""
        assigned = set()
        for i in range(self._group_tree.topLevelItemCount()):
            grp = self._group_tree.topLevelItem(i)
            for j in range(grp.childCount()):
                assigned.add(grp.child(j).text(0))

        self._avail_types.clear()
        for t in self._all_discovered_types:
            if t not in assigned:
                self._avail_types.addItem(t)

    def _new_group(self):
        name, ok = QInputDialog.getText(self, "New Group", "Group name:")
        if ok and name.strip():
            item = QTreeWidgetItem([name.strip()])
            self._group_tree.addTopLevelItem(item)
            self._group_tree.expandItem(item)
            self._group_tree.setCurrentItem(item)

    def _delete_group(self):
        current = self._group_tree.currentItem()
        if not current:
            return
        target = current if current.parent() is None else current.parent()
        idx = self._group_tree.indexOfTopLevelItem(target)
        if idx >= 0:
            self._group_tree.takeTopLevelItem(idx)
            self._refresh_available_types()

    def _assign_to_group(self):
        """Move selected available types into the selected group."""
        selected = [item.text() for item in self._avail_types.selectedItems()]
        if not selected:
            return

        current = self._group_tree.currentItem()
        if current is None:
            QMessageBox.information(
                self, "No Group", "Select or create a group first."
            )
            return
        group_item = current if current.parent() is None else current.parent()

        for t in selected:
            QTreeWidgetItem(group_item, [t])

        self._group_tree.expandItem(group_item)
        self._refresh_available_types()

    def _unassign_from_group(self):
        """Remove selected type from its group back to available."""
        current = self._group_tree.currentItem()
        if current and current.parent() is not None:
            parent = current.parent()
            parent.removeChild(current)
            self._refresh_available_types()

    def _on_types_dropped(self, type_codes: list):
        """Handle types dropped onto the group tree from the available list."""
        self._refresh_available_types()

    def _move_group_up(self):
        current = self._group_tree.currentItem()
        if not current:
            return
        target = current if current.parent() is None else current.parent()
        idx = self._group_tree.indexOfTopLevelItem(target)
        if idx <= 0:
            return
        item = self._group_tree.takeTopLevelItem(idx)
        self._group_tree.insertTopLevelItem(idx - 1, item)
        item.setExpanded(True)
        self._group_tree.setCurrentItem(item)

    def _move_group_down(self):
        current = self._group_tree.currentItem()
        if not current:
            return
        target = current if current.parent() is None else current.parent()
        idx = self._group_tree.indexOfTopLevelItem(target)
        if idx < 0 or idx >= self._group_tree.topLevelItemCount() - 1:
            return
        item = self._group_tree.takeTopLevelItem(idx)
        self._group_tree.insertTopLevelItem(idx + 1, item)
        item.setExpanded(True)
        self._group_tree.setCurrentItem(item)

    def _add_default_groups(self):
        """Pre-populate the group tree with default group names."""
        for name in self._DEFAULT_GROUPS:
            item = QTreeWidgetItem([name])
            self._group_tree.addTopLevelItem(item)

    def _get_groups(self) -> Dict[str, List[str]]:
        """Read groups from the tree widget."""
        groups = {}
        for i in range(self._group_tree.topLevelItemCount()):
            grp = self._group_tree.topLevelItem(i)
            name = grp.text(0).strip()
            codes = [grp.child(j).text(0) for j in range(grp.childCount())]
            if name and codes:
                groups[name] = codes
        return groups

    # ── Data filter ────────────────────────────────────────────────────────

    _FILTER_CATEGORIES = [
        "Highway Alignments",
        "Intersections",
        "Ramp Terminals",
        "Site Set Intersections",
        "Site Set Ramp Terminals",
    ]

    def _populate_filter_tree(self, items: Dict[str, set]):
        """Rebuild the filter tree with checkboxes for each data item."""
        self._filter_tree.blockSignals(True)
        self._filter_tree.clear()
        total = 0
        for cat in self._FILTER_CATEGORIES:
            names = sorted(items.get(cat, set()))
            if not names:
                continue
            parent = QTreeWidgetItem([cat])
            parent.setFlags(parent.flags() | Qt.ItemFlag.ItemIsUserCheckable)
            parent.setCheckState(0, Qt.CheckState.Checked)
            for name in names:
                child = QTreeWidgetItem([name])
                child.setFlags(child.flags() | Qt.ItemFlag.ItemIsUserCheckable)
                child.setCheckState(0, Qt.CheckState.Checked)
                parent.addChild(child)
                total += 1
            self._filter_tree.addTopLevelItem(parent)
            parent.setExpanded(False)
        self._filter_tree.blockSignals(False)

        if total:
            self._filter_status.setText(f"{total} items found. Uncheck to exclude.")
        else:
            self._filter_status.setText("No items found in loaded files.")

    def _on_filter_item_changed(self, item, column):
        """Propagate parent check state to children and vice versa."""
        self._filter_tree.blockSignals(True)
        if item.childCount() > 0:
            # Parent toggled -- set all children to match
            state = item.checkState(0)
            for i in range(item.childCount()):
                item.child(i).setCheckState(0, state)
        elif item.parent() is not None:
            # Child toggled -- update parent state
            parent = item.parent()
            checked = sum(
                1 for i in range(parent.childCount())
                if parent.child(i).checkState(0) == Qt.CheckState.Checked
            )
            if checked == parent.childCount():
                parent.setCheckState(0, Qt.CheckState.Checked)
            elif checked == 0:
                parent.setCheckState(0, Qt.CheckState.Unchecked)
            else:
                parent.setCheckState(0, Qt.CheckState.PartiallyChecked)
        self._filter_tree.blockSignals(False)

    def _get_excluded(self) -> Dict[str, set]:
        """Return dict of {category: set(excluded_names)} from the filter tree."""
        excluded: Dict[str, set] = {}
        for i in range(self._filter_tree.topLevelItemCount()):
            parent = self._filter_tree.topLevelItem(i)
            cat = parent.text(0)
            unchecked = set()
            for j in range(parent.childCount()):
                child = parent.child(j)
                if child.checkState(0) != Qt.CheckState.Checked:
                    unchecked.add(child.text(0))
            if unchecked:
                excluded[cat] = unchecked
        return excluded

    # ── Report generation ─────────────────────────────────────────────────

    def _generate_report(self):
        output_path = self._output_input.text().strip()
        if not output_path:
            QMessageBox.warning(self, "No Output", "Please set an output file path.")
            return

        is_single = self._single_radio.isChecked()
        is_kabco = self._kabco_radio.isChecked()

        # Collect file entries
        if is_single:
            path = self._single_path.text().strip()
            if not path or not os.path.isfile(path):
                QMessageBox.warning(
                    self, "No File", "Please select a valid Excel file."
                )
                return
            file_entries = [("Project", path)]
        else:
            file_entries = []
            for i in range(self._file_table.rowCount()):
                label_item = self._file_table.item(i, 0)
                path_item = self._file_table.item(i, 1)
                label = (label_item.text().strip() if label_item else "")
                fpath = (path_item.text().strip() if path_item else "")
                if fpath and os.path.isfile(fpath):
                    file_entries.append((label or f"Alt {i + 1}", fpath))
            if len(file_entries) < 2:
                QMessageBox.warning(
                    self,
                    "Need Files",
                    "Multi-project mode requires at least 2 valid Excel files.",
                )
                return

        groups = self._get_groups()
        excluded = self._get_excluded()
        title = self._title_input.text().strip() or "IHSDM Crash Prediction Report"
        subtitle = self._subtitle_input.text().strip() or "Crash Prediction Summary"
        years = self._years_input.text().strip()
        analyst = self._analyst_input.text().strip()
        logo_path = self._logo_input.text().strip() or None

        footer_text = ""
        if analyst:
            footer_text = f"Printed and reviewed by {analyst}"

        try:
            self.status_message.emit("Generating report...")
            self._generate_btn.setEnabled(False)

            if is_single:
                html_path = self._build_single_report(
                    file_entries[0][1], groups, excluded, is_kabco,
                    title, subtitle, years, logo_path, footer_text, output_path,
                )
            else:
                html_path = self._build_multi_report(
                    file_entries, groups, excluded, is_kabco,
                    title, subtitle, years, logo_path, footer_text, output_path,
                )

            self._last_output = html_path
            self._open_btn.setEnabled(True)
            self._status_label.setText(f"Report generated: {html_path}")
            self._status_label.setStyleSheet(f"color: {theme.ACCENT_GREEN};")
            self.status_message.emit("Report generated successfully")
            QMessageBox.information(
                self, "Done", f"Report saved to:\n{html_path}"
            )
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Report generation failed:\n{exc}")
            self.status_message.emit("Report generation failed")
        finally:
            self._generate_btn.setEnabled(True)

    def _open_in_browser(self):
        if self._last_output and os.path.isfile(self._last_output):
            webbrowser.open(f"file:///{self._last_output.replace(os.sep, '/')}")

    # ── Single project report ─────────────────────────────────────────────

    @staticmethod
    def _apply_filter(data: dict, excluded: Dict[str, set]) -> dict:
        """Return a filtered copy of project data based on excluded items."""
        hwy_ex = excluded.get("Highway Alignments", set())
        int_ex = excluded.get("Intersections", set())
        ramp_ex = excluded.get("Ramp Terminals", set())
        ss_int_ex = excluded.get("Site Set Intersections", set())
        ss_ramp_ex = excluded.get("Site Set Ramp Terminals", set())
        return {
            "highway_rows": [r for r in data["highway_rows"]
                             if str(r.get("segment") or "").strip() not in hwy_ex],
            "int_rows": [r for r in data["int_rows"]
                         if str(r.get("title") or "").strip() not in int_ex],
            "ramp_rows": [r for r in data["ramp_rows"]
                          if str(r.get("title") or "").strip() not in ramp_ex],
            "ss_int_rows": [r for r in data["ss_int_rows"]
                            if str(r.get("title") or "").strip() not in ss_int_ex],
            "ss_ramp_rows": [r for r in data["ss_ramp_rows"]
                             if str(r.get("title") or "").strip() not in ss_ramp_ex],
        }

    def _build_single_report(
        self,
        excel_path: str,
        groups: Dict[str, List[str]],
        excluded: Dict[str, set],
        is_kabco: bool,
        title: str,
        subtitle: str,
        years: str,
        logo_path: Optional[str],
        footer_text: str,
        output_path: str,
    ) -> str:
        data = self._apply_filter(load_project_data(excel_path), excluded)
        hwy_kabco, hwy_fi = group_highway_by_type(data["highway_rows"], groups)

        all_int = data["int_rows"] + data["ss_int_rows"]
        all_ramp = data["ramp_rows"] + data["ss_ramp_rows"]

        report = Report(
            title=title,
            subtitle=subtitle,
            project_id=f"Analysis: {years}" if years else "",
            logo_path=logo_path,
            footer_text=footer_text,
        )

        if years:
            report.add_note(f"Predicted crashes from IHSDM ({years})")

        if is_kabco:
            self._add_single_kabco(report, hwy_kabco, all_int, all_ramp, groups)
        else:
            self._add_single_fi(report, hwy_fi, all_int, all_ramp, groups)

        report.generate(output_path)
        return output_path

    def _add_single_kabco(self, report, hwy_groups, all_int, all_ramp, groups):
        int_sum = summarize_int_rows_kabco(all_int) if all_int else None
        ramp_sum = summarize_int_rows_kabco(all_ramp) if all_ramp else None

        # Grand totals
        grand = {"K": 0, "A": 0, "B": 0, "C": 0, "PD": 0, "Total": 0, "L": 0}
        summary_rows = []
        for group_name in groups:
            g = hwy_groups.get(group_name)
            if g:
                summary_rows.append(g)
                for k in ("K", "A", "B", "C", "PD", "Total", "L"):
                    grand[k] += g.get(k, 0)
        if int_sum:
            summary_rows.append({"name": "Intersections", "L": 0, **int_sum})
            for k in ("K", "A", "B", "C", "PD", "Total"):
                grand[k] += int_sum.get(k, 0)
        if ramp_sum:
            summary_rows.append({"name": "Ramp Terminals", "L": 0, **ramp_sum})
            for k in ("K", "A", "B", "C", "PD", "Total"):
                grand[k] += ramp_sum.get(k, 0)

        grand["name"] = "TOTAL"

        # Overall bar chart
        report.add_bar_chart([{
            "label": "Total Crashes",
            "value": grand["Total"],
            "segments": [
                (grand["K"], "k"), (grand["A"], "a"), (grand["B"], "b"),
                (grand["C"], "c"), (grand["PD"], "pd"),
            ],
        }])

        # Metric cards
        seg_total = grand["Total"] - (int_sum or {}).get("Total", 0) - (ramp_sum or {}).get("Total", 0)
        report.add_metric_cards([
            {"label": "Total Predicted", "value": f"{grand['Total']:.1f}", "style": "default"},
            {"label": "Segments", "value": f"{seg_total:.1f}", "style": "blue"},
            {"label": "Intersections", "value": f"{(int_sum or {}).get('Total', 0):.1f}", "style": "blue"},
            {"label": "Ramp Terminals", "value": f"{(ramp_sum or {}).get('Total', 0):.1f}", "style": "blue"},
        ])

        report.add_note(_KABCO_LEGEND)

        report.add_table(
            "Crash Prediction Summary",
            _KABCO_COLS, summary_rows, total_row=grand,
            distribution_fn=severity_bar,
        )

        # Intersection / ramp detail (aggregated by title)
        if all_int:
            self._add_int_detail_kabco(report, all_int, "Intersection Detail")
        if all_ramp:
            self._add_int_detail_kabco(report, all_ramp, "Ramp Terminal Detail")

    def _add_single_fi(self, report, hwy_groups, all_int, all_ramp, groups):
        int_sum = summarize_int_rows_fi(all_int) if all_int else None
        ramp_sum = summarize_int_rows_fi(all_ramp) if all_ramp else None

        grand = {"FI": 0, "PDO": 0, "Total": 0, "L": 0}
        summary_rows = []
        for group_name in groups:
            g = hwy_groups.get(group_name)
            if g:
                summary_rows.append(g)
                for k in ("FI", "PDO", "Total", "L"):
                    grand[k] += g.get(k, 0)
        if int_sum:
            summary_rows.append({"name": "Intersections", "L": 0, **int_sum})
            for k in ("FI", "PDO", "Total"):
                grand[k] += int_sum.get(k, 0)
        if ramp_sum:
            summary_rows.append({"name": "Ramp Terminals", "L": 0, **ramp_sum})
            for k in ("FI", "PDO", "Total"):
                grand[k] += ramp_sum.get(k, 0)

        grand["name"] = "TOTAL"

        # Overall bar chart
        report.add_bar_chart([{
            "label": "Total Crashes",
            "value": grand["Total"],
            "segments": [
                (grand["FI"], "c4"), (grand["PDO"], "c2"),
            ],
        }])

        seg_total = grand["Total"] - (int_sum or {}).get("Total", 0) - (ramp_sum or {}).get("Total", 0)
        report.add_metric_cards([
            {"label": "Total Predicted", "value": f"{grand['Total']:.1f}", "style": "default"},
            {"label": "Segments", "value": f"{seg_total:.1f}", "style": "blue"},
            {"label": "Intersections", "value": f"{(int_sum or {}).get('Total', 0):.1f}", "style": "blue"},
            {"label": "Ramp Terminals", "value": f"{(ramp_sum or {}).get('Total', 0):.1f}", "style": "blue"},
        ])

        def fi_bar(row):
            return generic_bar(row, ["FI", "PDO"], ["c4", "c2"])

        report.add_note(_FI_LEGEND)

        report.add_table(
            "Crash Prediction Summary",
            _FI_COLS, summary_rows, total_row=grand,
            distribution_fn=fi_bar,
        )

        # Intersection / ramp detail (aggregated by title)
        if all_int:
            self._add_int_detail_fi(report, all_int, "Intersection Detail")
        if all_ramp:
            self._add_int_detail_fi(report, all_ramp, "Ramp Terminal Detail")

    # ── Multi-project comparison report ───────────────────────────────────

    def _build_multi_report(
        self,
        file_entries: List[Tuple[str, str]],
        groups: Dict[str, List[str]],
        excluded: Dict[str, set],
        is_kabco: bool,
        title: str,
        subtitle: str,
        years: str,
        logo_path: Optional[str],
        footer_text: str,
        output_path: str,
    ) -> str:
        report = Report(
            title=title,
            subtitle=subtitle,
            project_id=f"Analysis: {years}" if years else "",
            logo_path=logo_path,
            footer_text=footer_text,
        )

        if years:
            report.add_note(f"Predicted crashes from IHSDM ({years})")

        # Load all projects
        project_summaries = []
        for label, fpath in file_entries:
            data = self._apply_filter(load_project_data(fpath), excluded)
            hwy_kabco, hwy_fi = group_highway_by_type(data["highway_rows"], groups)
            all_int = data["int_rows"] + data["ss_int_rows"]
            all_ramp = data["ramp_rows"] + data["ss_ramp_rows"]
            project_summaries.append({
                "label": label,
                "hwy_kabco": hwy_kabco,
                "hwy_fi": hwy_fi,
                "int_rows": all_int,
                "ramp_rows": all_ramp,
            })

        if is_kabco:
            self._add_multi_kabco(report, project_summaries, groups)
        else:
            self._add_multi_fi(report, project_summaries, groups)

        report.generate(output_path)
        return output_path

    def _add_multi_kabco(self, report, projects, groups):
        bars = []
        all_tables = []
        for proj in projects:
            grand = {"K": 0, "A": 0, "B": 0, "C": 0, "PD": 0, "Total": 0, "L": 0}
            summary_rows = []
            for group_name in groups:
                g = proj["hwy_kabco"].get(group_name)
                if g:
                    summary_rows.append(g)
                    for k in ("K", "A", "B", "C", "PD", "Total", "L"):
                        grand[k] += g.get(k, 0)

            int_sum = summarize_int_rows_kabco(proj["int_rows"]) if proj["int_rows"] else None
            ramp_sum = summarize_int_rows_kabco(proj["ramp_rows"]) if proj["ramp_rows"] else None

            if int_sum:
                summary_rows.append({"name": "Intersections", "L": 0, **int_sum})
                for k in ("K", "A", "B", "C", "PD", "Total"):
                    grand[k] += int_sum.get(k, 0)
            if ramp_sum:
                summary_rows.append({"name": "Ramp Terminals", "L": 0, **ramp_sum})
                for k in ("K", "A", "B", "C", "PD", "Total"):
                    grand[k] += ramp_sum.get(k, 0)

            grand["name"] = "TOTAL"
            all_tables.append((proj["label"], summary_rows, grand))

            bars.append({
                "label": proj["label"],
                "value": grand["Total"],
                "segments": [
                    (grand["K"], "k"), (grand["A"], "a"), (grand["B"], "b"),
                    (grand["C"], "c"), (grand["PD"], "pd"),
                ],
            })

        report.add_bar_chart(bars)
        report.add_note(_KABCO_LEGEND)

        for label, rows, grand in all_tables:
            report.add_table(
                label, _KABCO_COLS, rows, total_row=grand,
                distribution_fn=severity_bar,
            )

    def _add_multi_fi(self, report, projects, groups):
        bars = []
        all_tables = []
        for proj in projects:
            grand = {"FI": 0, "PDO": 0, "Total": 0, "L": 0}
            summary_rows = []
            for group_name in groups:
                g = proj["hwy_fi"].get(group_name)
                if g:
                    summary_rows.append(g)
                    for k in ("FI", "PDO", "Total", "L"):
                        grand[k] += g.get(k, 0)

            int_sum = summarize_int_rows_fi(proj["int_rows"]) if proj["int_rows"] else None
            ramp_sum = summarize_int_rows_fi(proj["ramp_rows"]) if proj["ramp_rows"] else None

            if int_sum:
                summary_rows.append({"name": "Intersections", "L": 0, **int_sum})
                for k in ("FI", "PDO", "Total"):
                    grand[k] += int_sum.get(k, 0)
            if ramp_sum:
                summary_rows.append({"name": "Ramp Terminals", "L": 0, **ramp_sum})
                for k in ("FI", "PDO", "Total"):
                    grand[k] += ramp_sum.get(k, 0)

            grand["name"] = "TOTAL"
            all_tables.append((proj["label"], summary_rows, grand))

            bars.append({
                "label": proj["label"],
                "value": grand["Total"],
                "segments": [
                    (grand["FI"], "c4"), (grand["PDO"], "c2"),
                ],
            })

        report.add_bar_chart(bars)
        report.add_note(_FI_LEGEND)

        def fi_bar(row):
            return generic_bar(row, ["FI", "PDO"], ["c4", "c2"])

        for label, rows, grand in all_tables:
            report.add_table(
                label, _FI_COLS, rows, total_row=grand,
                distribution_fn=fi_bar,
            )

    # ── Intersection detail helpers (aggregated by title) ────────────────

    @staticmethod
    def _aggregate_by_type(int_rows):
        """Aggregate intersection/ramp rows by type, summing across years."""
        agg = {}
        for r in int_rows:
            type_key = str(r.get("type") or "Unknown").strip() or "Unknown"
            if type_key not in agg:
                agg[type_key] = {"K": 0, "A": 0, "B": 0, "C": 0, "O": 0, "FI": 0}
            agg[type_key]["K"] += r["K"]
            agg[type_key]["A"] += r["A"]
            agg[type_key]["B"] += r["B"]
            agg[type_key]["C"] += r["C"]
            agg[type_key]["O"] += r["O"]
            agg[type_key]["FI"] += r["FI"]
        return agg

    def _add_int_detail_kabco(self, report, int_rows, title):
        agg = self._aggregate_by_type(int_rows)
        detail = []
        for name, vals in agg.items():
            total = vals["K"] + vals["A"] + vals["B"] + vals["C"] + vals["O"]
            detail.append({
                "name": name,
                "K": vals["K"], "A": vals["A"], "B": vals["B"], "C": vals["C"],
                "PD": vals["O"], "Total": total,
            })
        if not detail:
            return
        total_row = {
            "name": "TOTAL",
            "K": sum(d["K"] for d in detail),
            "A": sum(d["A"] for d in detail),
            "B": sum(d["B"] for d in detail),
            "C": sum(d["C"] for d in detail),
            "PD": sum(d["PD"] for d in detail),
            "Total": sum(d["Total"] for d in detail),
        }
        report.add_table(
            title, _KABCO_COLS_NO_LEN, detail, total_row=total_row,
            distribution_fn=severity_bar,
        )

    def _add_int_detail_fi(self, report, int_rows, title):
        agg = self._aggregate_by_type(int_rows)
        detail = []
        for name, vals in agg.items():
            detail.append({
                "name": name,
                "FI": vals["FI"], "PDO": vals["O"],
                "Total": vals["FI"] + vals["O"],
            })
        if not detail:
            return
        total_row = {
            "name": "TOTAL",
            "FI": sum(d["FI"] for d in detail),
            "PDO": sum(d["PDO"] for d in detail),
            "Total": sum(d["Total"] for d in detail),
        }

        def fi_bar(row):
            return generic_bar(row, ["FI", "PDO"], ["c4", "c2"])

        report.add_table(
            title, _FI_COLS_NO_LEN, detail, total_row=total_row,
            distribution_fn=fi_bar,
        )
