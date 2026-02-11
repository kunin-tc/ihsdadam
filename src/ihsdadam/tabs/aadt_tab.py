"""AADT Input tab -- 5-step wizard for managing highway AADT values."""

import csv
import os
import re
import xml.etree.ElementTree as ET
from dataclasses import asdict
from pathlib import Path
from typing import Dict, List, Optional, Set

from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QFont
from PySide6.QtWidgets import (
    QApplication,
    QComboBox,
    QDialog,
    QFileDialog,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QInputDialog,
    QLabel,
    QLineEdit,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QTextEdit,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
)

from .. import theme
from ..models import AADTSection
from ..workers import AADTScanWorker, _format_station


# ── Row-colour constants ────────────────────────────────────────────────────
_REVIEWED_BG = QColor("#d4edda")
_PENDING_BG = QColor("#fff3cd")
_NEW_YEAR_BG = QColor("#cce5ff")

_TREE_COLUMNS = [
    "Alignment",
    "Section",
    "Start Sta",
    "End Sta",
    "Year",
    "Current",
    "Forecast IDs",
    "Calculated",
]

_COL_WIDTHS = [180, 55, 85, 85, 50, 60, 200, 75]


class AADTTab(QWidget):
    """AADT Input tab -- five-step wizard for mapping forecast IDs to AADT."""

    status_message = Signal(str)
    progress_update = Signal(int, str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._project_path = ""
        self._sections: list = []  # list of dicts (section data)
        self._forecast_ids: dict = {}  # id_str -> float value
        self._reviewed_alignments: set = set()
        self._alignment_nodes: dict = {}  # title -> QTreeWidgetItem
        self._years_in_project: list = []
        self._sort_reverse: dict = {}
        self._worker: Optional[AADTScanWorker] = None
        self._setup_ui()

    # ── public API ───────────────────────────────────────────────────────

    def set_project_path(self, path: str):
        self._project_path = path

    # ── UI construction ──────────────────────────────────────────────────

    def _setup_ui(self):
        outer_layout = QVBoxLayout(self)
        outer_layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QScrollArea.NoFrame)
        outer_layout.addWidget(scroll)

        container = QWidget()
        self._root = QVBoxLayout(container)
        self._root.setContentsMargins(12, 12, 12, 12)
        self._root.setSpacing(10)
        scroll.setWidget(container)

        self._build_notice()
        self._build_step_indicator()
        self._build_step1()
        self._build_step2()
        self._build_step3()
        self._build_step4()
        self._build_step5()

    # -- compatibility notice -------------------------------------------------

    def _build_notice(self):
        notice = QLabel(
            "NOTE: This tool is designed for use with HNTB Wisconsin's "
            "forecasting spreadsheet format. Supports multiple evaluation "
            "years (e.g., 2028 base year + 2050 design year)."
        )
        notice.setWordWrap(True)
        notice.setStyleSheet(
            "background-color: #fff3cd; color: #856404; "
            "padding: 8px 10px; border-radius: 4px; font-style: italic;"
        )
        self._root.addWidget(notice)

    # -- step indicator -------------------------------------------------------

    def _build_step_indicator(self):
        row = QHBoxLayout()
        steps = [
            "1. Setup IHSDM",
            "2. Link Forecast",
            "3. Scan Project",
            "4. Map Sections",
            "5. Apply AADT",
        ]
        self._step_labels: list = []
        for text in steps:
            lbl = QLabel(text)
            lbl.setAlignment(Qt.AlignCenter)
            row.addWidget(lbl)
            self._step_labels.append(lbl)
        self._root.addLayout(row)

    # -- step 1 ---------------------------------------------------------------

    def _build_step1(self):
        grp = QGroupBox("Step 1: Initial IHSDM Setup (Do This First!)")
        lay = QVBoxLayout(grp)

        text = (
            "BEFORE using this tool, you must set up AADT station ranges in IHSDM:\n\n"
            "1. Open your IHSDM project in the IHSDM application\n"
            "2. For each highway alignment, go to Traffic Data -> Annual Average Daily Traffic\n"
            "3. Add AADT ranges for each segment where traffic volumes change:\n"
            "   - Set the Start Station and End Station for each range\n"
            "   - Enter a DEFAULT VALUE of 1 for the AADT (this tool will update the actual values)\n"
            "   - Enter the Year you are evaluating (e.g., 2028)\n\n"
            "Example: If you have a mainline with 3 different volume sections:\n"
            "   Range 1: Sta 1000 to 2500, Year 2028, AADT = 1\n"
            "   Range 2: Sta 2500 to 4000, Year 2028, AADT = 1\n"
            "   Range 3: Sta 4000 to 5500, Year 2028, AADT = 1\n\n"
            "Once you've set up all your AADT ranges in IHSDM, proceed to Step 2."
        )
        lbl = QLabel(text)
        lbl.setWordWrap(True)
        lay.addWidget(lbl)
        self._root.addWidget(grp)

    # -- step 2 ---------------------------------------------------------------

    def _build_step2(self):
        grp = QGroupBox("Step 2: Link Project Forecast Workbook")
        lay = QVBoxLayout(grp)

        # Forecast workbook row
        wb_row = QHBoxLayout()
        wb_row.addWidget(QLabel("Forecast Workbook:"))
        self._forecast_path_edit = QLineEdit()
        self._forecast_path_edit.setPlaceholderText("Select .xlsx or .xlsb file...")
        wb_row.addWidget(self._forecast_path_edit, stretch=1)
        browse_btn = QPushButton("Browse...")
        browse_btn.clicked.connect(self._browse_forecast)
        wb_row.addWidget(browse_btn)
        lay.addLayout(wb_row)

        # Named range row
        nr_row = QHBoxLayout()
        nr_row.addWidget(QLabel("Select Named Range (from column A):"))
        self._named_range_combo = QComboBox()
        self._named_range_combo.setMinimumWidth(250)
        nr_row.addWidget(self._named_range_combo)
        hint = QLabel("(Auto-populated from workbook)")
        hint.setProperty("caption", True)
        nr_row.addWidget(hint)
        nr_row.addStretch()
        lay.addLayout(nr_row)

        # Working year row
        yr_row = QHBoxLayout()
        yr_row.addWidget(QLabel("Working Year:"))
        self._year_combo = QComboBox()
        self._year_combo.setMinimumWidth(80)
        self._year_combo.currentTextChanged.connect(self._filter_by_year)
        yr_row.addWidget(self._year_combo)
        add_year_btn = QPushButton("Add Year...")
        add_year_btn.clicked.connect(self._add_year)
        yr_row.addWidget(add_year_btn)
        self._year_status_label = QLabel("(Scan project first to see available years)")
        self._year_status_label.setProperty("caption", True)
        yr_row.addWidget(self._year_status_label)
        yr_row.addStretch()
        lay.addLayout(yr_row)

        # Button row
        btn_row = QHBoxLayout()
        load_btn = QPushButton("Load Forecast Data")
        load_btn.setProperty("accent", True)
        load_btn.clicked.connect(self._load_forecast)
        btn_row.addWidget(load_btn)
        show_btn = QPushButton("Show Loaded IDs")
        show_btn.clicked.connect(self._show_loaded_ids)
        btn_row.addWidget(show_btn)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        # Status
        self._forecast_status_label = QLabel("No forecast loaded")
        self._forecast_status_label.setProperty("caption", True)
        lay.addWidget(self._forecast_status_label)

        self._root.addWidget(grp)

    # -- step 3 ---------------------------------------------------------------

    def _build_step3(self):
        grp = QGroupBox("Step 3: Scan Project for AADT Sections")
        lay = QVBoxLayout(grp)

        lay.addWidget(QLabel(
            "Scan the project to find all highway alignments and their AADT station ranges."
        ))

        row = QHBoxLayout()
        scan_btn = QPushButton("Scan for AADT Sections")
        scan_btn.setProperty("accent", True)
        scan_btn.clicked.connect(self._scan_sections)
        row.addWidget(scan_btn)
        self._scan_status_label = QLabel("No scan performed")
        self._scan_status_label.setProperty("caption", True)
        row.addWidget(self._scan_status_label)
        row.addStretch()
        lay.addLayout(row)

        self._root.addWidget(grp)

    # -- step 4 ---------------------------------------------------------------

    def _build_step4(self):
        grp = QGroupBox("Step 4: Map AADT Sections to Forecast IDs")
        lay = QVBoxLayout(grp)

        desc = QLabel(
            "For each AADT section, enter the Forecast IDs that contribute to that "
            "section's volume.\nClick on an alignment to expand and see its sections. "
            "Enter IDs, then click \"Mark Reviewed\" when done with each alignment."
        )
        desc.setWordWrap(True)
        lay.addWidget(desc)

        # Body: tree (left) + checklist (right)
        body = QHBoxLayout()

        # Left: tree
        self._tree = QTreeWidget()
        self._tree.setHeaderLabels(_TREE_COLUMNS)
        self._tree.setAlternatingRowColors(False)
        self._tree.setRootIsDecorated(True)
        self._tree.setSelectionMode(QTreeWidget.ExtendedSelection)
        self._tree.header().setStretchLastSection(True)
        self._tree.header().setSectionsClickable(True)
        self._tree.header().sectionClicked.connect(self._sort_tree)
        for i, w in enumerate(_COL_WIDTHS):
            if i < self._tree.columnCount():
                self._tree.setColumnWidth(i, w)
        self._tree.itemSelectionChanged.connect(self._on_tree_select)
        body.addWidget(self._tree, stretch=1)

        # Right: review checklist
        checklist_grp = QGroupBox("Review Checklist")
        checklist_grp.setFixedWidth(220)
        checklist_lay = QVBoxLayout(checklist_grp)

        checklist_scroll = QScrollArea()
        checklist_scroll.setWidgetResizable(True)
        checklist_scroll.setFrameShape(QScrollArea.NoFrame)
        self._checklist_widget = QWidget()
        self._checklist_layout = QVBoxLayout(self._checklist_widget)
        self._checklist_layout.setContentsMargins(4, 4, 4, 4)
        self._checklist_layout.setSpacing(2)
        self._checklist_layout.addStretch()
        checklist_scroll.setWidget(self._checklist_widget)
        checklist_lay.addWidget(checklist_scroll)

        body.addWidget(checklist_grp)
        lay.addLayout(body)

        # ID entry row 1
        id_row1 = QHBoxLayout()
        id_row1.addWidget(QLabel("Forecast IDs (+/-):"))

        self._sign_combos: list = []
        self._id_edits: list = []
        for _ in range(3):
            sign = QComboBox()
            sign.addItems(["+", "-"])
            sign.setFixedWidth(45)
            id_row1.addWidget(sign)
            self._sign_combos.append(sign)

            edit = QLineEdit()
            edit.setFixedWidth(70)
            id_row1.addWidget(edit)
            self._id_edits.append(edit)

        id_row1.addStretch()
        lay.addLayout(id_row1)

        # ID entry row 2
        id_row2 = QHBoxLayout()
        id_row2.addSpacing(100)  # spacer to align under row 1

        for _ in range(3):
            sign = QComboBox()
            sign.addItems(["+", "-"])
            sign.setFixedWidth(45)
            id_row2.addWidget(sign)
            self._sign_combos.append(sign)

            edit = QLineEdit()
            edit.setFixedWidth(70)
            id_row2.addWidget(edit)
            self._id_edits.append(edit)

        apply_sel_btn = QPushButton("Apply to Selected")
        apply_sel_btn.clicked.connect(self._apply_ids_to_selected)
        id_row2.addWidget(apply_sel_btn)
        id_row2.addStretch()
        lay.addLayout(id_row2)

        # Action buttons
        action_row = QHBoxLayout()
        mark_btn = QPushButton("Mark Alignment Reviewed")
        mark_btn.setProperty("accent", True)
        mark_btn.clicked.connect(self._mark_reviewed)
        action_row.addWidget(mark_btn)

        calc_btn = QPushButton("Calculate All")
        calc_btn.clicked.connect(self._calculate_all)
        action_row.addWidget(calc_btn)

        expand_btn = QPushButton("Expand All")
        expand_btn.clicked.connect(self._expand_all)
        action_row.addWidget(expand_btn)

        collapse_btn = QPushButton("Collapse All")
        collapse_btn.clicked.connect(self._collapse_all)
        action_row.addWidget(collapse_btn)

        action_row.addStretch()
        lay.addLayout(action_row)

        # Session buttons
        session_row = QHBoxLayout()
        session_lbl = QLabel("Session:")
        session_lbl.setStyleSheet("font-weight: 600;")
        session_row.addWidget(session_lbl)

        save_btn = QPushButton("Save Mapping to CSV")
        save_btn.clicked.connect(self._save_mapping_csv)
        session_row.addWidget(save_btn)

        load_btn = QPushButton("Load Mapping from CSV")
        load_btn.clicked.connect(self._load_mapping_csv)
        session_row.addWidget(load_btn)

        session_row.addStretch()
        lay.addLayout(session_row)

        self._root.addWidget(grp)

    # -- step 5 ---------------------------------------------------------------

    def _build_step5(self):
        grp = QGroupBox("Step 5: Apply AADT Values to IHSDM Project")
        lay = QVBoxLayout(grp)

        desc = QLabel(
            "Once all sections have calculated AADT values, click below to update "
            "the highway XML files.\nThis will modify the adtRate attribute in each "
            "AnnualAveDailyTraffic element."
        )
        desc.setWordWrap(True)
        lay.addWidget(desc)

        btn_row = QHBoxLayout()
        preview_btn = QPushButton("Preview Changes")
        preview_btn.clicked.connect(self._preview_changes)
        btn_row.addWidget(preview_btn)

        apply_btn = QPushButton("Apply AADT to XML Files")
        apply_btn.setProperty("accent", True)
        apply_btn.clicked.connect(self._apply_to_xml)
        btn_row.addWidget(apply_btn)

        export_btn = QPushButton("Export Mapping to Excel")
        export_btn.clicked.connect(self._export_mapping)
        btn_row.addWidget(export_btn)

        btn_row.addStretch()
        lay.addLayout(btn_row)

        self._apply_status_label = QLabel("")
        self._apply_status_label.setStyleSheet(f"color: {theme.ACCENT_GREEN};")
        lay.addWidget(self._apply_status_label)

        self._root.addWidget(grp)

    # ====================================================================
    # Step 2 -- Forecast workbook
    # ====================================================================

    def _browse_forecast(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select Forecast Workbook",
            "",
            "Excel files (*.xlsx *.xlsb);;Excel Workbook (*.xlsx);;Excel Binary (*.xlsb);;All files (*)",
        )
        if path:
            self._forecast_path_edit.setText(path)
            self._load_named_ranges(path)

    def _load_named_ranges(self, file_path: str):
        """Read row 1 headers of BalancedOutput sheet and populate combo."""
        try:
            is_xlsb = file_path.lower().endswith(".xlsb")
            named_ranges: list = []

            if is_xlsb:
                try:
                    from pyxlsb import open_workbook as open_xlsb
                except ImportError:
                    QMessageBox.critical(
                        self,
                        "Error",
                        "The pyxlsb library is required to read .xlsb files.\n\n"
                        "Install with: pip install pyxlsb",
                    )
                    return

                with open_xlsb(file_path) as wb:
                    sheet_name = None
                    for name in wb.sheets:
                        if "balancedoutput" in name.lower():
                            sheet_name = name
                            break
                    if not sheet_name:
                        QMessageBox.warning(
                            self, "Warning",
                            "Could not find 'BalancedOutput' sheet in workbook",
                        )
                        return
                    with wb.get_sheet(sheet_name) as sheet:
                        for row in sheet.rows():
                            for cell in row:
                                if cell.v is not None:
                                    val = cell.v
                                    if isinstance(val, str) and val.strip():
                                        val_str = val.strip()
                                        if val_str not in named_ranges:
                                            named_ranges.append(val_str)
                            break
            else:
                from openpyxl import load_workbook

                wb = load_workbook(file_path, data_only=True, read_only=True)
                sheet_name = None
                for name in wb.sheetnames:
                    if "balancedoutput" in name.lower():
                        sheet_name = name
                        break
                if not sheet_name:
                    QMessageBox.warning(
                        self, "Warning",
                        "Could not find 'BalancedOutput' sheet in workbook",
                    )
                    wb.close()
                    return
                ws = wb[sheet_name]
                for row in ws.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        if cell.value is not None:
                            val = cell.value
                            if isinstance(val, str) and val.strip():
                                val_str = val.strip()
                                if val_str not in named_ranges:
                                    named_ranges.append(val_str)
                wb.close()

            if named_ranges:
                self._named_range_combo.clear()
                self._named_range_combo.addItems(named_ranges)
                self._named_range_combo.setCurrentIndex(-1)
                self.status_message.emit(
                    f"Found {len(named_ranges)} named ranges in row 1"
                )
            else:
                self._named_range_combo.clear()
                QMessageBox.information(
                    self, "Info",
                    "No named ranges found in row 1 of BalancedOutput sheet",
                )
        except Exception as exc:
            QMessageBox.critical(
                self, "Error", f"Error reading workbook: {exc}"
            )

    def _load_forecast(self):
        """Load forecast data from workbook under selected header column range."""
        forecast_path = self._forecast_path_edit.text().strip()
        selected_header = self._named_range_combo.currentText()

        if not forecast_path or not os.path.isfile(forecast_path):
            QMessageBox.critical(
                self, "Error", "Please select a valid forecast workbook"
            )
            return
        if not selected_header:
            QMessageBox.critical(
                self, "Error", "Please select a named range from the dropdown"
            )
            return

        try:
            is_xlsb = forecast_path.lower().endswith(".xlsb")
            self._forecast_ids = {}
            sheet_name = None

            if is_xlsb:
                try:
                    from pyxlsb import open_workbook as open_xlsb
                except ImportError:
                    QMessageBox.critical(
                        self, "Error",
                        "The pyxlsb library is required to read .xlsb files.\n\n"
                        "Install with: pip install pyxlsb",
                    )
                    return

                with open_xlsb(forecast_path) as wb:
                    for name in wb.sheets:
                        if "balancedoutput" in name.lower():
                            sheet_name = name
                            break
                    if not sheet_name:
                        QMessageBox.critical(
                            self, "Error",
                            "Could not find 'BalancedOutput' sheet in workbook",
                        )
                        return
                    with wb.get_sheet(sheet_name) as sheet:
                        rows = list(sheet.rows())
                        if not rows:
                            QMessageBox.critical(
                                self, "Error",
                                "BalancedOutput sheet is empty",
                            )
                            return
                        header_row = rows[0]
                        start_col = None
                        for col_idx, cell in enumerate(header_row):
                            if cell.v is not None and str(cell.v).strip() == selected_header:
                                start_col = col_idx
                                break
                        if start_col is None:
                            QMessageBox.critical(
                                self, "Error",
                                f"Could not find header '{selected_header}' in row 1",
                            )
                            return
                        end_col = len(header_row)
                        for col_idx in range(start_col + 1, len(header_row)):
                            cell = header_row[col_idx]
                            if cell.v is not None and isinstance(cell.v, str) and cell.v.strip():
                                end_col = col_idx
                                break
                        for row in rows[1:]:
                            col = start_col
                            while col + 1 < end_col and col + 1 < len(row):
                                id_cell = row[col].v if col < len(row) and row[col] else None
                                val_cell = row[col + 1].v if col + 1 < len(row) and row[col + 1] else None
                                if id_cell is not None:
                                    if isinstance(id_cell, float) and id_cell == int(id_cell):
                                        id_str = str(int(id_cell))
                                    else:
                                        id_str = str(id_cell).strip()
                                    if id_str:
                                        try:
                                            val = float(val_cell) if val_cell is not None else 0
                                        except (ValueError, TypeError):
                                            val = 0
                                        self._forecast_ids[id_str] = val
                                col += 2
            else:
                from openpyxl import load_workbook

                wb = load_workbook(forecast_path, data_only=True)
                for name in wb.sheetnames:
                    if "balancedoutput" in name.lower():
                        sheet_name = name
                        break
                if not sheet_name:
                    QMessageBox.critical(
                        self, "Error",
                        "Could not find 'BalancedOutput' sheet in workbook",
                    )
                    return
                ws = wb[sheet_name]

                start_col = None
                for col in range(1, ws.max_column + 1):
                    cell_val = ws.cell(row=1, column=col).value
                    if cell_val is not None and str(cell_val).strip() == selected_header:
                        start_col = col
                        break
                if start_col is None:
                    QMessageBox.critical(
                        self, "Error",
                        f"Could not find header '{selected_header}' in row 1",
                    )
                    wb.close()
                    return

                end_col = ws.max_column + 1
                for col in range(start_col + 1, ws.max_column + 1):
                    cell_val = ws.cell(row=1, column=col).value
                    if cell_val is not None and isinstance(cell_val, str) and cell_val.strip():
                        end_col = col
                        break

                for row in range(2, ws.max_row + 1):
                    col = start_col
                    while col + 1 < end_col:
                        id_cell = ws.cell(row=row, column=col).value
                        val_cell = ws.cell(row=row, column=col + 1).value
                        if id_cell is not None:
                            if isinstance(id_cell, float) and id_cell == int(id_cell):
                                id_str = str(int(id_cell))
                            else:
                                id_str = str(id_cell).strip()
                            if id_str:
                                try:
                                    val = float(val_cell) if val_cell is not None else 0
                                except (ValueError, TypeError):
                                    val = 0
                                self._forecast_ids[id_str] = val
                        col += 2
                wb.close()

            self._forecast_status_label.setText(
                f"Loaded {len(self._forecast_ids)} forecast IDs from '{selected_header}'"
            )
            self.status_message.emit(
                f"Forecast data loaded: {len(self._forecast_ids)} IDs"
            )
        except Exception as exc:
            QMessageBox.critical(
                self, "Error", f"Error loading forecast: {exc}"
            )
            self._forecast_status_label.setText(f"Error: {exc}")

    def _show_loaded_ids(self):
        """Open dialog with searchable list of all loaded forecast IDs."""
        if not self._forecast_ids:
            QMessageBox.information(
                self, "No Forecast Data",
                "No forecast data has been loaded yet.\n\nClick 'Load Forecast Data' first.",
            )
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Loaded Forecast IDs")
        dlg.resize(500, 600)
        lay = QVBoxLayout(dlg)

        lay.addWidget(QLabel(f"Loaded {len(self._forecast_ids)} Forecast IDs:"))

        search_row = QHBoxLayout()
        search_row.addWidget(QLabel("Search:"))
        search_edit = QLineEdit()
        search_edit.setPlaceholderText("Filter IDs...")
        search_edit.setClearButtonEnabled(True)
        search_row.addWidget(search_edit)
        lay.addLayout(search_row)

        list_tree = QTreeWidget()
        list_tree.setHeaderLabels(["ID", "Value"])
        list_tree.setColumnWidth(0, 200)
        list_tree.setRootIsDecorated(False)
        lay.addWidget(list_tree, stretch=1)

        def _populate(filter_text=""):
            list_tree.clear()
            ft = filter_text.lower()
            sorted_ids = sorted(
                self._forecast_ids.keys(),
                key=lambda x: (not x.isdigit(), x),
            )
            for id_str in sorted_ids:
                if ft and ft not in id_str.lower():
                    continue
                value = self._forecast_ids[id_str]
                disp = str(int(value)) if value == int(value) else str(value)
                item = QTreeWidgetItem([id_str, disp])
                list_tree.addTopLevelItem(item)

        search_edit.textChanged.connect(_populate)
        _populate()

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dlg.accept)
        lay.addWidget(close_btn)
        dlg.exec()

    # ====================================================================
    # Step 3 -- Scan project
    # ====================================================================

    def _scan_sections(self):
        if not self._project_path or not os.path.isdir(self._project_path):
            QMessageBox.critical(
                self, "Error", "Please select a valid project directory"
            )
            return

        self.status_message.emit("Scanning for AADT sections...")
        self._sections = []

        self._worker = AADTScanWorker(self._project_path, parent=self)
        self._worker.progress.connect(self._on_scan_progress)
        self._worker.finished.connect(self._on_scan_finished)
        self._worker.error.connect(self._on_scan_error)
        self._worker.start()

    def _on_scan_progress(self, pct: int, msg: str):
        self.progress_update.emit(pct, msg)
        self.status_message.emit(msg)

    def _on_scan_finished(self, aadt_sections: list):
        """Receive AADTSection list from worker, convert to dicts, populate tree."""
        self._sections = []
        for sec in aadt_sections:
            d = asdict(sec)
            self._sections.append(d)

        # Collect unique years
        self._years_in_project = sorted(
            {s["year"] for s in self._sections if s["year"]}
        )
        self._year_combo.blockSignals(True)
        self._year_combo.clear()
        self._year_combo.addItems(self._years_in_project)
        if self._years_in_project:
            self._year_combo.setCurrentIndex(0)
            self._year_status_label.setText(
                f"Years found: {', '.join(self._years_in_project)}"
            )
            self._year_status_label.setStyleSheet(f"color: {theme.ACCENT_GREEN};")
        else:
            self._year_status_label.setText("No years found in AADT data")
            self._year_status_label.setStyleSheet(f"color: {theme.WARNING};")
        self._year_combo.blockSignals(False)

        self._reviewed_alignments = set()
        self._populate_tree()
        self._update_checklist()

        working_year = self._year_combo.currentText()
        alignments = {s["roadway_title"] for s in self._sections
                      if not working_year or s["year"] == working_year}
        year_sections = [s for s in self._sections
                         if not working_year or s["year"] == working_year]

        self._scan_status_label.setText(
            f"Found {len(self._sections)} AADT sections across {len(alignments)} alignments"
        )
        self.status_message.emit(
            f"Scan complete: {len(self._sections)} AADT sections in {len(alignments)} alignments"
        )

    def _on_scan_error(self, msg: str):
        QMessageBox.critical(self, "Scan Error", msg)
        self.status_message.emit("Error during AADT scan")

    # ====================================================================
    # Tree population / filtering
    # ====================================================================

    def _populate_tree(self):
        """Build the tree grouped by alignment, filtered by working year."""
        self._tree.clear()
        self._alignment_nodes = {}

        working_year = self._year_combo.currentText()

        # Group sections by roadway title
        alignments: Dict[str, List[tuple]] = {}
        for i, section in enumerate(self._sections):
            if working_year and section["year"] != working_year:
                continue
            title = section["roadway_title"]
            alignments.setdefault(title, []).append((i, section))

        col_count = self._tree.columnCount()

        for title in sorted(alignments.keys()):
            sections_list = alignments[title]
            # Parent node
            parent = QTreeWidgetItem()
            parent.setText(0, f"{title} ({len(sections_list)} sections)")
            parent.setData(0, Qt.UserRole, f"align_{title}")
            # Bold font for alignment rows
            font = parent.font(0)
            font.setBold(True)
            parent.setFont(0, font)
            # Pending background
            is_reviewed = title in self._reviewed_alignments
            bg = _REVIEWED_BG if is_reviewed else _PENDING_BG
            for c in range(col_count):
                parent.setBackground(c, bg)
            self._tree.addTopLevelItem(parent)
            self._alignment_nodes[title] = parent

            for idx, section in sections_list:
                # Ensure all ID/sign slots exist
                for slot in range(1, 7):
                    section.setdefault(f"id{slot}", "")
                    section.setdefault(f"sign{slot}", "+")

                id_display = self._get_ids_display(section)
                child = QTreeWidgetItem()
                child.setText(1, str(section["section_num"]))
                child.setText(2, _format_station(section["start_station"]))
                child.setText(3, _format_station(section["end_station"]))
                child.setText(4, section["year"])
                child.setText(5, section["current_aadt"])
                child.setText(6, id_display)
                child.setText(7, section["calculated_aadt"])
                child.setData(0, Qt.UserRole, idx)

                # Alignment for center columns
                for c in range(1, col_count):
                    child.setTextAlignment(c, Qt.AlignCenter)

                # Row colouring
                if section.get("is_new", False):
                    row_bg = _NEW_YEAR_BG
                    italic_font = child.font(0)
                    italic_font.setItalic(True)
                    for c in range(col_count):
                        child.setFont(c, italic_font)
                elif is_reviewed:
                    row_bg = _REVIEWED_BG
                else:
                    row_bg = _PENDING_BG
                for c in range(col_count):
                    child.setBackground(c, row_bg)

                parent.addChild(child)

    def _filter_by_year(self):
        """Re-populate tree filtered by the selected working year."""
        if not self._sections:
            return
        self._populate_tree()
        self._update_checklist()

        working_year = self._year_combo.currentText()
        count = sum(
            1 for s in self._sections
            if not working_year or s["year"] == working_year
        )
        align_count = len(self._alignment_nodes)
        year_str = f" for year {working_year}" if working_year else ""
        self._scan_status_label.setText(
            f"Showing {count} sections across {align_count} alignments{year_str}"
        )

    def _add_year(self):
        """Add a new year by duplicating sections from the base year."""
        if not self._sections:
            QMessageBox.warning(
                self, "No Data",
                "Please scan the project first to find existing AADT sections.",
            )
            return
        if not self._years_in_project:
            QMessageBox.warning(
                self, "No Years",
                "No years found in the project. Set up AADT in IHSDM first.",
            )
            return

        new_year, ok = QInputDialog.getText(
            self, "Add Year", "Enter the new evaluation year (e.g., 2050):"
        )
        if not ok or not new_year:
            return
        if not new_year.isdigit() or len(new_year) != 4:
            QMessageBox.critical(
                self, "Invalid Year", "Please enter a valid 4-digit year."
            )
            return
        if new_year in self._years_in_project:
            QMessageBox.warning(
                self, "Year Exists",
                f"Year {new_year} already exists in the project.",
            )
            return

        base_year = self._year_combo.currentText() or self._years_in_project[0]
        base_sections = [s for s in self._sections if s["year"] == base_year]
        if not base_sections:
            QMessageBox.critical(
                self, "Error",
                f"No sections found for base year {base_year}.",
            )
            return

        base_with_ids = sum(1 for s in base_sections if s.get("id1", "").strip())
        if base_with_ids == 0:
            QMessageBox.warning(
                self, "No IDs Mapped",
                f"Base year {base_year} has no forecast IDs mapped.\n\n"
                "The new year sections will have empty IDs.\n"
                "You can map IDs after adding the year.",
            )

        new_count = 0
        ids_copied = 0
        for base in base_sections:
            new_section = {
                "roadway_title": base["roadway_title"],
                "highway_dir": base["highway_dir"],
                "xml_file": base["xml_file"],
                "section_num": base["section_num"],
                "start_station": base["start_station"],
                "end_station": base["end_station"],
                "year": new_year,
                "current_aadt": "1",
                "id1": base.get("id1", ""),
                "id2": base.get("id2", ""),
                "id3": base.get("id3", ""),
                "id4": base.get("id4", ""),
                "id5": base.get("id5", ""),
                "id6": base.get("id6", ""),
                "sign1": base.get("sign1", "+"),
                "sign2": base.get("sign2", "+"),
                "sign3": base.get("sign3", "+"),
                "sign4": base.get("sign4", "+"),
                "sign5": base.get("sign5", "+"),
                "sign6": base.get("sign6", "+"),
                "calculated_aadt": "",
                "is_new": True,
            }
            if base.get("id1", "").strip():
                ids_copied += 1
            self._sections.append(new_section)
            new_count += 1

        # Update year combo
        self._years_in_project = sorted(
            {s["year"] for s in self._sections if s["year"]}
        )
        self._year_combo.blockSignals(True)
        self._year_combo.clear()
        self._year_combo.addItems(self._years_in_project)
        idx = self._years_in_project.index(new_year) if new_year in self._years_in_project else 0
        self._year_combo.setCurrentIndex(idx)
        self._year_combo.blockSignals(False)
        self._year_status_label.setText(
            f"Years: {', '.join(self._years_in_project)}"
        )
        self._year_status_label.setStyleSheet(f"color: {theme.ACCENT_GREEN};")

        self._populate_tree()
        self._update_checklist()

        ids_msg = (
            f"Forecast IDs copied from {base_year} for {ids_copied} sections."
            if ids_copied > 0
            else f"No forecast IDs to copy from {base_year}."
        )
        QMessageBox.information(
            self, "Year Added",
            f"Added {new_count} sections for year {new_year}.\n\n"
            f"{ids_msg}\n\n"
            f"NEXT STEPS:\n"
            f"1. Load the forecast workbook data for {new_year}\n"
            f"2. Click 'Calculate All' to compute AADT values\n"
            f"3. Click 'Apply AADT' to write to XML\n\n"
            f"(New sections shown in blue/italic)",
        )

    # ====================================================================
    # Tree selection / ID entry
    # ====================================================================

    def _on_tree_select(self):
        """Populate ID entry fields from selected section."""
        items = self._tree.selectedItems()
        if not items:
            return
        item = items[0]
        raw = item.data(0, Qt.UserRole)
        if raw is None:
            return
        # Skip alignment parent nodes
        if isinstance(raw, str) and raw.startswith("align_"):
            return
        try:
            idx = int(raw)
        except (ValueError, TypeError):
            return
        if idx < 0 or idx >= len(self._sections):
            return
        section = self._sections[idx]
        for slot in range(6):
            self._id_edits[slot].setText(section.get(f"id{slot + 1}", ""))
            sign_val = section.get(f"sign{slot + 1}", "+")
            combo = self._sign_combos[slot]
            combo.setCurrentIndex(0 if sign_val == "+" else 1)

    def _apply_ids_to_selected(self):
        """Read ID/sign entries and apply to all selected tree items."""
        items = self._tree.selectedItems()
        if not items:
            QMessageBox.warning(
                self, "No Selection", "Please select a section row first"
            )
            return

        for item in items:
            raw = item.data(0, Qt.UserRole)
            if raw is None:
                continue
            if isinstance(raw, str) and raw.startswith("align_"):
                continue
            try:
                idx = int(raw)
            except (ValueError, TypeError):
                continue
            if idx < 0 or idx >= len(self._sections):
                continue

            section = self._sections[idx]
            for slot in range(6):
                section[f"id{slot + 1}"] = self._id_edits[slot].text().strip()
                section[f"sign{slot + 1}"] = self._sign_combos[slot].currentText()

            calculated = self._calculate_section_aadt(section)
            section["calculated_aadt"] = str(int(calculated)) if calculated > 0 else ""

            # Update tree item display
            id_display = self._get_ids_display(section)
            item.setText(6, id_display)
            item.setText(7, section["calculated_aadt"])

        self.status_message.emit("Forecast IDs applied and AADT calculated")

    # ====================================================================
    # AADT calculation
    # ====================================================================

    def _get_ids_display(self, section: dict) -> str:
        """Return compact string like '101, +202, -303'."""
        parts: list = []
        for i in range(1, 7):
            id_val = section.get(f"id{i}", "").strip()
            if id_val:
                sign = section.get(f"sign{i}", "+")
                if sign == "-":
                    parts.append(f"-{id_val}")
                else:
                    parts.append(f"+{id_val}" if parts else id_val)
        return ", ".join(parts) if parts else ""

    def _calculate_section_aadt(self, section: dict) -> float:
        """Sum/subtract forecast values based on IDs and signs."""
        total = 0.0
        for i in range(1, 7):
            forecast_id = section.get(f"id{i}", "").strip()
            sign = section.get(f"sign{i}", "+")
            if forecast_id and forecast_id in self._forecast_ids:
                value = self._forecast_ids[forecast_id]
                if sign == "-":
                    total -= value
                else:
                    total += value
        return max(0.0, total)

    def _calculate_all(self):
        """Calculate AADT for all sections in working year that have IDs."""
        if not self._forecast_ids:
            QMessageBox.warning(
                self, "No Forecast",
                "Please load forecast data first (Step 2)",
            )
            return

        working_year = self._year_combo.currentText()
        if not working_year:
            QMessageBox.warning(
                self, "No Year Selected",
                "Please select a working year first",
            )
            return

        calculated_count = 0
        missing_ids: set = set()
        total_for_year = 0

        for section in self._sections:
            if section["year"] != working_year:
                continue
            total_for_year += 1

            has_ids = False
            for j in range(1, 7):
                id_val = section.get(f"id{j}", "").strip()
                if id_val:
                    has_ids = True
                    if id_val not in self._forecast_ids:
                        missing_ids.add(id_val)

            if has_ids:
                calculated = self._calculate_section_aadt(section)
                section["calculated_aadt"] = str(int(calculated))
                calculated_count += 1
            else:
                section["calculated_aadt"] = ""

        # Refresh tree display
        self._populate_tree()
        self._update_checklist()

        status_msg = (
            f"Year {working_year}: Calculated AADT for "
            f"{calculated_count} of {total_for_year} sections"
        )
        if missing_ids:
            status_msg += f" (Warning: {len(missing_ids)} IDs not found in forecast)"
            if len(missing_ids) <= 10:
                missing_list = ", ".join(sorted(missing_ids))
            else:
                missing_list = (
                    ", ".join(sorted(list(missing_ids)[:10]))
                    + f"... and {len(missing_ids) - 10} more"
                )
            QMessageBox.warning(
                self, "Missing Forecast IDs",
                f"The following IDs were not found in the forecast workbook:\n\n"
                f"{missing_list}\n\n"
                f"Make sure the ID values match exactly (case-sensitive).\n"
                f"Loaded forecast contains {len(self._forecast_ids)} IDs.",
            )
        elif calculated_count == 0:
            QMessageBox.information(
                self, "No IDs Assigned",
                f"No forecast IDs have been assigned to any sections for year {working_year}.\n\n"
                "Select a section, enter Forecast IDs, then click 'Apply to Selected' "
                "before using Calculate All.",
            )
        else:
            QMessageBox.information(
                self, "Calculation Complete",
                f"Calculated AADT for {calculated_count} sections in year {working_year}.",
            )
        self.status_message.emit(status_msg)

    # ====================================================================
    # Review checklist
    # ====================================================================

    def _mark_reviewed(self):
        """Mark the selected alignment as reviewed."""
        items = self._tree.selectedItems()
        if not items:
            QMessageBox.warning(
                self, "No Selection",
                "Please select an alignment or section first",
            )
            return

        item = items[0]
        raw = item.data(0, Qt.UserRole)

        alignment_title = None
        if isinstance(raw, str) and raw.startswith("align_"):
            alignment_title = raw[6:]
        else:
            try:
                idx = int(raw)
                if 0 <= idx < len(self._sections):
                    alignment_title = self._sections[idx]["roadway_title"]
            except (ValueError, TypeError):
                pass

        if not alignment_title:
            return

        self._reviewed_alignments.add(alignment_title)

        # Recolour the alignment node and children
        col_count = self._tree.columnCount()
        if alignment_title in self._alignment_nodes:
            parent = self._alignment_nodes[alignment_title]
            for c in range(col_count):
                parent.setBackground(c, _REVIEWED_BG)
            for ci in range(parent.childCount()):
                child = parent.child(ci)
                for c in range(col_count):
                    child.setBackground(c, _REVIEWED_BG)

        self._update_checklist()
        self.status_message.emit(f"Marked '{alignment_title}' as reviewed")

    def _update_checklist(self):
        """Rebuild the checklist showing check/circle for each alignment."""
        # Clear existing items
        while self._checklist_layout.count():
            child = self._checklist_layout.takeAt(0)
            if child.widget():
                child.widget().deleteLater()

        if not self._alignment_nodes:
            self._checklist_layout.addStretch()
            return

        header = QLabel("Alignments:")
        header.setStyleSheet("font-weight: 600;")
        self._checklist_layout.addWidget(header)

        for title in sorted(self._alignment_nodes.keys()):
            is_reviewed = title in self._reviewed_alignments
            icon = "\u2713" if is_reviewed else "\u25CB"
            colour = theme.ACCENT_GREEN if is_reviewed else theme.TEXT_SECONDARY
            lbl = QLabel(f"{icon} {title}")
            lbl.setStyleSheet(f"color: {colour}; font-size: 9pt;")
            self._checklist_layout.addWidget(lbl)

        # Summary
        total = len(self._alignment_nodes)
        reviewed = len(self._reviewed_alignments)
        sep = QLabel("")
        sep.setFixedHeight(1)
        sep.setStyleSheet(f"background-color: {theme.BORDER};")
        self._checklist_layout.addWidget(sep)

        summary = QLabel(f"Progress: {reviewed}/{total}")
        summary.setStyleSheet("font-weight: 600;")
        self._checklist_layout.addWidget(summary)
        self._checklist_layout.addStretch()

    # ====================================================================
    # Tree helpers
    # ====================================================================

    def _expand_all(self):
        self._tree.expandAll()

    def _collapse_all(self):
        self._tree.collapseAll()

    def _sort_tree(self, col_index: int):
        """Sort tree -- alignments by name or sections within each alignment."""
        reverse = self._sort_reverse.get(col_index, False)

        if col_index == 0:
            # Sort top-level alignment nodes
            items: list = []
            while self._tree.topLevelItemCount():
                items.append(self._tree.takeTopLevelItem(0))
            items.sort(
                key=lambda it: it.text(0).lower(),
                reverse=reverse,
            )
            for it in items:
                self._tree.addTopLevelItem(it)
        else:
            # Sort children within each alignment
            for ai in range(self._tree.topLevelItemCount()):
                parent = self._tree.topLevelItem(ai)
                children: list = []
                while parent.childCount():
                    children.append(parent.takeChild(0))

                def _sort_key(child_item):
                    val = child_item.text(col_index)
                    try:
                        if isinstance(val, str) and "+" in val:
                            return float(val.replace("+", ""))
                        return float(val) if val else 0
                    except (ValueError, TypeError):
                        return val.lower() if val else ""

                children.sort(key=_sort_key, reverse=reverse)
                for child_item in children:
                    parent.addChild(child_item)

        self._sort_reverse[col_index] = not reverse
        direction = "descending" if reverse else "ascending"
        self.status_message.emit(f"Sorted by column {col_index} ({direction})")

    # ====================================================================
    # Step 4 -- Save / Load mapping CSV
    # ====================================================================

    def _save_mapping_csv(self):
        """Save section mappings to CSV."""
        if not self._sections:
            QMessageBox.warning(
                self, "No Data",
                "No AADT sections to save. Please scan the project first.",
            )
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Save AADT Mapping Progress", "",
            "CSV files (*.csv);;All files (*)",
        )
        if not path:
            return

        try:
            with open(path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                headers = [
                    "Roadway_Title", "Section_Num", "Start_Station", "End_Station",
                    "Year", "Current_AADT",
                    "ID1", "Sign1", "ID2", "Sign2", "ID3", "Sign3",
                    "ID4", "Sign4", "ID5", "Sign5", "ID6", "Sign6",
                    "Calculated_AADT", "Reviewed", "Highway_Dir", "XML_File", "Is_New",
                ]
                writer.writerow(headers)

                for section in self._sections:
                    is_reviewed = section["roadway_title"] in self._reviewed_alignments
                    row = [
                        section["roadway_title"],
                        section["section_num"],
                        section["start_station"],
                        section["end_station"],
                        section["year"],
                        section["current_aadt"],
                        section.get("id1", ""), section.get("sign1", "+"),
                        section.get("id2", ""), section.get("sign2", "+"),
                        section.get("id3", ""), section.get("sign3", "+"),
                        section.get("id4", ""), section.get("sign4", "+"),
                        section.get("id5", ""), section.get("sign5", "+"),
                        section.get("id6", ""), section.get("sign6", "+"),
                        section["calculated_aadt"],
                        "Yes" if is_reviewed else "No",
                        section["highway_dir"],
                        section["xml_file"],
                        "Yes" if section.get("is_new", False) else "No",
                    ]
                    writer.writerow(row)

            QMessageBox.information(
                self, "Save Complete",
                f"AADT mapping saved to:\n{path}\n\n"
                "You can load this file later to resume your work.",
            )
            self.status_message.emit(f"Mapping saved to {os.path.basename(path)}")
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Error saving mapping: {exc}")

    def _load_mapping_csv(self):
        """Read CSV and restore forecast IDs/signs to matching sections."""
        if not self._sections:
            QMessageBox.warning(
                self, "Scan Required",
                "Please scan the project first (Step 3) before loading a mapping.\n\n"
                "This ensures AADT values come from the actual XML files.",
            )
            return

        path, _ = QFileDialog.getOpenFileName(
            self, "Load AADT Mapping Progress", "",
            "CSV files (*.csv);;All files (*)",
        )
        if not path:
            return

        try:
            def _normalize_station(sta):
                try:
                    return float(sta)
                except (ValueError, TypeError):
                    return sta

            # Build lookup from scanned sections
            section_lookup: dict = {}
            for i, section in enumerate(self._sections):
                key = (
                    section["roadway_title"],
                    _normalize_station(section["start_station"]),
                    _normalize_station(section["end_station"]),
                    section["year"],
                )
                section_lookup[key] = i

            matched_count = 0
            unmatched_rows: list = []
            reviewed_alignments: set = set()
            new_sections_added = 0

            with open(path, "r", encoding="utf-8") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    key = (
                        row.get("Roadway_Title", ""),
                        _normalize_station(row.get("Start_Station", "")),
                        _normalize_station(row.get("End_Station", "")),
                        row.get("Year", ""),
                    )
                    if key in section_lookup:
                        idx = section_lookup[key]
                        for slot in range(1, 7):
                            self._sections[idx][f"id{slot}"] = row.get(f"ID{slot}", "")
                            self._sections[idx][f"sign{slot}"] = row.get(f"Sign{slot}", "+")
                        matched_count += 1
                        if row.get("Reviewed", "No") == "Yes":
                            reviewed_alignments.add(row.get("Roadway_Title", ""))
                    else:
                        if row.get("Is_New", "No") == "Yes":
                            new_section = {
                                "roadway_title": row.get("Roadway_Title", ""),
                                "highway_dir": row.get("Highway_Dir", ""),
                                "xml_file": row.get("XML_File", ""),
                                "section_num": int(row.get("Section_Num", 0)),
                                "start_station": row.get("Start_Station", ""),
                                "end_station": row.get("End_Station", ""),
                                "year": row.get("Year", ""),
                                "current_aadt": row.get("Current_AADT", "1"),
                                "calculated_aadt": row.get("Calculated_AADT", ""),
                                "is_new": True,
                            }
                            for slot in range(1, 7):
                                new_section[f"id{slot}"] = row.get(f"ID{slot}", "")
                                new_section[f"sign{slot}"] = row.get(f"Sign{slot}", "+")
                            self._sections.append(new_section)
                            new_sections_added += 1
                            matched_count += 1
                            if row.get("Reviewed", "No") == "Yes":
                                reviewed_alignments.add(row.get("Roadway_Title", ""))
                        else:
                            unmatched_rows.append(
                                f"{row.get('Roadway_Title', '')} @ {row.get('Start_Station', '')}"
                            )

            if matched_count == 0:
                QMessageBox.warning(
                    self, "No Matches",
                    "No CSV rows matched the scanned sections.\n\n"
                    "Make sure the CSV was saved from the same project structure.",
                )
                return

            self._reviewed_alignments = reviewed_alignments

            # Refresh years
            self._years_in_project = sorted(
                {s["year"] for s in self._sections if s["year"]}
            )
            self._year_combo.blockSignals(True)
            self._year_combo.clear()
            self._year_combo.addItems(self._years_in_project)
            if self._years_in_project:
                if self._year_combo.currentText() not in self._years_in_project:
                    self._year_combo.setCurrentIndex(0)
            self._year_combo.blockSignals(False)
            self._year_status_label.setText(
                f"Years: {', '.join(self._years_in_project)}"
            )
            self._year_status_label.setStyleSheet(f"color: {theme.ACCENT_GREEN};")

            self._populate_tree()
            self._update_checklist()

            new_msg = (
                f" ({new_sections_added} new year entries restored)"
                if new_sections_added > 0
                else ""
            )
            self._scan_status_label.setText(
                f"Merged IDs for {matched_count} sections{new_msg}"
            )
            self.status_message.emit(
                f"Loaded forecast IDs from {os.path.basename(path)}"
            )

            unmatched_msg = ""
            if unmatched_rows:
                unmatched_msg = (
                    f"\n\n{len(unmatched_rows)} rows in CSV did not match scanned sections."
                )
            new_year_msg = (
                f"\n{new_sections_added} new year entries restored."
                if new_sections_added > 0
                else ""
            )
            QMessageBox.information(
                self, "Load Complete",
                f"Merged forecast IDs for {matched_count} sections.{new_year_msg}\n"
                f"{len(reviewed_alignments)} alignments marked as reviewed.\n"
                f"Current AADT values preserved from XML scan.{unmatched_msg}\n\n"
                "Note: Load the forecast workbook (Step 2) to calculate AADT values.",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Error loading mapping: {exc}")

    # ====================================================================
    # Step 5 -- Preview / Apply / Export
    # ====================================================================

    def _preview_changes(self):
        """Show dialog listing all sections with calculated AADT values."""
        changes: list = []
        new_entries: list = []
        for section in self._sections:
            if section["calculated_aadt"]:
                entry = (
                    f"{section['roadway_title']} Section {section['section_num']}: "
                    f"Sta {_format_station(section['start_station'])} to "
                    f"{_format_station(section['end_station'])} -> "
                    f"AADT {section['calculated_aadt']}"
                )
                if section.get("is_new", False):
                    new_entries.append(f"[NEW] {entry}")
                else:
                    changes.append(entry)
        changes = new_entries + changes

        if not changes:
            QMessageBox.information(
                self, "No Changes",
                "No AADT values calculated yet. "
                "Please enter forecast IDs and calculate first.",
            )
            return

        dlg = QDialog(self)
        dlg.setWindowTitle("Preview AADT Changes")
        dlg.resize(700, 500)
        lay = QVBoxLayout(dlg)

        lay.addWidget(QLabel(
            f"The following {len(changes)} changes will be made:"
        ))

        text_edit = QTextEdit()
        text_edit.setReadOnly(True)
        text_edit.setPlainText("\n".join(changes))
        lay.addWidget(text_edit, stretch=1)

        close_btn = QPushButton("Close")
        close_btn.clicked.connect(dlg.accept)
        lay.addWidget(close_btn)
        dlg.exec()

    def _apply_to_xml(self):
        """Modify highway XML files with calculated AADT values."""
        if not self._project_path or not os.path.isdir(self._project_path):
            QMessageBox.critical(
                self, "Error", "Please select a valid project folder first."
            )
            return

        project_dir = Path(self._project_path)

        # Group sections by XML file, reconstructing paths from current project
        file_sections: Dict[str, list] = {}
        path_errors: list = []

        for section in self._sections:
            if not section["calculated_aadt"]:
                continue
            stored_path = Path(section["xml_file"])
            highway_folder = stored_path.parent.name

            xml_file = None

            # Try direct path
            highway_dir = project_dir / highway_folder
            if highway_dir.exists():
                highway_xmls = list(highway_dir.glob("highway.*.xml"))
                if not highway_xmls:
                    highway_xmls = list(highway_dir.glob("highway.xml"))
                if highway_xmls:
                    xml_file = sorted(highway_xmls)[-1]

            # Search in interchange containers
            if xml_file is None:
                for c_dir in project_dir.glob("c*"):
                    potential_dir = c_dir / highway_folder
                    if potential_dir.exists():
                        highway_xmls = list(potential_dir.glob("highway.*.xml"))
                        if not highway_xmls:
                            highway_xmls = list(potential_dir.glob("highway.xml"))
                        if highway_xmls:
                            xml_file = sorted(highway_xmls)[-1]
                            break

            # Last resort: search anywhere
            if xml_file is None:
                matches = list(project_dir.glob(f"**/{highway_folder}/highway.*.xml"))
                if not matches:
                    matches = list(project_dir.glob(f"**/{highway_folder}/highway.xml"))
                if matches:
                    xml_file = sorted(matches)[-1]
                else:
                    path_errors.append(highway_folder)
                    continue

            xml_str = str(xml_file)
            file_sections.setdefault(xml_str, []).append(section)

        if path_errors:
            QMessageBox.warning(
                self, "Path Errors",
                f"Could not find {len(path_errors)} XML files in current project:\n"
                + "\n".join(path_errors[:10])
                + ("\n..." if len(path_errors) > 10 else "")
                + f"\n\nMake sure the project path is correct:\n{self._project_path}",
            )

        if not file_sections:
            QMessageBox.warning(
                self, "No Changes",
                "No AADT values to apply. Please calculate AADT values first.",
            )
            return

        total_changes = sum(len(v) for v in file_sections.values())
        reply = QMessageBox.question(
            self, "Confirm Changes",
            f"This will modify {total_changes} AADT values in {len(file_sections)} XML files.\n\n"
            "It is recommended to backup your project before proceeding.\n\n"
            "Continue?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            return

        self.status_message.emit("Applying AADT values to XML files...")
        QApplication.processEvents()

        updated_count = 0
        error_count = 0
        changes_list: list = []

        for xml_file, sections in file_sections.items():
            try:
                with open(xml_file, "r", encoding="utf-8") as f:
                    content = f.read()
                original_content = content

                for section in sections:
                    if section.get("is_new", False):
                        continue  # Handled separately below

                    new_aadt = section["calculated_aadt"]
                    start_sta = section["start_station"]
                    end_sta = section["end_station"]

                    def _station_pattern(sta_str):
                        try:
                            val = float(sta_str)
                            int_part = int(val)
                            frac_part = val - int_part
                            if frac_part == 0:
                                return rf"{int_part}(?:\.0+)?"
                            else:
                                decimal_str = str(val).split(".")[1].rstrip("0") or "0"
                                return rf"{int_part}\.{decimal_str}0*"
                        except (ValueError, TypeError):
                            return re.escape(sta_str)

                    start_pat = _station_pattern(start_sta)
                    end_pat = _station_pattern(end_sta)

                    pattern = (
                        r"(<AnnualAveDailyTraffic\s+"
                        rf'startStation="{start_pat}"\s+'
                        rf'endStation="{end_pat}"\s*'
                        r'[^>]*?adtRate=")(\d+)(")'
                    )

                    old_value = [None]

                    def _replace_aadt(match):
                        old_value[0] = match.group(2)
                        return match.group(1) + new_aadt + match.group(3)

                    content, count = re.subn(pattern, _replace_aadt, content)

                    if count == 0:
                        pattern2 = (
                            r"(<AnnualAveDailyTraffic\s+[^>]*?"
                            rf'startStation="{start_pat}"[^>]*?'
                            rf'endStation="{end_pat}"[^>]*?'
                            r'adtRate=")(\d+)(")'
                        )
                        content, count = re.subn(pattern2, _replace_aadt, content)

                    if count > 0 and old_value[0] and old_value[0] != new_aadt:
                        updated_count += 1
                        changes_list.append((
                            section["roadway_title"],
                            _format_station(start_sta),
                            _format_station(end_sta),
                            old_value[0],
                            new_aadt,
                        ))

                # Handle new sections (is_new=True)
                new_sections = [s for s in sections if s.get("is_new", False)]
                for section in new_sections:
                    new_aadt = section["calculated_aadt"]
                    if not new_aadt:
                        continue

                    start_sta = section["start_station"]
                    end_sta = section["end_station"]
                    year = section["year"]

                    new_element = (
                        f'  <AnnualAveDailyTraffic startStation="{start_sta}" '
                        f'endStation="{end_sta}" \n'
                        f'    adtYear="{year}" adtRate="{new_aadt}" />\n'
                    )

                    aadt_matches = list(
                        re.finditer(r"<AnnualAveDailyTraffic[^>]*/>", content)
                    )
                    if aadt_matches:
                        last_match = aadt_matches[-1]
                        insert_pos = last_match.end()
                        while insert_pos < len(content) and content[insert_pos] != "\n":
                            insert_pos += 1
                        insert_pos += 1

                        content = content[:insert_pos] + new_element + content[insert_pos:]
                        updated_count += 1
                        changes_list.append((
                            section["roadway_title"],
                            _format_station(start_sta),
                            _format_station(end_sta),
                            "NEW",
                            new_aadt,
                        ))
                        section["is_new"] = False
                        section["current_aadt"] = new_aadt

                if content != original_content:
                    with open(xml_file, "w", encoding="utf-8") as f:
                        f.write(content)
            except Exception as exc:
                error_count += 1

        # Refresh tree
        self._populate_tree()
        self._update_checklist()

        self._apply_status_label.setText(
            f"Updated {updated_count} AADT values. Errors: {error_count}"
        )
        self.status_message.emit(
            f"AADT update complete: {updated_count} values updated"
        )

        if error_count > 0:
            QMessageBox.warning(
                self, "Partial Success",
                f"Updated {updated_count} values with {error_count} errors.\n"
                "Check the console for error details.",
            )
        elif updated_count == 0:
            QMessageBox.information(
                self, "No Changes",
                "No AADT values needed updating - all values already match.",
            )
        else:
            changes_text = f"Updated {updated_count} AADT values:\n\n"
            for alignment, start, end, old_val, new_val in changes_list[:20]:
                changes_text += f"  {alignment}\n    {start} to {end}: {old_val} -> {new_val}\n"
            if len(changes_list) > 20:
                changes_text += f"\n  ... and {len(changes_list) - 20} more changes"
            QMessageBox.information(self, "Success", changes_text)

    def _export_mapping(self):
        """Export all section mappings to Excel with openpyxl."""
        if not self._sections:
            QMessageBox.warning(
                self, "No Data",
                "No AADT sections to export. Please scan the project first.",
            )
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export AADT Mapping", "",
            "Excel files (*.xlsx);;All files (*)",
        )
        if not path:
            return

        try:
            from openpyxl import Workbook

            wb = Workbook()
            ws = wb.active
            ws.title = "AADT Mapping"

            headers = [
                "Roadway_Title", "Section", "Start_Station", "End_Station", "Year",
                "Current_AADT", "Sign1", "ID1", "Sign2", "ID2", "Sign3", "ID3",
                "Sign4", "ID4", "Sign5", "ID5", "Sign6", "ID6",
                "Calculated_AADT", "Reviewed", "XML_File",
            ]
            ws.append(headers)

            for section in self._sections:
                is_reviewed = section["roadway_title"] in self._reviewed_alignments
                row = [
                    section["roadway_title"],
                    section["section_num"],
                    section["start_station"],
                    section["end_station"],
                    section["year"],
                    section["current_aadt"],
                    section.get("sign1", "+"), section.get("id1", ""),
                    section.get("sign2", "+"), section.get("id2", ""),
                    section.get("sign3", "+"), section.get("id3", ""),
                    section.get("sign4", "+"), section.get("id4", ""),
                    section.get("sign5", "+"), section.get("id5", ""),
                    section.get("sign6", "+"), section.get("id6", ""),
                    section["calculated_aadt"],
                    "Yes" if is_reviewed else "No",
                    section["xml_file"],
                ]
                ws.append(row)

            wb.save(path)
            QMessageBox.information(
                self, "Export Complete",
                f"AADT mapping exported to:\n{path}",
            )
        except Exception as exc:
            QMessageBox.critical(self, "Error", f"Error exporting: {exc}")
