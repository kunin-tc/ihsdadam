"""Evaluation Info tab -- CMF / calibration factor scanner and Excel exporter."""

import os
from typing import List

from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QTreeWidgetItem, QFileDialog, QMessageBox,
)
from PySide6.QtCore import Qt, Signal
from PySide6.QtGui import QColor, QFont

from ..models import CMFEntry
from ..workers import CMFScanWorker
from ..widgets import ScrollableTree
from .. import theme


class CMFTab(QWidget):
    """Evaluation Info tab displaying calibration factors per alignment."""

    status_message = Signal(str)
    progress_update = Signal(int, str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self._project_path = ""
        self._cmf_data: List[CMFEntry] = []
        self._worker = None
        self._setup_ui()

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    def set_project_path(self, path: str):
        self._project_path = path

    # ------------------------------------------------------------------
    # UI construction
    # ------------------------------------------------------------------

    def _setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(8)

        # -- top row: buttons + summary ---------------------------------
        top_row = QHBoxLayout()
        top_row.setSpacing(8)

        self._scan_btn = QPushButton("Scan for Evaluation Info")
        self._scan_btn.setProperty("accent", True)
        self._scan_btn.clicked.connect(self._scan_cmf)
        top_row.addWidget(self._scan_btn)

        self._export_btn = QPushButton("Export to Excel")
        self._export_btn.setEnabled(False)
        self._export_btn.clicked.connect(self._export_to_excel)
        top_row.addWidget(self._export_btn)

        self._summary_label = QLabel("No scan performed yet")
        self._summary_label.setProperty("subheader", True)
        top_row.addWidget(self._summary_label)

        top_row.addStretch()
        layout.addLayout(top_row)

        # -- tree -------------------------------------------------------
        self._tree = ScrollableTree(
            ["Type", "ID", "Name", "Evaluation", "Evaluation Years", "Calibration Factor"],
        )
        self._tree.set_column_widths([120, 80, 280, 100, 140, 180])
        layout.addWidget(self._tree, stretch=1)

    # ------------------------------------------------------------------
    # Scan
    # ------------------------------------------------------------------

    def _scan_cmf(self):
        if not self._project_path or not os.path.isdir(self._project_path):
            QMessageBox.warning(
                self, "Invalid Project",
                "Please select a valid project directory first.",
            )
            return

        self._scan_btn.setEnabled(False)
        self._export_btn.setEnabled(False)
        self._tree.clear()
        self._cmf_data = []
        self._summary_label.setText("Scanning...")
        self.status_message.emit("Scanning for evaluation info...")

        self._worker = CMFScanWorker(self._project_path, parent=self)
        self._worker.progress.connect(self._on_scan_progress)
        self._worker.finished.connect(self._on_scan_finished)
        self._worker.error.connect(self._on_scan_error)
        self._worker.start()

    def _on_scan_progress(self, pct: int, msg: str):
        self.progress_update.emit(pct, msg)

    def _on_scan_finished(self, entries: list):
        self._cmf_data = list(entries)
        self._scan_btn.setEnabled(True)
        self._populate_tree(entries)

        highway_count = sum(1 for e in entries if e.type == "Highway")
        intersection_count = sum(1 for e in entries if e.type == "Intersection")
        ramp_count = sum(1 for e in entries if e.type == "Ramp Terminal")

        summary = (
            f"Total: {len(entries)}  |  "
            f"Highways: {highway_count}  |  "
            f"Intersections: {intersection_count}  |  "
            f"Ramp Terminals: {ramp_count}"
        )
        self._summary_label.setText(summary)
        self.status_message.emit(f"Found {len(entries)} evaluation info entries")

        self._export_btn.setEnabled(bool(entries))
        self._worker = None

    def _on_scan_error(self, msg: str):
        self._scan_btn.setEnabled(True)
        self._summary_label.setText("Scan failed")
        self.status_message.emit(f"Evaluation info scan failed: {msg}")
        QMessageBox.critical(self, "Scan Error", f"Error scanning evaluation info:\n{msg}")
        self._worker = None

    # ------------------------------------------------------------------
    # Tree population
    # ------------------------------------------------------------------

    def _populate_tree(self, entries: List[CMFEntry]):
        self._tree.clear()

        highways = [e for e in entries if e.type == "Highway"]
        intersections = [e for e in entries if e.type == "Intersection"]
        ramps = [e for e in entries if e.type == "Ramp Terminal"]

        groups = [
            (f"HIGHWAYS ({len(highways)})", highways),
            (f"INTERSECTIONS ({len(intersections)})", intersections),
            (f"RAMP TERMINALS ({len(ramps)})", ramps),
        ]

        bold_font = QFont()
        bold_font.setBold(True)

        for group_label, group_entries in groups:
            if not group_entries:
                continue

            parent_item = QTreeWidgetItem([group_label, "", "", "", "", ""])
            parent_item.setFont(0, bold_font)
            parent_item.setForeground(0, QColor(theme.PRIMARY))
            self._tree.tree.addTopLevelItem(parent_item)

            for entry in sorted(group_entries, key=lambda e: e.id):
                child = QTreeWidgetItem([
                    entry.type,
                    entry.id,
                    entry.name,
                    entry.evaluation,
                    entry.years,
                    entry.calibration,
                ])
                child.setTextAlignment(1, Qt.AlignCenter)
                child.setTextAlignment(4, Qt.AlignCenter)
                parent_item.addChild(child)

            parent_item.setExpanded(True)

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------

    def _export_to_excel(self):
        if not self._cmf_data:
            QMessageBox.warning(
                self, "No Data",
                "No evaluation info data to export. Please scan first.",
            )
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font as XlFont, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(
                self, "Missing Dependency",
                "openpyxl is required for Excel export.\n"
                "Install it with:  pip install openpyxl",
            )
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Export Evaluation Info",
            "Evaluation_Info.xlsx",
            "Excel Files (*.xlsx);;All Files (*)",
        )
        if not path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Evaluation Info"

            headers = [
                "Type", "ID", "Name", "Evaluation",
                "Evaluation Years", "Calibration Factor", "File Path",
            ]
            ws.append(headers)

            header_fill = PatternFill(
                start_color="366092", end_color="366092", fill_type="solid",
            )
            header_font = XlFont(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for entry in self._cmf_data:
                ws.append([
                    entry.type,
                    entry.id,
                    entry.name,
                    entry.evaluation,
                    entry.years,
                    entry.calibration,
                    entry.path,
                ])

            ws.column_dimensions["A"].width = 15
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 50
            ws.column_dimensions["D"].width = 15
            ws.column_dimensions["E"].width = 25
            ws.column_dimensions["F"].width = 60

            wb.save(path)

            QMessageBox.information(
                self, "Export Successful",
                f"Evaluation info exported to:\n{path}",
            )
            self.status_message.emit(
                f"Exported evaluation info to {os.path.basename(path)}"
            )
        except Exception as exc:
            QMessageBox.critical(
                self, "Export Error",
                f"Failed to export to Excel:\n{exc}",
            )
