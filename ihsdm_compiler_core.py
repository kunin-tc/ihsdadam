"""
IHSDM Compiler Core Functions
Extraction and compilation functions for IHSDM crash prediction data.

Original Author: Adam Engbring (aengbring@hntb.com)
Integrated by: Claude Code
Date: 2025-12-17
"""

import os
import csv
import xml.etree.ElementTree as ET
from collections import Counter
from openpyxl import Workbook, load_workbook


# =============================================================================
# XML EVALUATION NAME EXTRACTION
# =============================================================================

def get_evaluation_title_from_xml(eval_dir):
    """Extract evaluationTitle from evaluation.1.result.xml in the given directory.

    Args:
        eval_dir: Directory containing the evaluation files (e.g., .../h1/e1/)

    Returns:
        The evaluationTitle attribute value, or "Unknown" if not found.
    """
    xml_path = os.path.join(eval_dir, "evaluation.1.result.xml")

    if not os.path.exists(xml_path):
        return "Unknown"

    try:
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # Try with namespace first
        ns = "{http://www.ihsdm.org/schema/Highway-1.0}"
        eval_info = root.find(f".//{ns}EvaluationInfo")

        # Try without namespace if not found
        if eval_info is None:
            eval_info = root.find(".//EvaluationInfo")

        if eval_info is not None:
            return eval_info.get("evaluationTitle", "Unknown")

        return "Unknown"
    except Exception as e:
        print(f"Error reading evaluation XML: {e}")
        return "Unknown"


# =============================================================================
# HSM SEVERITY DISTRIBUTIONS
# =============================================================================

HSM_SEVERITY_K = 0.0146      # 1.46% fatal (highway)
HSM_SEVERITY_A = 0.044764    # 4.48% incapacitating injury (highway)
HSM_SEVERITY_B = 0.2469      # 24.69% non-incapacitating injury (highway)
HSM_SEVERITY_C = 0.69172     # 69.17% possible injury (highway)

INT_SEVERITY_K = 0.002575    # 0.26% fatal (intersection/ramp)
INT_SEVERITY_A = 0.053525    # 5.35% incapacitating injury (intersection/ramp)
INT_SEVERITY_B = 0.276415    # 27.64% non-incapacitating injury (intersection/ramp)
INT_SEVERITY_C = 0.667485    # 66.75% possible injury (intersection/ramp)


# =============================================================================
# OUTPUT HEADERS
# =============================================================================

HIGHWAY_HEADER = ["Evaluation Name", "Segment", "Segment #", "Year", "AADT", "Start_Location", "Type",
                  "Length miles", "Total K", "Total A", "Total B", "Total C", "Total PD", "FI", "PDO"]

INTERSECTION_HEADER = ["Evaluation Name", "Inter. #", "Intersection Type", "Title", "Year", "Major AADT", "Minor AADT",
                       "Fatal (K) Crashes", "Incapacitating Injury (A) Crashes",
                       "Non-Incapacitating Injury (B) Crashes", "Possible Injury (C) Crashes",
                       "No Injury (O) Crashes", "Fatal and Injury (FI) Crashes"]

RAMP_TERMINAL_HEADER = ["Evaluation Name", "Ramp Terminal #", "Ramp Terminal Type", "Title", "Year", "Exit AADT", "Entrance AADT",
                        "Fatal (K) Crashes", "Incapacitating Injury (A) Crashes",
                        "Non-Incapacitating Injury (B) Crashes", "Possible Injury (C) Crashes",
                        "No Injury (O) Crashes", "Fatal and Injury (FI) Crashes"]

# Site set headers - use Calibrated FI as the FI source for KABC calculation
SITESET_INT_HEADER = ["Evaluation Name", "Inter. #", "Intersection Type", "Title", "Year", "Major AADT", "Minor AADT",
                      "Fatal (K) Crashes", "Incapacitating Injury (A) Crashes",
                      "Non-Incapacitating Injury (B) Crashes", "Possible Injury (C) Crashes",
                      "No Injury (O) Crashes", "Calibrated FI Predicted Crashes Per Year"]

SITESET_RAMP_HEADER = ["Evaluation Name", "Ramp Terminal #", "Ramp Terminal Type", "Title", "Year", "Exit AADT", "Entrance AADT",
                       "Fatal (K) Crashes", "Incapacitating Injury (A) Crashes",
                       "Non-Incapacitating Injury (B) Crashes", "Possible Injury (C) Crashes",
                       "No Injury (O) Crashes", "Calibrated FI Predicted Crashes Per Year"]


# =============================================================================
# UTILITY FUNCTIONS
# =============================================================================

def find_folders_with_file(root_dir, target_filename):
    """Recursively search for folders containing the target file."""
    matching_folders = []
    for root, dirs, files in os.walk(root_dir):
        if target_filename in files:
            parent_folder = os.path.dirname(root)
            matching_folders.append(parent_folder)
    return matching_folders


def remove_duplicates(rows):
    """Remove duplicate rows while preserving order."""
    row_counts = Counter(tuple(row) for row in rows)
    duplicates = {row: count for row, count in row_counts.items() if count > 1}

    seen = set()
    unique_rows = []
    for row in rows:
        row_tuple = tuple(row)
        if row_tuple not in seen:
            seen.add(row_tuple)
            unique_rows.append(row)

    return unique_rows, duplicates


def deduplicate_by_title(rows, title_column_index=3):
    """Remove duplicate rows by title, keeping first occurrence.

    Used for intersection deduplication where the same intersection may appear
    in multiple evaluation files.

    Args:
        rows: List of data rows
        title_column_index: Column index containing the title (default 3 for Intersection)
                           With "Evaluation Name" at index 0, Title is at index 3

    Returns:
        List of unique rows (first occurrence of each title kept)
    """
    seen_titles = set()
    unique_rows = []
    for row in rows:
        if len(row) > title_column_index:
            title = row[title_column_index]
            if title and title not in seen_titles:
                seen_titles.add(title)
                unique_rows.append(row)
        else:
            # Keep rows that don't have enough columns (header rows, etc.)
            unique_rows.append(row)
    return unique_rows


# =============================================================================
# HIGHWAY SEGMENT EXTRACTION
# =============================================================================

def extract_highway_segments_from_csv(file_path, start_index=5, end_index=None):
    """Extract highway segment rows from IHSDM evaluation CSV file."""
    extracted_rows = []
    try:
        with open(file_path, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            lines = list(reader)

            if len(lines) < start_index + 2:
                return []

            if end_index is None:
                for i in range(start_index, len(lines)):
                    if len(lines[i]) > 1 and (lines[i][1] == "" or "Crash Proportion" in lines[i][1]):
                        end_index = i
                        break
                else:
                    end_index = len(lines)

            for row in lines[start_index:end_index]:
                if row:
                    extracted_rows.append(row)

    except Exception as e:
        print(f"Error reading file {file_path}: {e}")

    return extracted_rows


def should_process_highway_row(row):
    """Determine if a highway row should be processed."""
    if not row or len(row) == 0:
        return False
    if row[0] == "Highway Title":
        return False
    if "*************" in row:
        return False
    if "Status" in row or "Crash predictionXYZ" in row or "Crash Proportion" in row:
        return False
    if "Seg. No." in row or row[0] == "Seg. No.":
        return False
    if "USAIntersection" in row[0]:
        return False
    return True


def extract_highway_row_data(row, eval_name="", debug=False):
    """Extract relevant columns from highway row based on facility type.

    Args:
        row: The CSV row data
        eval_name: Evaluation name to prepend to the output row
        debug: Enable debug output
    """
    filtered_row = None

    try:
        # Urban/Rural Arterials
        if "Arterial" in row[0]:
            try:
                filtered_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[9],
                                row[83], row[84], row[85], row[86], row[87], row[88], row[89]]
                if debug: print("Classified as urban arterial")
            except IndexError:
                try:
                    filtered_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[9],
                                    0, 0, 0, 0, 0, row[55], row[56]]
                    if debug: print("Classified as rural arterial (55-56)")
                except IndexError:
                    filtered_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[9],
                                    0, 0, 0, 0, 0, row[44], row[45]]
                    if debug: print("Classified as rural arterial (44-45)")

        elif "Ramp" in row[0] or "Exit" in row[0] or "Entrance" in row[0]:
            filtered_row = [row[0], row[1], row[3], row[4], row[5], row[6], row[10],
                            row[78], row[79], row[80], row[81], row[82], row[83]]
            if debug: print("Classified as ramp/entrance/exit")

        elif "FR" in row[0]:
            try:
                filtered_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[9],
                                row[83], row[84], row[85], row[86], row[87], row[88], row[89]]
                if debug: print("Classified frontage road as arterial")
            except IndexError:
                filtered_row = [row[0], row[1], row[3], row[4], row[5], row[6], row[10],
                                row[78], row[79], row[80], row[81], row[82], row[83]]
                if debug: print("Classified frontage road as ramp")

        elif "Mainline" in row[0] and "urban" in row[2] and ("entrance" in row[3] or "exit" in row[3]):
            filtered_row = [row[0], row[1], row[4], row[5], row[8], row[9], row[13],
                            row[49], row[50], row[51], row[52], row[53]]
            if debug: print("Classified as urban mainline speed change")

        elif "Mainline" in row[0] and "urban" in row[2]:
            filtered_row = [row[0], row[1], row[3], row[4], row[10], row[11], row[15],
                            row[82], row[83], row[84], row[85], row[86]]
            if debug: print("Classified as urban mainline")

        elif "Mainline" in row[0] and "rural" in row[2] and ("entrance" in row[3] or "exit" in row[3]):
            try:
                filtered_row = [row[0], row[1], row[4], row[5], row[8], row[9], row[13],
                                0, 0, 0, 0, 0, row[52], row[53]]
                if debug: print("Classified as rural mainline speed change")
            except IndexError:
                if debug: print("Rural mainline speed change - column index error")
                return None

        elif "Mainline" in row[0] and "rural" in row[2]:
            try:
                filtered_row = [row[0], row[1], row[3], row[4], row[10], row[11], row[15],
                                row[82], row[83], row[84], row[85], row[86]]
                if debug: print("Classified as rural mainline")
            except IndexError:
                filtered_row = [row[0], row[1], row[3], row[4], row[10], row[11], row[14],
                                row[82], row[83], row[84], row[85], row[86], row[87]]
                if debug: print("Classified as rural mainline (alternate)")

        elif "(Special Eval)" in row[0]:
            filtered_row = [row[0], row[1], row[3], row[4], row[5], row[6], row[10],
                            row[78], row[79], row[80], row[81], row[82], row[83], row[84]]
            if debug: print("Classified as special evaluation")

        elif "" in row[0] and any(xsec in row[5] for xsec in ["4U", "6U", "8U", "10U", "2U", "4D", "6D", "8D", "10D", "2O", "4O", "3O", "5T", "3T"]):
            try:
                filtered_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[9],
                                row[83], row[84], row[85], row[86], row[87], row[88], row[89]]
                if debug: print("Classified unnamed as arterial by cross-section")
            except IndexError:
                try:
                    filtered_row = [row[0], row[1], row[3], row[4], row[5], row[6], row[10],
                                    row[78], row[79], row[80], row[81], row[82]]
                    if debug: print("Classified unnamed as ramp")
                except IndexError:
                    try:
                        filtered_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[9],
                                        0, 0, 0, 0, 0, row[55], row[56]]
                        if debug: print("Classified unnamed as rural arterial (55-56)")
                    except IndexError:
                        filtered_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[9],
                                        0, 0, 0, 0, 0, row[44], row[45]]
                        if debug: print("Classified unnamed as rural arterial (44-45)")

        elif "CD" in row[0] and "rural" in row[2]:
            try:
                filtered_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[9],
                                row[83], row[84], row[85], row[86], row[87], row[88], row[89]]
                if debug: print("Classified as rural CD road")
            except IndexError:
                return None

        elif ("2CD" in row[6] or "1CD" in row[6]) and "urban" in row[2]:
            try:
                filtered_row = [row[0], row[1], row[3], row[4], row[5], row[6], row[10],
                                row[78], row[79], row[80], row[81], row[82]]
                if debug: print("Classified as urban CD road")
            except IndexError:
                return None

        elif "" in row[0] and any(ramp_type in row[6] for ramp_type in ["1EX", "1EN", "2EX", "2EN"]):
            try:
                filtered_row = [row[0], row[1], row[3], row[4], row[5], row[6], row[10],
                                row[78], row[79], row[80], row[81], row[82], row[83]]
                if debug: print("Classified unnamed as ramp by type indicator")
            except IndexError:
                return None

        else:
            try:
                filtered_row = [row[0], row[1], row[2], row[3], row[4], row[5], row[9],
                                0, 0, 0, 0, 0, row[44], row[45]]
                if debug: print("Using default fallback structure")
            except IndexError:
                filtered_row = [row[0], row[1], row[3], row[4], row[5], row[6], row[10],
                                row[78], row[79], row[80], row[81], row[82], row[83]]
                if debug: print("Using final fallback (ramp structure)")

    except Exception as e:
        print(f"Error extracting highway row data: {e}")
        return None

    # Prepend evaluation name to the row
    if filtered_row is not None:
        filtered_row = [eval_name] + filtered_row

    return filtered_row


# =============================================================================
# INTERSECTION & RAMP TERMINAL EXTRACTION
# =============================================================================

def extract_by_headers_from_csv(file_path, target_headers, first_file=False, multi_year=True, eval_name=""):
    """Extract rows from CSV by matching column headers.

    Args:
        file_path: Path to the CSV file
        target_headers: List of headers to extract
        first_file: If True, include header row in output
        multi_year: If True, extract 20 years of data
        eval_name: Evaluation name to prepend to each data row
    """
    extracted_rows = []
    try:
        with open(file_path, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            lines = list(reader)

            if len(lines) < 7:
                print(f"Warning: File {file_path} does not have enough rows. Found: {len(lines)}")
                return []

            header_row = lines[5]  # Row 6 (index 5) contains headers

            # Determine rows to extract
            if multi_year:
                if first_file:
                    rows_to_extract = lines[5:26]  # Include header + 20 years
                else:
                    rows_to_extract = lines[6:26]  # Skip header, just data
            else:
                if first_file:
                    rows_to_extract = lines[5:7]  # Header + first data row
                else:
                    rows_to_extract = [lines[6]]  # Single data row

            # Build header index map
            header_index_map = {header: [] for header in target_headers}
            for idx, header in enumerate(header_row):
                if header in target_headers:
                    header_index_map[header].append(idx)

            # Extract data
            is_first_row = True
            for row in rows_to_extract:
                if len(row) < len(header_row):
                    continue

                # Skip invalid rows
                if not row or len(row) == 0:
                    continue
                if any(marker in str(row) for marker in ["*************", "Crash Proportion", "Seg. No."]):
                    continue
                if row[0] == "" or row[0] == "Type":
                    continue

                extracted_row = []
                for header in target_headers:
                    if header in header_index_map:
                        indices = header_index_map[header]
                        for idx in indices:
                            extracted_row.append(row[idx] if idx < len(row) else "")

                # Prepend "Evaluation Name" header text for first row (header),
                # or actual eval_name value for data rows
                if first_file and is_first_row:
                    extracted_row = ["Evaluation Name"] + extracted_row
                    is_first_row = False
                else:
                    extracted_row = [eval_name] + extracted_row
                extracted_rows.append(extracted_row)

    except Exception as e:
        print(f"Error reading file {file_path}: {e}")

    return extracted_rows


def extract_site_set_data(file_path, target_headers, section_marker, first_file=False, eval_name=""):
    """Extract data from site set CSV files.

    Site set CSVs have multiple sections, each with a marker line like
    '*************","USA Intersection Debug Result' followed by a header row and data row.

    Args:
        file_path: Path to the CSV file
        target_headers: List of headers to extract
        section_marker: Marker text to identify sections (e.g., "USA Intersection Debug Result")
        first_file: If True, include header row in output
        eval_name: Evaluation name to prepend to each data row

    Returns:
        List of extracted rows
    """
    extracted_rows = []
    try:
        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            lines = list(reader)

            i = 0
            include_header = first_file

            while i < len(lines):
                row = lines[i]

                # Look for section marker
                if len(row) >= 2 and "*************" in str(row[0]) and section_marker in str(row[1]):
                    # Next row should be the header
                    if i + 1 < len(lines):
                        header_row = lines[i + 1]

                        # Build header index map
                        header_index_map = {header: [] for header in target_headers}
                        for idx, header in enumerate(header_row):
                            if header in target_headers:
                                header_index_map[header].append(idx)

                        # Include header row if this is first file and we haven't included it yet
                        if include_header:
                            extracted_row = []
                            for header in target_headers:
                                if header in header_index_map:
                                    indices = header_index_map[header]
                                    for idx in indices:
                                        extracted_row.append(header_row[idx] if idx < len(header_row) else "")
                            # Prepend "Evaluation Name" header text to header row
                            extracted_row = ["Evaluation Name"] + extracted_row
                            extracted_rows.append(extracted_row)
                            include_header = False

                        # Extract data row (row after header)
                        if i + 2 < len(lines):
                            data_row = lines[i + 2]

                            # Skip if empty row
                            if data_row and len(data_row) > 0 and data_row[0] != "":
                                extracted_row = []
                                for header in target_headers:
                                    if header in header_index_map:
                                        indices = header_index_map[header]
                                        for idx in indices:
                                            extracted_row.append(data_row[idx] if idx < len(data_row) else "")
                                # Prepend evaluation name to data row
                                extracted_row = [eval_name] + extracted_row
                                extracted_rows.append(extracted_row)

                        i += 3  # Skip marker, header, and data rows
                        continue

                i += 1

    except Exception as e:
        print(f"Error reading site set file {file_path}: {e}")

    return extracted_rows


def scan_site_set_sections(file_path):
    """Scan a site set CSV file and return all section types found.

    Returns:
        dict with keys: 'intersections', 'ramp_terminals', 'unknown'
        Each value is a list of section names found
    """
    sections = {
        'intersections': [],
        'ramp_terminals': [],
        'unknown': []
    }

    # Known section markers
    known_intersection_markers = ["USA Intersection Debug Result"]
    known_ramp_markers = ["Ramp Terminal CMF"]
    # These are category headers, not data sections
    ignore_markers = ["Urban/Suburban Arterial", "Freeway Ramp Terminal", "Rural Two-Lane",
                      "Rural Multilane", "Urban/Suburban Freeway", ""]

    try:
        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            lines = list(reader)

            for i, row in enumerate(lines):
                if len(row) >= 2 and "*************" in str(row[0]):
                    section_name = row[1].strip() if len(row) > 1 else ""

                    if section_name in ignore_markers:
                        continue
                    elif section_name in known_intersection_markers:
                        if section_name not in sections['intersections']:
                            sections['intersections'].append(section_name)
                    elif section_name in known_ramp_markers:
                        if section_name not in sections['ramp_terminals']:
                            sections['ramp_terminals'].append(section_name)
                    else:
                        # Unknown section type
                        if section_name not in sections['unknown']:
                            sections['unknown'].append(section_name)

    except Exception as e:
        print(f"Error scanning site set file {file_path}: {e}")

    return sections


def extract_unknown_site_set_sections(file_path, known_markers):
    """Extract data from unknown section types in site set CSV.

    Attempts to extract any section that isn't in known_markers list.
    Uses a generic approach - grabs all columns from header and data rows.

    Args:
        file_path: Path to CSV file
        known_markers: List of section markers to skip (already handled)

    Returns:
        List of tuples: [(section_name, header_row, data_rows), ...]
    """
    results = []

    # Category headers to ignore (not data sections)
    ignore_markers = ["Urban/Suburban Arterial", "Freeway Ramp Terminal", "Rural Two-Lane",
                      "Rural Multilane", "Urban/Suburban Freeway", ""]

    try:
        with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
            reader = csv.reader(csvfile)
            lines = list(reader)

            i = 0
            while i < len(lines):
                row = lines[i]

                # Look for section marker
                if len(row) >= 2 and "*************" in str(row[0]):
                    section_name = row[1].strip() if len(row) > 1 else ""

                    # Skip known and ignored markers
                    if section_name in ignore_markers or section_name in known_markers:
                        i += 1
                        continue

                    # Found an unknown section - extract header and data
                    if i + 1 < len(lines):
                        header_row = lines[i + 1]

                        # Collect data rows until next marker or empty row
                        data_rows = []
                        j = i + 2
                        while j < len(lines):
                            data_row = lines[j]
                            # Stop at next marker or empty row
                            if not data_row or len(data_row) == 0:
                                break
                            if len(data_row) >= 1 and "*************" in str(data_row[0]):
                                break
                            if data_row[0] == "":
                                break
                            data_rows.append(data_row)
                            j += 1

                        if data_rows:
                            results.append((section_name, header_row, data_rows))

                        i = j
                        continue

                i += 1

    except Exception as e:
        print(f"Error extracting unknown sections from {file_path}: {e}")

    return results


# =============================================================================
# EXCEL WRITING FUNCTIONS
# =============================================================================

def write_rows_to_excel(rows, excel_path, sheet_name):
    """Write rows to an Excel sheet."""
    try:
        if not os.path.exists(excel_path):
            workbook = Workbook()
        else:
            workbook = load_workbook(excel_path)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(title=sheet_name)

        for row in rows:
            if row:
                sheet.append(row)

        workbook.save(excel_path)
    except Exception as e:
        print(f"Error writing to file {excel_path}: {e}")


def add_header_to_excel(excel_path, sheet_name, header):
    """Insert header row at the top of the Excel sheet."""
    try:
        workbook = load_workbook(excel_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            sheet.insert_rows(1)
            for col_index, value in enumerate(header, start=1):
                sheet.cell(row=1, column=col_index, value=value)
        workbook.save(excel_path)
    except Exception as e:
        print(f"Error adding header to file {excel_path}: {e}")


def fill_missing_highway_values(excel_path):
    """Apply HSM severity distributions to highway segments."""
    try:
        workbook = load_workbook(excel_path)

        if "Highway" in workbook.sheetnames:
            sheet = workbook["Highway"]
            for row in sheet.iter_rows(min_row=2):  # Skip header
                if len(row) > 13:
                    # Column indices shifted by 1 due to "Evaluation Name" in column 0
                    cell_k = row[8]
                    cell_a = row[9]
                    cell_b = row[10]
                    cell_c = row[11]
                    cell_pd = row[12]
                    cell_fi = row[13]
                    cell_pdo = row[14]

                    if cell_fi.value:
                        fi_value = float(cell_fi.value)

                        if not cell_k.value:
                            cell_k.value = fi_value * HSM_SEVERITY_K
                        if not cell_a.value:
                            cell_a.value = fi_value * HSM_SEVERITY_A
                        if not cell_b.value:
                            cell_b.value = fi_value * HSM_SEVERITY_B
                        if not cell_c.value:
                            cell_c.value = fi_value * HSM_SEVERITY_C
                        if not cell_pd.value:
                            cell_pd.value = cell_pdo.value if cell_pdo.value else ""

            workbook.save(excel_path)

    except Exception as e:
        print(f"Error filling highway values: {e}")


def fill_missing_intersection_values(excel_path, sheet_name="Intersection"):
    """Apply intersection severity distributions."""
    try:
        workbook = load_workbook(excel_path)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2):
                if len(row) > 13:
                    # Column indices shifted by 1 due to "Evaluation Name" in column 0
                    cell_f = row[6]   # Minor AADT
                    cell_g = row[7]   # Fatal (K)
                    cell_h = row[8]   # Incapacitating (A)
                    cell_i = row[9]   # Non-Incapacitating (B)
                    cell_j = row[10]  # Possible (C)
                    cell_k = row[11]  # No Injury (O)
                    cell_l = row[12]  # No Injury (O) duplicate
                    cell_m = row[13]  # FI

                    if cell_f.value and cell_l.value:
                        l_value = float(cell_l.value)
                        m_value = float(cell_m.value)

                        if not cell_g.value:
                            cell_g.value = m_value * INT_SEVERITY_K
                        if not cell_h.value:
                            cell_h.value = m_value * INT_SEVERITY_A
                        if not cell_i.value:
                            cell_i.value = m_value * INT_SEVERITY_B
                        if not cell_j.value:
                            cell_j.value = m_value * INT_SEVERITY_C

                        if not cell_k.value:
                            cell_k.value = l_value

            workbook.save(excel_path)

    except Exception as e:
        print(f"Error filling intersection values: {e}")


def fill_missing_ramp_terminal_values(excel_path, sheet_name="RampTerminal"):
    """Apply severity distributions to ramp terminals."""
    try:
        workbook = load_workbook(excel_path)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2):
                if len(row) > 12:
                    # Column indices shifted by 1 due to "Evaluation Name" in column 0
                    cell_k = row[7]
                    cell_a = row[8]
                    cell_b = row[9]
                    cell_c = row[10]
                    cell_o = row[11]
                    cell_fi = row[12]

                    if cell_fi.value:
                        fi_value = float(cell_fi.value)

                        if not cell_k.value:
                            cell_k.value = fi_value * INT_SEVERITY_K
                        if not cell_a.value:
                            cell_a.value = fi_value * INT_SEVERITY_A
                        if not cell_b.value:
                            cell_b.value = fi_value * INT_SEVERITY_B
                        if not cell_c.value:
                            cell_c.value = fi_value * INT_SEVERITY_C

            workbook.save(excel_path)

    except Exception as e:
        print(f"Error filling ramp terminal values: {e}")


def scrub_duplicate_columns(excel_path, sheet_name):
    """Remove duplicate columns from intersection/ramp sheets."""
    try:
        workbook = load_workbook(excel_path)

        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # Delete columns 16, 15, 14, 13 (reverse order)
            # Column indices shifted by 1 due to "Evaluation Name" in column 1
            for col in range(16, 12, -1):
                sheet.delete_cols(col)
            workbook.save(excel_path)

    except Exception as e:
        print(f"Error scrubbing columns in {sheet_name}: {e}")
