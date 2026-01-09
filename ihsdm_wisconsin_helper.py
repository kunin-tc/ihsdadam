"""
IHSDadaM
Comprehensive tool for IHSDM project analysis and data compilation.

Features:
- Warning Message Extraction (supports highways, intersections, ramp terminals)
- Data Compilation to Excel (crash predictions with HSM severity distributions)

Author: Claude Code
Original IHSDM Compiler by: Adam Engbring (aengbring@hntb.com)
Date: 2025-12-17
Version: 1.0
"""

import os
import csv
import xml.etree.ElementTree as ET
from pathlib import Path
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
from typing import List, Dict, Tuple
from dataclasses import dataclass, field
from collections import defaultdict, Counter
import subprocess
import time
import threading
import webbrowser
import json
from urllib.request import urlopen, Request
from urllib.error import URLError

# Import version info
try:
    from version import __version__, __app_name__, GITHUB_API_URL, GITHUB_RELEASES_URL
except ImportError:
    __version__ = "1.0.0"
    __app_name__ = "IHSDM Wisconsin Helper"
    GITHUB_API_URL = None
    GITHUB_RELEASES_URL = None

try:
    from openpyxl import Workbook, load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl not found. Data Compiler tab will be disabled.")
    print("Install with: pip install openpyxl")


# =============================================================================
# DATA STRUCTURES
# =============================================================================

@dataclass
class ResultMessage:
    """Container for a result message from IHSDM evaluation"""
    alignment_type: str  # 'h' for highway, 'i' for intersection, 'r' for ramp terminal
    alignment_id: str    # e.g., 'h74', 'i100', 'r25'
    alignment_name: str  # e.g., 'Mainline - I-375 Centerline'
    evaluation: str      # e.g., 'e20'
    start_sta: str
    end_sta: str
    message: str
    status: str          # 'info', 'warning', 'error', 'fault', 'CRITICAL'
    file_path: str
    is_critical: bool = False  # True if contains "no crash prediction supported"


# =============================================================================
# HSM SEVERITY DISTRIBUTIONS
# =============================================================================

HSM_SEVERITY_K = 0.0146      # 1.46% fatal (highway)
HSM_SEVERITY_A = 0.044764    # 4.48% incapacitating injury (highway)
HSM_SEVERITY_B = 0.2469      # 24.69% non-incapacitating injury (highway)
HSM_SEVERITY_C = 0.69172     # 69.17% possible injury (highway)

INT_SEVERITY_K = 0.002575    # 0.26% fatal (intersection/ramp)
INT_SEVERITY_A = 0.053525    # 5.35% incapacitating injury (intersection/ramp)
INT_SEVERITY_B = 0.276415    # 27.64% non-incapacitating injury (intersection/ramp)
INT_SEVERITY_C = 0.667485    # 66.75% possible injury (intersection/ramp)


# =============================================================================
# MAIN APPLICATION CLASS
# =============================================================================

class IHSDMWisconsinHelper:
    """Main application class combining warning extraction and data compilation"""

    def __init__(self):
        self.root = tk.Tk()
        self.root.title(f"{__app_name__} v{__version__}")
        self.root.geometry("1400x800")

        # Set color scheme
        self.colors = {
            'primary': '#1e3a8a',      # Deep blue
            'secondary': '#3b82f6',    # Bright blue
            'accent': '#10b981',       # Green
            'warning': '#f59e0b',      # Orange
            'danger': '#ef4444',       # Red
            'bg_light': '#f8fafc',     # Light gray
            'bg_dark': '#1e293b',      # Dark gray
            'text_light': '#ffffff',   # White
            'text_dark': '#1f2937',    # Dark text
        }

        # Configure root window
        self.root.configure(bg=self.colors['bg_light'])

        # Setup custom styles
        self.setup_styles()

        # Shared state
        self.project_path = tk.StringVar()
        self.messages: List[ResultMessage] = []
        self.filtered_messages: List[ResultMessage] = []

        # Compiler state
        self.target_file = tk.StringVar(value="evaluation.1.diag.csv")
        self.excel_output = tk.StringVar()
        self.multi_year_mode = tk.BooleanVar(value=False)  # Default to single-year mode
        self.debug_mode = tk.BooleanVar(value=False)

        self.setup_ui()

    def setup_styles(self):
        """Setup custom ttk styles for a modern look"""
        style = ttk.Style()

        # Use a modern theme as base
        try:
            style.theme_use('clam')  # Modern, customizable theme
        except:
            pass  # Fall back to default if clam not available

        # Configure notebook (tabs)
        style.configure('TNotebook', background=self.colors['bg_light'], borderwidth=0)
        style.configure('TNotebook.Tab',
                       padding=[15, 8],
                       font=('Segoe UI', 9),
                       background=self.colors['bg_dark'],
                       foreground=self.colors['text_light'])
        style.map('TNotebook.Tab',
                 background=[('selected', self.colors['primary'])],
                 foreground=[('selected', self.colors['text_light'])],
                 padding=[('selected', [20, 12])],
                 font=[('selected', ('Segoe UI', 11, 'bold'))],
                 expand=[('selected', [1, 1, 1, 0])])

        # Configure frames
        style.configure('TFrame', background=self.colors['bg_light'])
        style.configure('TLabelframe',
                       background=self.colors['bg_light'],
                       borderwidth=2,
                       relief='solid')
        style.configure('TLabelframe.Label',
                       background=self.colors['bg_light'],
                       foreground=self.colors['primary'],
                       font=('Segoe UI', 10, 'bold'))

        # Configure buttons
        style.configure('TButton',
                       font=('Segoe UI', 9),
                       borderwidth=1,
                       relief='flat',
                       padding=[10, 5])
        style.map('TButton',
                 background=[('active', self.colors['secondary'])],
                 relief=[('pressed', 'sunken')])

        # Accent button style (for primary actions)
        style.configure('Accent.TButton',
                       font=('Segoe UI', 10, 'bold'),
                       background=self.colors['accent'],
                       foreground=self.colors['text_light'],
                       borderwidth=0,
                       padding=[15, 8])
        style.map('Accent.TButton',
                 background=[('active', '#059669')],
                 foreground=[('active', self.colors['text_light'])])

        # Configure labels
        style.configure('TLabel',
                       background=self.colors['bg_light'],
                       font=('Segoe UI', 9))

        style.configure('Header.TLabel',
                       background=self.colors['bg_light'],
                       font=('Segoe UI', 12, 'bold'),
                       foreground=self.colors['primary'])

        # Configure entries
        style.configure('TEntry',
                       fieldbackground='white',
                       borderwidth=1,
                       relief='solid')

        # Configure combobox
        style.configure('TCombobox',
                       fieldbackground='white',
                       background='white',
                       borderwidth=1)

        # Configure treeview
        style.configure('Treeview',
                       background='white',
                       fieldbackground='white',
                       font=('Segoe UI', 9),
                       rowheight=45)  # Increased row height for multi-line messages
        style.configure('Treeview.Heading',
                       font=('Segoe UI', 9, 'bold'),
                       background=self.colors['primary'],
                       foreground=self.colors['text_light'],
                       borderwidth=1,
                       relief='flat')
        style.map('Treeview.Heading',
                 background=[('active', self.colors['secondary'])])

        # Configure checkbuttons
        style.configure('TCheckbutton',
                       background=self.colors['bg_light'],
                       font=('Segoe UI', 9))

    def setup_ui(self):
        """Setup the tabbed user interface"""
        # Main container
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

        # Header banner
        header_frame = tk.Frame(main_frame, bg=self.colors['primary'], height=60)
        header_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 15))
        header_frame.grid_propagate(False)

        title_label = tk.Label(header_frame,
                               text="IHSDadaM",
                               font=('Segoe UI', 18, 'bold'),
                               bg=self.colors['primary'],
                               fg=self.colors['text_light'])
        title_label.pack(side=tk.LEFT, padx=20, pady=10)

        self.version_label = tk.Label(header_frame,
                                text=f"v{__version__}",
                                font=('Segoe UI', 10),
                                bg=self.colors['primary'],
                                fg=self.colors['text_light'],
                                cursor="hand2")
        self.version_label.pack(side=tk.LEFT, pady=10)
        self.version_label.bind("<Button-1>", lambda e: self.check_for_updates(show_current=True))

        # Add tooltip for version label
        self.create_tooltip(self.version_label, "Click to check for updates")

        subtitle_label = tk.Label(header_frame,
                                 text="Warning Extraction & Data Compilation",
                                 font=('Segoe UI', 10, 'italic'),
                                 bg=self.colors['primary'],
                                 fg=self.colors['text_light'])
        subtitle_label.pack(side=tk.RIGHT, padx=20, pady=10)

        # Project path selection (shared across tabs)
        path_frame = ttk.LabelFrame(main_frame, text="IHSDM Project Directory", padding="10")
        path_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        # Project path entry with placeholder
        self.path_entry = ttk.Entry(path_frame, textvariable=self.project_path, width=100)
        self.path_entry.grid(row=0, column=0, padx=(0, 10))

        # Setup placeholder text
        self.placeholder_text = r"C:\Users\aengbring\Downloads\ihsdm_17_0_0_full\IHSDM2021\users\aengbring\Projects_V5\p259"
        self.placeholder_active = True
        self.setup_placeholder()

        ttk.Button(path_frame, text="Browse...", command=self.browse_project).grid(row=0, column=1)

        # Create notebook (tabbed interface)
        self.notebook = ttk.Notebook(main_frame)
        self.notebook.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Tab 1: Warning Message Extractor
        self.warning_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.warning_tab, text="  Warning Extractor  ")
        self.setup_warning_tab()

        # Tab 2: Data Compiler
        self.compiler_tab = ttk.Frame(self.notebook)
        if OPENPYXL_AVAILABLE:
            self.notebook.add(self.compiler_tab, text="  Data Compiler  ")
            self.setup_compiler_tab()
        else:
            self.notebook.add(self.compiler_tab, text="  Data Compiler (Disabled)  ")
            self.setup_compiler_disabled_tab()

        # Tab 3: Appendix Generator
        self.appendix_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.appendix_tab, text="  Appendix Generator  ")
        self.setup_appendix_tab()

        # Tab 4: Visual View
        self.visual_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.visual_tab, text="  Visual View  ")
        self.setup_visual_tab()

        # Tab 5: Calibration Scanner
        self.cmf_tab = ttk.Frame(self.notebook)
        self.notebook.add(self.cmf_tab, text="  Calibration Scanner  ")
        self.setup_cmf_tab()

        # Tab 6: AADT Input
        self.aadt_tab = ttk.Frame(self.notebook)
        if OPENPYXL_AVAILABLE:
            self.notebook.add(self.aadt_tab, text="  AADT Input  ")
            self.setup_aadt_tab()
        else:
            self.notebook.add(self.aadt_tab, text="  AADT Input (Disabled)  ")
            self.setup_aadt_disabled_tab()

        # Configure grid weights
        main_frame.columnconfigure(0, weight=1)
        main_frame.rowconfigure(2, weight=1)

        # Status bar (shared) - enhanced styling
        status_frame = tk.Frame(self.root, bg=self.colors['bg_dark'], height=35)
        status_frame.grid(row=1, column=0, sticky=(tk.W, tk.E))
        status_frame.grid_propagate(False)

        # Left side - status message
        self.status_var = tk.StringVar(value="Ready - Select a project directory above")
        status_bar = tk.Label(status_frame,
                             textvariable=self.status_var,
                             bg=self.colors['bg_dark'],
                             fg=self.colors['text_light'],
                             font=('Segoe UI', 9),
                             anchor=tk.W,
                             padx=15)
        status_bar.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Right side - credits and suggestions
        credits_text = "Created by Adam Engbring • aengbring@hntb.com • Message for suggestions/bugs"
        credits_label = tk.Label(status_frame,
                                text=credits_text,
                                bg=self.colors['bg_dark'],
                                fg=self.colors['text_light'],
                                font=('Segoe UI', 8),
                                anchor=tk.E,
                                padx=15)
        credits_label.pack(side=tk.RIGHT)

    # =========================================================================
    # TOOLTIP HELPER
    # =========================================================================

    def create_tooltip(self, widget, text):
        """Create a tooltip that appears on hover"""
        tooltip = None

        def show_tooltip(event):
            nonlocal tooltip
            x, y, _, _ = widget.bbox("insert")
            x += widget.winfo_rootx() + 25
            y += widget.winfo_rooty() + 25

            tooltip = tk.Toplevel(widget)
            tooltip.wm_overrideredirect(True)
            tooltip.wm_geometry(f"+{x}+{y}")

            label = tk.Label(tooltip, text=text,
                           background="#FFFFE0",
                           relief="solid",
                           borderwidth=1,
                           font=('Segoe UI', 9),
                           padx=5, pady=3)
            label.pack()

        def hide_tooltip(event):
            nonlocal tooltip
            if tooltip:
                tooltip.destroy()
                tooltip = None

        widget.bind("<Enter>", show_tooltip)
        widget.bind("<Leave>", hide_tooltip)

    # =========================================================================
    # WARNING EXTRACTOR TAB
    # =========================================================================

    def setup_warning_tab(self):
        """Setup the warning message extractor tab"""
        tab_frame = ttk.Frame(self.warning_tab, padding="10")
        tab_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.warning_tab.columnconfigure(0, weight=1)
        self.warning_tab.rowconfigure(0, weight=1)

        # Scan button
        scan_frame = ttk.Frame(tab_frame)
        scan_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))

        ttk.Button(scan_frame, text="Scan for Warning Messages",
                  command=self.scan_warnings, style='Accent.TButton').pack(side=tk.LEFT, padx=5)

        # Summary - enhanced visual styling
        summary_frame = tk.Frame(tab_frame, bg='white', relief=tk.RIDGE, borderwidth=2)
        summary_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))

        summary_inner = tk.Frame(summary_frame, bg='white', padx=15, pady=10)
        summary_inner.pack(fill=tk.BOTH, expand=True)

        summary_title = tk.Label(summary_inner,
                                text="Scan Summary",
                                font=('Segoe UI', 10, 'bold'),
                                bg='white',
                                fg=self.colors['primary'])
        summary_title.pack(anchor=tk.W, pady=(0, 5))

        self.summary_text = tk.StringVar(value="No project scanned yet")
        summary_label = tk.Label(summary_inner,
                                textvariable=self.summary_text,
                                font=('Segoe UI', 9),
                                bg='white',
                                fg=self.colors['text_dark'],
                                justify=tk.LEFT)
        summary_label.pack(anchor=tk.W)

        # Filter panel
        filter_frame = ttk.LabelFrame(tab_frame, text="Filters", padding="10")
        filter_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))

        # Alignment Type filter
        ttk.Label(filter_frame, text="Type:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.alignment_type_var = tk.StringVar(value="All")
        alignment_type_combo = ttk.Combobox(filter_frame, textvariable=self.alignment_type_var,
                                           values=['All', 'Highways (h)', 'Intersections (i)', 'Ramp Terminals (r)'],
                                           state='readonly', width=30)
        alignment_type_combo.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        alignment_type_combo.bind('<<ComboboxSelected>>', lambda e: self.apply_filters())

        # Specific Alignment filter
        ttk.Label(filter_frame, text="Alignment:").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.alignment_var = tk.StringVar(value="All")
        self.alignment_combo = ttk.Combobox(filter_frame, textvariable=self.alignment_var, state='readonly', width=30)
        self.alignment_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        self.alignment_combo.bind('<<ComboboxSelected>>', lambda e: self.apply_filters())

        # Message type filter
        ttk.Label(filter_frame, text="Message Type:").grid(row=2, column=0, sticky=tk.W, pady=5)
        self.type_var = tk.StringVar(value="All")
        type_combo = ttk.Combobox(filter_frame, textvariable=self.type_var,
                                  values=['All', 'CRITICAL', 'error', 'warning', 'fault', 'info'],
                                  state='readonly', width=30)
        type_combo.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)
        type_combo.bind('<<ComboboxSelected>>', lambda e: self.apply_filters())

        # Search filter
        ttk.Label(filter_frame, text="Search Text:").grid(row=3, column=0, sticky=tk.W, pady=5)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self.apply_filters())
        search_entry = ttk.Entry(filter_frame, textvariable=self.search_var, width=30)
        search_entry.grid(row=3, column=1, sticky=(tk.W, tk.E), pady=5, padx=5)

        ttk.Button(filter_frame, text="Clear Filters", command=self.clear_filters).grid(row=4, column=0, columnspan=2, pady=10)

        filter_frame.columnconfigure(1, weight=1)

        # Results display
        results_frame = ttk.LabelFrame(tab_frame, text="Messages", padding="10")
        results_frame.grid(row=2, column=1, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0))

        # Treeview for messages
        columns = ('Type', 'Alignment', 'Eval', 'Status', 'Start Sta', 'End Sta', 'Message')
        self.tree = ttk.Treeview(results_frame, columns=columns, show='tree headings', height=20)

        # Define column headings
        self.tree.heading('#0', text='')
        self.tree.heading('Type', text='Type')
        self.tree.heading('Alignment', text='Alignment')
        self.tree.heading('Eval', text='Eval')
        self.tree.heading('Status', text='Status')
        self.tree.heading('Start Sta', text='Start Station')
        self.tree.heading('End Sta', text='End Station')
        self.tree.heading('Message', text='Message')

        # Define column widths
        self.tree.column('#0', width=0, stretch=False)
        self.tree.column('Type', width=40, minwidth=40, stretch=False)
        self.tree.column('Alignment', width=250, minwidth=200)
        self.tree.column('Eval', width=50, minwidth=50, stretch=False)
        self.tree.column('Status', width=90, minwidth=80, stretch=False)
        self.tree.column('Start Sta', width=110, minwidth=100)
        self.tree.column('End Sta', width=110, minwidth=100)
        self.tree.column('Message', width=600, minwidth=400)

        # Scrollbars
        vsb = ttk.Scrollbar(results_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(results_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))

        # Configure styling
        self.tree.tag_configure('critical', background='#ffcccc', foreground='#cc0000')
        self.tree.tag_configure('error', background='#ffe6e6', foreground='#cc0000')
        self.tree.tag_configure('warning', background='#fff4e6', foreground='#cc6600')

        results_frame.columnconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)

        # Action buttons
        button_frame = ttk.Frame(tab_frame)
        button_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(10, 0))

        ttk.Button(button_frame, text="Export to CSV", command=self.export_warnings_csv).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="Copy Selected", command=self.copy_selected).pack(side=tk.LEFT, padx=5)

        # Configure grid weights
        tab_frame.columnconfigure(1, weight=2)
        tab_frame.columnconfigure(2, weight=1)
        tab_frame.rowconfigure(2, weight=1)

    # =========================================================================
    # DATA COMPILER TAB
    # =========================================================================

    def setup_compiler_tab(self):
        """Setup the data compiler tab"""
        tab_frame = ttk.Frame(self.compiler_tab, padding="10")
        tab_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.compiler_tab.columnconfigure(0, weight=1)
        self.compiler_tab.rowconfigure(0, weight=1)

        # Configuration section
        config_frame = ttk.LabelFrame(tab_frame, text="Compiler Configuration", padding="10")
        config_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        # Target file
        ttk.Label(config_frame, text="Target CSV File:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(config_frame, textvariable=self.target_file, width=40).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        ttk.Label(config_frame, text="(default: evaluation.1.diag.csv)", font=('Arial', 8, 'italic')).grid(row=0, column=2, sticky=tk.W, padx=5)

        # Excel output
        ttk.Label(config_frame, text="Excel Output:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(config_frame, textvariable=self.excel_output, width=40).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)
        ttk.Button(config_frame, text="Browse...", command=self.browse_excel_output).grid(row=1, column=2, padx=5, pady=5)

        # Options
        ttk.Checkbutton(config_frame, text="Multi-year extraction (20 years)",
                       variable=self.multi_year_mode).grid(row=2, column=0, columnspan=2, sticky=tk.W, padx=5, pady=5)
        ttk.Checkbutton(config_frame, text="Debug mode (verbose output)",
                       variable=self.debug_mode).grid(row=3, column=0, columnspan=2, sticky=tk.W, padx=5, pady=5)

        config_frame.columnconfigure(1, weight=1)

        # Compile button
        compile_frame = ttk.Frame(tab_frame)
        compile_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        ttk.Button(compile_frame, text="Compile Data to Excel",
                  command=self.run_compiler, style='Accent.TButton').pack(side=tk.LEFT, padx=5)

        # Instructions
        instructions_frame = ttk.LabelFrame(tab_frame, text="Instructions & Information", padding="10")
        instructions_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        instructions_text = scrolledtext.ScrolledText(instructions_frame, wrap=tk.WORD, height=15, font=('Consolas', 9))
        instructions_text.pack(fill=tk.BOTH, expand=True)

        instructions_text.insert('1.0', """IHSDM DATA COMPILER - Instructions

This tool extracts crash prediction data from IHSDM evaluation files and compiles them into Excel.

WHAT IT DOES:
• Processes Highway Segments (h folders)
• Processes Intersections (i folders)
• Processes Ramp Terminals (r folders)
• Applies HSM severity distributions (K, A, B, C crashes)
• Removes duplicates
• Outputs to Excel workbook with separate sheets

FOLDER NAMING:
• h1, h2, h74, etc. → Highway alignments
• i1, i2, i100, etc. → Intersection alignments
• r1, r2, r25, etc. → Ramp terminal alignments
• c1, c2, etc. → Interchanges (can contain h/i/r subfolders)

ALIGNMENT NAMING BEST PRACTICES:
• Mainline freeways → Include "Mainline" in alignment name
• Ramps → Include "Entrance" or "Exit" in alignment name
• Ensure single functional class per alignment

OUTPUT:
• Excel file with 3 sheets: Highway, Intersection, RampTerminal
• Includes crash predictions with severity breakdown
• Single-year mode (default): Extracts one year of predictions
• Multi-year mode: Extracts 20 years of predictions (enable checkbox above)

HSM SEVERITY DISTRIBUTIONS:
• Highway: K=1.46%, A=4.48%, B=24.69%, C=69.17%
• Intersection/Ramp: K=0.26%, A=5.35%, B=27.64%, C=66.75%

STEPS:
1. Select project directory above (shared with Warning Extractor)
2. Set Excel output file path
3. Configure options (multi-year, debug)
4. Click "Compile Data to Excel"
5. Review output in Excel file

NOTE: Consult your state's predictive modeling practices for calibration factors.
Original script by Adam Engbring (aengbring@hntb.com)
""")
        instructions_text.config(state=tk.DISABLED)

        tab_frame.rowconfigure(2, weight=1)

    def setup_compiler_disabled_tab(self):
        """Show message when openpyxl is not available"""
        message_frame = ttk.Frame(self.compiler_tab, padding="50")
        message_frame.pack(expand=True)

        ttk.Label(message_frame, text="Data Compiler Unavailable",
                 font=('Arial', 14, 'bold')).pack(pady=10)
        ttk.Label(message_frame, text="The openpyxl library is required for the Data Compiler feature.",
                 font=('Arial', 10)).pack(pady=5)
        ttk.Label(message_frame, text="Install it using:", font=('Arial', 10)).pack(pady=5)
        ttk.Label(message_frame, text="pip install openpyxl",
                 font=('Consolas', 11, 'bold'), foreground='blue').pack(pady=10)
        ttk.Label(message_frame, text="Then restart this application.", font=('Arial', 10)).pack(pady=5)

    # =========================================================================
    # APPENDIX GENERATOR TAB
    # =========================================================================

    def setup_appendix_tab(self):
        """Setup the appendix generator tab"""
        tab_frame = ttk.Frame(self.appendix_tab, padding="10")
        tab_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.appendix_tab.columnconfigure(0, weight=1)
        self.appendix_tab.rowconfigure(0, weight=1)

        # Left panel - Alignment selection
        left_frame = ttk.Frame(tab_frame)
        left_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(0, 5))

        # Description
        desc_frame = ttk.LabelFrame(left_frame, text="Appendix Generator", padding="10")
        desc_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        desc_text = """Combines evaluation.1.report.pdf files into a single PDF.
Select which alignments to include below."""
        desc_label = tk.Label(desc_frame,
                             text=desc_text,
                             font=('Segoe UI', 10),
                             justify=tk.LEFT,
                             bg=self.colors['bg_light'])
        desc_label.pack(anchor=tk.W)

        # Alignment selection buttons
        btn_frame = ttk.Frame(left_frame)
        btn_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 5))

        ttk.Button(btn_frame, text="Scan for Reports",
                  command=self.scan_appendix_alignments, style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Select All",
                  command=self.appendix_select_all).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Deselect All",
                  command=self.appendix_deselect_all).pack(side=tk.LEFT, padx=5)

        # Alignment selection frame with scrollbar
        alignment_frame = ttk.LabelFrame(left_frame, text="Select Reports to Include", padding="10")
        alignment_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        # Create canvas with scrollbar for alignment checkboxes
        appendix_canvas = tk.Canvas(alignment_frame, bg='white', highlightthickness=0, width=350)
        appendix_scrollbar = ttk.Scrollbar(alignment_frame, orient="vertical", command=appendix_canvas.yview)
        self.appendix_alignment_container = ttk.Frame(appendix_canvas)

        self.appendix_alignment_container.bind(
            "<Configure>",
            lambda e: appendix_canvas.configure(scrollregion=appendix_canvas.bbox("all"))
        )

        appendix_canvas.create_window((0, 0), window=self.appendix_alignment_container, anchor="nw")
        appendix_canvas.configure(yscrollcommand=appendix_scrollbar.set)

        appendix_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        appendix_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Bind mousewheel to scroll
        def _on_mousewheel(event):
            appendix_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        appendix_canvas.bind("<Enter>", lambda e: appendix_canvas.bind_all("<MouseWheel>", _on_mousewheel))
        appendix_canvas.bind("<Leave>", lambda e: appendix_canvas.unbind_all("<MouseWheel>"))

        # Initialize alignment tracking
        self.appendix_alignments = []  # List of discovered PDF files
        self.appendix_selected = {}  # Dict of path: BooleanVar

        left_frame.columnconfigure(0, weight=1)
        left_frame.rowconfigure(2, weight=1)

        # Right panel - Output and log
        right_frame = ttk.Frame(tab_frame)
        right_frame.grid(row=0, column=1, sticky=(tk.W, tk.E, tk.N, tk.S), padx=(5, 0))

        # Output file selection
        output_frame = ttk.LabelFrame(right_frame, text="Output PDF File", padding="10")
        output_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        self.appendix_output = tk.StringVar()
        ttk.Entry(output_frame, textvariable=self.appendix_output, width=50).grid(row=0, column=0, padx=(0, 10))
        ttk.Button(output_frame, text="Browse...", command=self.browse_appendix_output).grid(row=0, column=1)

        output_frame.columnconfigure(0, weight=1)

        # Generate button
        button_frame = ttk.Frame(right_frame)
        button_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        ttk.Button(button_frame, text="Generate Appendix PDF",
                  command=self.generate_appendix, style='Accent.TButton').pack(side=tk.LEFT, padx=5)

        # Status/Log area
        log_frame = ttk.LabelFrame(right_frame, text="Status", padding="10")
        log_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.appendix_log = scrolledtext.ScrolledText(log_frame, wrap=tk.WORD, height=15, font=('Consolas', 9))
        self.appendix_log.pack(fill=tk.BOTH, expand=True)
        self.appendix_log.insert('1.0', "Click 'Scan for Reports' to find available evaluation reports.\n")
        self.appendix_log.config(state=tk.DISABLED)

        right_frame.columnconfigure(0, weight=1)
        right_frame.rowconfigure(2, weight=1)

        # Configure main grid
        tab_frame.columnconfigure(0, weight=1)
        tab_frame.columnconfigure(1, weight=1)
        tab_frame.rowconfigure(0, weight=1)

    # =========================================================================
    # VISUAL VIEW TAB
    # =========================================================================

    def setup_visual_tab(self):
        """Setup the visual view tab"""
        tab_frame = ttk.Frame(self.visual_tab, padding="10")
        tab_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.visual_tab.columnconfigure(0, weight=1)
        self.visual_tab.rowconfigure(0, weight=1)

        # Description
        desc_frame = ttk.LabelFrame(tab_frame, text="Visual View", padding="10")
        desc_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        desc_text = """This tool displays highway alignment data visually from highway.1.xml files.
View horizontal and vertical alignment profiles for any highway alignment in your project."""
        desc_label = tk.Label(desc_frame,
                             text=desc_text,
                             font=('Segoe UI', 10),
                             justify=tk.LEFT,
                             bg=self.colors['bg_light'])
        desc_label.pack(anchor=tk.W)

        # Alignment selector
        selector_frame = ttk.LabelFrame(tab_frame, text="Select Alignment", padding="10")
        selector_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        # Load button on top row
        ttk.Button(selector_frame, text="Load Alignments",
                  command=self.refresh_visual_alignments).grid(row=0, column=0, columnspan=2, sticky=tk.W, padx=5, pady=5)

        # Alignment dropdown on second row
        ttk.Label(selector_frame, text="Highway Alignment:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.visual_alignment_var = tk.StringVar(value="No alignments found")
        self.visual_alignment_combo = ttk.Combobox(selector_frame, textvariable=self.visual_alignment_var,
                                                   state='readonly', width=60)
        self.visual_alignment_combo.grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5, pady=5)

        ttk.Button(selector_frame, text="Display Alignment",
                  command=self.display_alignment, style='Accent.TButton').grid(row=1, column=2, padx=5, pady=5)

        selector_frame.columnconfigure(1, weight=1)

        # Visualization area (placeholder for now)
        viz_frame = ttk.LabelFrame(tab_frame, text="Alignment Visualization", padding="10")
        viz_frame.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))

        self.visual_canvas_frame = tk.Frame(viz_frame, bg='white', relief=tk.SUNKEN, borderwidth=2)
        self.visual_canvas_frame.pack(fill=tk.BOTH, expand=True)

        # Placeholder text
        placeholder_label = tk.Label(self.visual_canvas_frame,
                                     text="Select an alignment and click 'Display Alignment' to view visualization",
                                     font=('Segoe UI', 12),
                                     bg='white',
                                     fg='gray')
        placeholder_label.pack(expand=True)

        tab_frame.columnconfigure(0, weight=1)
        tab_frame.rowconfigure(2, weight=1)

    # =========================================================================
    # SHARED FUNCTIONS
    # =========================================================================

    def setup_placeholder(self):
        """Setup placeholder text for project path entry"""
        # Insert placeholder initially
        self.path_entry.insert(0, self.placeholder_text)
        self.path_entry.config(foreground='gray')

        # Bind focus events
        self.path_entry.bind('<FocusIn>', self.on_path_entry_focus_in)
        self.path_entry.bind('<FocusOut>', self.on_path_entry_focus_out)

    def on_path_entry_focus_in(self, event):
        """Handle focus in - auto-fill with default if placeholder is active"""
        if self.placeholder_active and self.project_path.get() == self.placeholder_text:
            # Keep the text but change to black (auto-fill behavior)
            self.path_entry.config(foreground='black')
            self.placeholder_active = False

    def on_path_entry_focus_out(self, event):
        """Handle focus out - restore placeholder if empty"""
        if not self.project_path.get():
            self.project_path.set(self.placeholder_text)
            self.path_entry.config(foreground='gray')
            self.placeholder_active = True

    def format_station(self, station_str: str) -> str:
        """Format station number with + notation (e.g., 400.00 → 4+00.00, 50.00 → 50.00)"""
        if not station_str or station_str.strip() == '':
            return station_str

        try:
            station_float = float(station_str)

            # If less than 100, no + is needed
            if station_float < 100:
                return station_str

            # Convert to string and split at decimal
            station_parts = str(station_float).split('.')
            integer_part = station_parts[0]
            decimal_part = station_parts[1] if len(station_parts) > 1 else '00'

            # Ensure decimal part has at least 2 digits
            decimal_part = decimal_part.ljust(2, '0')

            # Insert + two digits from the right of integer part
            if len(integer_part) >= 2:
                left_part = integer_part[:-2]
                right_part = integer_part[-2:]
                return f"{left_part}+{right_part}.{decimal_part}"
            else:
                # Edge case: integer part has only 1 digit
                return f"{integer_part}.{decimal_part}"

        except (ValueError, TypeError):
            return station_str

    def wrap_text(self, text: str, width: int = 60) -> str:
        """Wrap text to fit in treeview cell with line breaks"""
        if not text or len(text) <= width:
            return text

        words = text.split()
        lines = []
        current_line = []
        current_length = 0

        for word in words:
            word_length = len(word)
            if current_length + word_length + len(current_line) <= width:
                current_line.append(word)
                current_length += word_length
            else:
                if current_line:
                    lines.append(' '.join(current_line))
                current_line = [word]
                current_length = word_length

        if current_line:
            lines.append(' '.join(current_line))

        return '\n'.join(lines)

    def browse_project(self):
        """Open directory browser dialog for project selection"""
        # Get initial directory, avoiding placeholder text
        initial_dir = self.project_path.get() if not self.placeholder_active else os.path.expanduser("~")

        directory = filedialog.askdirectory(
            title="Select IHSDM Project Directory",
            initialdir=initial_dir
        )
        if directory:
            self.project_path.set(directory)
            self.path_entry.config(foreground='black')
            self.placeholder_active = False
            self.status_var.set(f"Project directory set: {directory}")

    def browse_excel_output(self):
        """Open file browser for Excel output"""
        file_path = filedialog.asksaveasfilename(
            title="Select Excel Output File",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if file_path:
            self.excel_output.set(file_path)

    # =========================================================================
    # WARNING EXTRACTOR FUNCTIONS
    # =========================================================================

    def scan_warnings(self):
        """Scan project for warning messages"""
        project_path = self.project_path.get()

        # Check if placeholder is active
        if self.placeholder_active or not project_path or not os.path.isdir(project_path):
            messagebox.showerror("Error", "Please select a valid project directory")
            return

        self.status_var.set("Scanning project for warnings...")
        self.root.update()

        try:
            self.messages = []
            project_dir = Path(project_path)
            alignment_dirs = []

            # Find direct alignment directories (h*, i*, r*)
            direct_alignments = [d for d in project_dir.iterdir()
                                if d.is_dir() and d.name[0].lower() in ['h', 'i', 'r']]
            alignment_dirs.extend(direct_alignments)

            # Find interchange directories (c*) and their nested alignments
            interchange_dirs = [d for d in project_dir.iterdir()
                               if d.is_dir() and d.name[0].lower() == 'c']

            for interchange_dir in interchange_dirs:
                # Inside each interchange, look for h*, i*, r* subdirectories
                nested_alignments = [d for d in interchange_dir.iterdir()
                                    if d.is_dir() and d.name[0].lower() in ['h', 'i', 'r']]
                alignment_dirs.extend(nested_alignments)

            total_dirs = len(alignment_dirs)
            processed = 0

            for alignment_dir in alignment_dirs:
                processed += 1
                self.status_var.set(f"Scanning {alignment_dir.name}... ({processed}/{total_dirs})")
                self.root.update()

                # Get alignment name
                alignment_name = self.get_alignment_name(alignment_dir)

                # Find all evaluation directories (e*)
                eval_dirs = [d for d in alignment_dir.iterdir()
                           if d.is_dir() and d.name.startswith('e')]

                for eval_dir in eval_dirs:
                    # Look for evaluation result XML files
                    result_files = list(eval_dir.glob("evaluation.*.result.xml"))

                    for result_file in result_files:
                        messages = self.parse_result_file(
                            result_file,
                            alignment_dir.name,
                            alignment_name,
                            eval_dir.name
                        )
                        self.messages.extend(messages)

            # Count critical messages
            critical_count = sum(1 for msg in self.messages if msg.is_critical)

            self.update_summary()
            self.update_alignment_filter()
            self.apply_filters()

            self.status_var.set(f"Scan complete. Found {len(self.messages)} messages ({critical_count} critical).")

            msg_text = f"Found {len(self.messages)} messages across {total_dirs} alignments"
            if critical_count > 0:
                msg_text += f"\n\nWARNING: {critical_count} CRITICAL messages found!\n(Messages containing 'no crash prediction supported')"
                messagebox.showwarning("Scan Complete - Critical Messages Found", msg_text)
            else:
                messagebox.showinfo("Scan Complete", msg_text)

        except Exception as e:
            messagebox.showerror("Error", f"Error scanning project: {str(e)}")
            self.status_var.set("Error during scan")

    def get_alignment_name(self, alignment_dir: Path) -> str:
        """Extract alignment name from XML files"""
        try:
            # Try highway.xml
            highway_file = alignment_dir / "highway.xml"
            if highway_file.exists():
                tree = ET.parse(highway_file)
                root = tree.getroot()
                return root.get('title', alignment_dir.name)

            # Try intersection.xml
            intersection_file = alignment_dir / "intersection.xml"
            if intersection_file.exists():
                tree = ET.parse(intersection_file)
                root = tree.getroot()
                return root.get('title', alignment_dir.name)

            # Try rampterminal.xml
            ramp_file = alignment_dir / "rampterminal.xml"
            if ramp_file.exists():
                tree = ET.parse(ramp_file)
                root = tree.getroot()
                return root.get('title', alignment_dir.name)

        except Exception:
            pass

        return alignment_dir.name

    def parse_result_file(self, file_path: Path, alignment_id: str,
                         alignment_name: str, evaluation: str) -> List[ResultMessage]:
        """Parse evaluation result XML file for ResultMessage elements"""
        messages = []

        try:
            tree = ET.parse(file_path)
            root = tree.getroot()

            for msg_elem in root.iter('ResultMessage'):
                start_sta = msg_elem.get('startSta', '')
                end_sta = msg_elem.get('endSta', '')
                message = msg_elem.get('message', '')
                status = msg_elem.get('ResultMessage.status', 'info')

                alignment_type = alignment_id[0].lower()  # 'h', 'i', or 'r'

                # Check for critical phrase
                is_critical = 'no crash prediction supported' in message.lower()
                if is_critical:
                    status = 'CRITICAL'

                msg = ResultMessage(
                    alignment_type=alignment_type,
                    alignment_id=alignment_id,
                    alignment_name=alignment_name,
                    evaluation=evaluation,
                    start_sta=start_sta,
                    end_sta=end_sta,
                    message=message,
                    status=status,
                    file_path=str(file_path),
                    is_critical=is_critical
                )
                messages.append(msg)

        except Exception as e:
            print(f"Error parsing {file_path}: {e}")

        return messages

    def update_summary(self):
        """Update the summary text"""
        if not self.messages:
            self.summary_text.set("No messages found")
            return

        type_counts = defaultdict(int)
        highway_alignments = set()
        intersection_alignments = set()
        ramp_alignments = set()

        for msg in self.messages:
            type_counts[msg.status] += 1
            alignment_str = f"{msg.alignment_id} - {msg.alignment_name}"

            if msg.alignment_type == 'h':
                highway_alignments.add(alignment_str)
            elif msg.alignment_type == 'i':
                intersection_alignments.add(alignment_str)
            elif msg.alignment_type == 'r':
                ramp_alignments.add(alignment_str)

        summary = f"Total: {len(self.messages)} | "
        summary += f"Highways: {len(highway_alignments)} | "
        summary += f"Intersections: {len(intersection_alignments)} | "
        summary += f"Ramp Terminals: {len(ramp_alignments)} | "

        if type_counts['CRITICAL'] > 0:
            summary += f"CRITICAL: {type_counts['CRITICAL']} | "

        summary += f"Errors: {type_counts['error']} | "
        summary += f"Warnings: {type_counts['warning']} | "
        summary += f"Info: {type_counts['info']}"

        self.summary_text.set(summary)

    def update_alignment_filter(self):
        """Update alignment filter dropdown"""
        highway_alignments = set()
        intersection_alignments = set()
        ramp_alignments = set()

        for msg in self.messages:
            alignment_str = f"{msg.alignment_id} - {msg.alignment_name}"
            if msg.alignment_type == 'h':
                highway_alignments.add(alignment_str)
            elif msg.alignment_type == 'i':
                intersection_alignments.add(alignment_str)
            elif msg.alignment_type == 'r':
                ramp_alignments.add(alignment_str)

        values = ['All']

        if highway_alignments:
            values.append('--- HIGHWAYS ---')
            values.extend(sorted(highway_alignments))

        if intersection_alignments:
            values.append('--- INTERSECTIONS ---')
            values.extend(sorted(intersection_alignments))

        if ramp_alignments:
            values.append('--- RAMP TERMINALS ---')
            values.extend(sorted(ramp_alignments))

        self.alignment_combo['values'] = values
        self.alignment_combo.set('All')

    def apply_filters(self):
        """Apply filters and update treeview"""
        for item in self.tree.get_children():
            self.tree.delete(item)

        alignment_type_filter = self.alignment_type_var.get()
        alignment_filter = self.alignment_var.get()
        type_filter = self.type_var.get()
        search_text = self.search_var.get().lower()

        self.filtered_messages = []

        for msg in self.messages:
            # Alignment type filter
            if alignment_type_filter == 'Highways (h)' and msg.alignment_type != 'h':
                continue
            elif alignment_type_filter == 'Intersections (i)' and msg.alignment_type != 'i':
                continue
            elif alignment_type_filter == 'Ramp Terminals (r)' and msg.alignment_type != 'r':
                continue

            # Specific alignment filter
            if alignment_filter not in ['All', '--- HIGHWAYS ---', '--- INTERSECTIONS ---', '--- RAMP TERMINALS ---']:
                msg_alignment = f"{msg.alignment_id} - {msg.alignment_name}"
                if msg_alignment != alignment_filter:
                    continue

            # Type filter
            if type_filter != 'All' and msg.status != type_filter:
                continue

            # Search filter
            if search_text:
                searchable = f"{msg.message} {msg.alignment_name}".lower()
                if search_text not in searchable:
                    continue

            self.filtered_messages.append(msg)

        # Group by type
        highways = defaultdict(list)
        intersections = defaultdict(list)
        ramps = defaultdict(list)

        for msg in self.filtered_messages:
            key = f"{msg.alignment_id} - {msg.alignment_name}"
            if msg.alignment_type == 'h':
                highways[key].append(msg)
            elif msg.alignment_type == 'i':
                intersections[key].append(msg)
            elif msg.alignment_type == 'r':
                ramps[key].append(msg)

        # Populate tree
        self.populate_tree_section('HIGHWAYS', highways, 'h')
        self.populate_tree_section('INTERSECTIONS', intersections, 'i')
        self.populate_tree_section('RAMP TERMINALS', ramps, 'r')

        critical_count = sum(1 for msg in self.filtered_messages if msg.is_critical)
        status_msg = f"Showing {len(self.filtered_messages)} of {len(self.messages)} messages"
        if critical_count > 0:
            status_msg += f" ({critical_count} CRITICAL)"
        self.status_var.set(status_msg)

    def populate_tree_section(self, section_name, alignments_dict, type_char):
        """Populate a section of the tree view"""
        if not alignments_dict:
            return

        section_parent = self.tree.insert('', 'end', text=section_name,
                                         values=('', f'({len(alignments_dict)} alignments)', '', '', '', '', ''),
                                         open=True)

        for alignment, msgs in sorted(alignments_dict.items()):
            critical_count = sum(1 for m in msgs if m.is_critical)
            error_count = sum(1 for m in msgs if m.status == 'error')
            warning_count = sum(1 for m in msgs if m.status == 'warning')

            label = f"{alignment} ({len(msgs)} msgs"
            if critical_count > 0:
                label += f", {critical_count} CRITICAL"
            if error_count > 0:
                label += f", {error_count} errors"
            if warning_count > 0:
                label += f", {warning_count} warnings"
            label += ")"

            parent = self.tree.insert(section_parent, 'end', text=alignment,
                                    values=(type_char, alignment, '', '', '', '', label))

            for msg in sorted(msgs, key=lambda x: (x.evaluation, x.start_sta)):
                tags = []
                if msg.is_critical:
                    tags.append('critical')
                elif msg.status == 'error':
                    tags.append('error')
                elif msg.status == 'warning':
                    tags.append('warning')

                # Format station numbers with + notation
                formatted_start = self.format_station(msg.start_sta)
                formatted_end = self.format_station(msg.end_sta)

                # Wrap message text for multi-line display (wider wrap for readability)
                wrapped_message = self.wrap_text(msg.message, width=90)

                self.tree.insert(parent, 'end',
                               values=('', '', msg.evaluation, msg.status,
                                      formatted_start, formatted_end, wrapped_message),
                               tags=tags)

    def clear_filters(self):
        """Clear all filters"""
        self.alignment_type_var.set('All')
        self.alignment_var.set('All')
        self.type_var.set('All')
        self.search_var.set('')
        self.apply_filters()

    def export_warnings_csv(self):
        """Export filtered messages to CSV"""
        if not self.filtered_messages:
            messagebox.showwarning("No Data", "No messages to export")
            return

        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            title="Export Messages to CSV"
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.writer(csvfile)
                writer.writerow(['Type', 'Alignment ID', 'Alignment Name', 'Evaluation',
                               'Status', 'Is Critical', 'Start Station', 'End Station', 'Message', 'File Path'])

                for msg in self.filtered_messages:
                    type_name = {'h': 'Highway', 'i': 'Intersection', 'r': 'Ramp Terminal'}[msg.alignment_type]
                    writer.writerow([
                        type_name,
                        msg.alignment_id,
                        msg.alignment_name,
                        msg.evaluation,
                        msg.status,
                        'YES' if msg.is_critical else 'NO',
                        msg.start_sta,
                        msg.end_sta,
                        msg.message,
                        msg.file_path
                    ])

            messagebox.showinfo("Export Complete", f"Exported {len(self.filtered_messages)} messages to CSV")
            self.status_var.set(f"Exported to {file_path}")

        except Exception as e:
            messagebox.showerror("Export Error", f"Error exporting to CSV: {str(e)}")

    def copy_selected(self):
        """Copy selected messages to clipboard"""
        selection = self.tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select messages to copy")
            return

        text_lines = []
        for item in selection:
            values = self.tree.item(item)['values']
            if len(values) >= 6:
                text_lines.append('\t'.join(str(v) for v in values))

        if text_lines:
            self.root.clipboard_clear()
            self.root.clipboard_append('\n'.join(text_lines))
            self.status_var.set(f"Copied {len(text_lines)} messages to clipboard")
        else:
            messagebox.showinfo("No Data", "No message data to copy")

    # =========================================================================
    # DATA COMPILER FUNCTIONS (stub - will add full implementation)
    # =========================================================================

    def run_compiler(self):
        """Run the data compiler"""
        import ihsdm_compiler_core as compiler

        project_path = self.project_path.get()
        excel_path = self.excel_output.get()
        target_file = self.target_file.get()
        multi_year = self.multi_year_mode.get()
        debug = self.debug_mode.get()

        # Validation
        if self.placeholder_active or not project_path or not os.path.isdir(project_path):
            messagebox.showerror("Error", "Please select a valid project directory")
            return

        if not excel_path:
            messagebox.showerror("Error", "Please select an Excel output file")
            return

        self.status_var.set("Compiling data...")
        self.root.update()

        try:
            # Find folders
            parent_folders = compiler.find_folders_with_file(project_path, target_file)

            if not parent_folders:
                messagebox.showerror("Error", f"No folders containing '{target_file}' found in project directory")
                self.status_var.set("Error: No target files found")
                return

            # Process highways
            h_folders = [f for f in parent_folders if 'h' in os.path.basename(f).lower()]
            all_highway_rows = []

            for parent_folder in h_folders:
                for subfolder in os.listdir(parent_folder):
                    subfolder_path = os.path.join(parent_folder, subfolder)
                    if not os.path.isdir(subfolder_path):
                        continue
                    file_path = os.path.join(subfolder_path, target_file)

                    if os.path.isfile(file_path):
                        rows = compiler.extract_highway_segments_from_csv(file_path, start_index=5)
                        for row in rows:
                            if compiler.should_process_highway_row(row):
                                filtered_row = compiler.extract_highway_row_data(row, debug)
                                if filtered_row:
                                    all_highway_rows.append(filtered_row)

            # Remove duplicates and write
            unique_highway_rows, _ = compiler.remove_duplicates(all_highway_rows)
            unique_highway_rows.sort(key=lambda x: (x[0], x[1], x[2]), reverse=False)
            compiler.write_rows_to_excel(unique_highway_rows, excel_path, "Highway")
            compiler.add_header_to_excel(excel_path, "Highway", compiler.HIGHWAY_HEADER)
            compiler.fill_missing_highway_values(excel_path)

            # Process intersections
            i_folders = [f for f in parent_folders if 'i' in os.path.basename(f).lower()]
            all_intersection_rows = []
            first_file = False

            for parent_folder in i_folders:
                for subfolder in os.listdir(parent_folder):
                    subfolder_path = os.path.join(parent_folder, subfolder)
                    if not os.path.isdir(subfolder_path):
                        continue
                    file_path = os.path.join(subfolder_path, target_file)

                    if os.path.isfile(file_path):
                        rows = compiler.extract_by_headers_from_csv(
                            file_path, compiler.INTERSECTION_HEADER[:-1] + ["Fatal and Injury (FI) Crashes"],
                            first_file=not first_file, multi_year=multi_year
                        )
                        if rows:
                            all_intersection_rows.extend(rows)
                            first_file = True

            if all_intersection_rows:
                compiler.write_rows_to_excel(all_intersection_rows, excel_path, "Intersection")
                compiler.fill_missing_intersection_values(excel_path)
                compiler.scrub_duplicate_columns(excel_path, "Intersection")

            # Process ramp terminals
            r_folders = [f for f in parent_folders if 'r' in os.path.basename(f).lower()]
            all_ramp_rows = []
            first_file = False

            for parent_folder in r_folders:
                for subfolder in os.listdir(parent_folder):
                    subfolder_path = os.path.join(parent_folder, subfolder)
                    if not os.path.isdir(subfolder_path):
                        continue
                    file_path = os.path.join(subfolder_path, target_file)

                    if os.path.isfile(file_path):
                        rows = compiler.extract_by_headers_from_csv(
                            file_path, compiler.RAMP_TERMINAL_HEADER[:-1] + ["Fatal and Injury (FI) Crashes"],
                            first_file=not first_file, multi_year=multi_year
                        )
                        if rows:
                            all_ramp_rows.extend(rows)
                            first_file = True

            if all_ramp_rows:
                compiler.write_rows_to_excel(all_ramp_rows, excel_path, "RampTerminal")
                compiler.fill_missing_ramp_terminal_values(excel_path)
                compiler.scrub_duplicate_columns(excel_path, "RampTerminal")

            self.status_var.set("Compilation complete!")

            summary = f"Compilation Complete!\n\n"
            summary += f"Results written to:\n{excel_path}\n\n"
            summary += f"Summary:\n"
            summary += f"  - Highway Segments: {len(unique_highway_rows)} rows\n"
            summary += f"  - Intersections: {len(all_intersection_rows)} rows\n"
            summary += f"  - Ramp Terminals: {len(all_ramp_rows)} rows"

            messagebox.showinfo("Success", summary)

        except Exception as e:
            messagebox.showerror("Error", f"Error during compilation: {str(e)}")
            self.status_var.set("Error during compilation")
            import traceback
            traceback.print_exc()

    # =========================================================================
    # APPENDIX GENERATOR FUNCTIONS
    # =========================================================================

    def browse_appendix_output(self):
        """Open file browser for appendix PDF output"""
        file_path = filedialog.asksaveasfilename(
            title="Select Output PDF File",
            defaultextension=".pdf",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        if file_path:
            self.appendix_output.set(file_path)

    def log_appendix(self, message):
        """Add message to appendix log"""
        self.appendix_log.config(state=tk.NORMAL)
        self.appendix_log.insert(tk.END, message + '\n')
        self.appendix_log.see(tk.END)
        self.appendix_log.config(state=tk.DISABLED)
        self.root.update()

    def scan_appendix_alignments(self):
        """Scan project for all evaluation report PDFs"""
        project_path = self.project_path.get()

        if self.placeholder_active or not project_path or not os.path.isdir(project_path):
            messagebox.showerror("Error", "Please select a valid project directory")
            return

        self.status_var.set("Scanning for evaluation reports...")
        self.appendix_log.config(state=tk.NORMAL)
        self.appendix_log.delete('1.0', tk.END)
        self.appendix_log.config(state=tk.DISABLED)

        self.log_appendix(f"Scanning project: {project_path}")
        self.log_appendix("")

        try:
            project_dir = Path(project_path)
            # Search for all evaluation report PDFs (e1, e2, e3, etc.)
            pdf_files = list(project_dir.glob("**/e*/evaluation.*.report.pdf"))

            if not pdf_files:
                # Fallback to old pattern
                pdf_files = list(project_dir.glob("**/evaluation.*.report.pdf"))

            if not pdf_files:
                self.log_appendix("No evaluation.*.report.pdf files found.")
                messagebox.showwarning("No Files", "No evaluation report PDFs found in the project.")
                return

            # Sort PDFs and organize by alignment
            self.appendix_alignments = sorted(pdf_files)

            self.log_appendix(f"Found {len(self.appendix_alignments)} evaluation reports:")

            # Clear existing checkboxes
            for widget in self.appendix_alignment_container.winfo_children():
                widget.destroy()

            self.appendix_selected = {}

            # Group by type (h=highway, i=intersection, r=ramp)
            highways = []
            intersections = []
            ramps = []
            other = []

            for pdf_file in self.appendix_alignments:
                rel_path = pdf_file.relative_to(project_dir)
                parts = rel_path.parts

                # Get alignment folder name (e.g., h1, i2, r3)
                alignment_folder = None
                for part in parts:
                    if part and part[0].lower() in ['h', 'i', 'r', 'c']:
                        alignment_folder = part
                        break

                # Get alignment name from highway.xml if possible
                alignment_name = self.get_alignment_name_from_pdf_path(pdf_file)
                display_name = f"{'/'.join(parts[:-2])} - {alignment_name}" if alignment_name else str(rel_path.parent.parent)

                entry = {'path': pdf_file, 'display': display_name, 'rel_path': rel_path}

                if alignment_folder and alignment_folder[0].lower() == 'h':
                    highways.append(entry)
                elif alignment_folder and alignment_folder[0].lower() == 'i':
                    intersections.append(entry)
                elif alignment_folder and alignment_folder[0].lower() == 'r':
                    ramps.append(entry)
                else:
                    other.append(entry)

            row = 0

            # Highways section
            if highways:
                header = ttk.Label(self.appendix_alignment_container, text="HIGHWAYS",
                                 font=('Segoe UI', 10, 'bold'), foreground=self.colors['primary'])
                header.grid(row=row, column=0, sticky=tk.W, pady=(5, 2))
                row += 1

                for entry in highways:
                    var = tk.BooleanVar(value=True)
                    self.appendix_selected[str(entry['path'])] = var
                    cb = ttk.Checkbutton(self.appendix_alignment_container,
                                        text=entry['display'],
                                        variable=var)
                    cb.grid(row=row, column=0, sticky=tk.W, padx=10)
                    row += 1
                    self.log_appendix(f"  [H] {entry['display']}")

            # Intersections section
            if intersections:
                header = ttk.Label(self.appendix_alignment_container, text="INTERSECTIONS",
                                 font=('Segoe UI', 10, 'bold'), foreground=self.colors['primary'])
                header.grid(row=row, column=0, sticky=tk.W, pady=(10, 2))
                row += 1

                for entry in intersections:
                    var = tk.BooleanVar(value=True)
                    self.appendix_selected[str(entry['path'])] = var
                    cb = ttk.Checkbutton(self.appendix_alignment_container,
                                        text=entry['display'],
                                        variable=var)
                    cb.grid(row=row, column=0, sticky=tk.W, padx=10)
                    row += 1
                    self.log_appendix(f"  [I] {entry['display']}")

            # Ramps section
            if ramps:
                header = ttk.Label(self.appendix_alignment_container, text="RAMP TERMINALS",
                                 font=('Segoe UI', 10, 'bold'), foreground=self.colors['primary'])
                header.grid(row=row, column=0, sticky=tk.W, pady=(10, 2))
                row += 1

                for entry in ramps:
                    var = tk.BooleanVar(value=True)
                    self.appendix_selected[str(entry['path'])] = var
                    cb = ttk.Checkbutton(self.appendix_alignment_container,
                                        text=entry['display'],
                                        variable=var)
                    cb.grid(row=row, column=0, sticky=tk.W, padx=10)
                    row += 1
                    self.log_appendix(f"  [R] {entry['display']}")

            # Other section
            if other:
                header = ttk.Label(self.appendix_alignment_container, text="OTHER",
                                 font=('Segoe UI', 10, 'bold'), foreground=self.colors['primary'])
                header.grid(row=row, column=0, sticky=tk.W, pady=(10, 2))
                row += 1

                for entry in other:
                    var = tk.BooleanVar(value=True)
                    self.appendix_selected[str(entry['path'])] = var
                    cb = ttk.Checkbutton(self.appendix_alignment_container,
                                        text=entry['display'],
                                        variable=var)
                    cb.grid(row=row, column=0, sticky=tk.W, padx=10)
                    row += 1
                    self.log_appendix(f"  [?] {entry['display']}")

            self.log_appendix("")
            self.log_appendix(f"All {len(self.appendix_alignments)} reports selected by default.")
            self.status_var.set(f"Found {len(self.appendix_alignments)} evaluation reports")

        except Exception as e:
            self.log_appendix(f"Error: {str(e)}")
            messagebox.showerror("Error", f"Error scanning for reports: {str(e)}")

    def get_alignment_name_from_pdf_path(self, pdf_path):
        """Try to get alignment name from highway.xml near the PDF"""
        try:
            # PDF is in e*/evaluation.1.report.pdf, alignment is parent of e*
            alignment_dir = pdf_path.parent.parent
            # Look for highway.1.xml or similar
            for xml_file in alignment_dir.glob("*.1.xml"):
                if 'highway' in xml_file.name or 'intersection' in xml_file.name or 'ramp' in xml_file.name:
                    return self.get_alignment_name(alignment_dir)
            return self.get_alignment_name(alignment_dir)
        except:
            return None

    def appendix_select_all(self):
        """Select all alignments for appendix"""
        for var in self.appendix_selected.values():
            var.set(True)

    def appendix_deselect_all(self):
        """Deselect all alignments for appendix"""
        for var in self.appendix_selected.values():
            var.set(False)

    def generate_appendix(self):
        """Generate combined PDF from selected evaluation.1.report.pdf files"""
        project_path = self.project_path.get()
        output_path = self.appendix_output.get()

        # Validation
        if self.placeholder_active or not project_path or not os.path.isdir(project_path):
            messagebox.showerror("Error", "Please select a valid project directory")
            return

        if not output_path:
            messagebox.showerror("Error", "Please select an output PDF file")
            return

        # Check for PyPDF2
        try:
            from PyPDF2 import PdfMerger
        except ImportError:
            msg = "PyPDF2 library is required for PDF merging.\n\nInstall with: pip install PyPDF2"
            messagebox.showerror("Missing Dependency", msg)
            self.log_appendix("ERROR: PyPDF2 not installed. Run: pip install PyPDF2")
            return

        self.status_var.set("Generating appendix...")
        self.appendix_log.config(state=tk.NORMAL)
        self.appendix_log.delete('1.0', tk.END)
        self.appendix_log.config(state=tk.DISABLED)

        self.log_appendix("Starting appendix generation...")
        self.log_appendix(f"Project: {project_path}")
        self.log_appendix(f"Output: {output_path}")
        self.log_appendix("")

        try:
            project_dir = Path(project_path)

            # Get selected PDFs from checkbox selection, or scan if not done yet
            if self.appendix_selected:
                pdf_files = [Path(path) for path, var in self.appendix_selected.items() if var.get()]
            else:
                # Fallback: scan for all PDFs if user didn't scan first (all e* folders)
                pdf_files = list(project_dir.glob("**/e*/evaluation.*.report.pdf"))
                if not pdf_files:
                    pdf_files = list(project_dir.glob("**/evaluation.*.report.pdf"))

            if not pdf_files:
                self.log_appendix("No evaluation reports selected or found.")
                messagebox.showwarning("No Files", "No evaluation reports selected. Please scan and select reports first.")
                self.status_var.set("No PDF files selected")
                return

            self.log_appendix(f"Selected {len(pdf_files)} PDF files to merge:")
            for pdf_file in sorted(pdf_files):
                try:
                    rel_path = pdf_file.relative_to(project_dir)
                    self.log_appendix(f"  - {rel_path}")
                except ValueError:
                    self.log_appendix(f"  - {pdf_file}")

            self.log_appendix("")
            self.log_appendix("Merging PDFs...")

            # Merge PDFs
            merger = PdfMerger()
            added_count = 0
            for pdf_file in sorted(pdf_files):
                try:
                    merger.append(str(pdf_file))
                    try:
                        rel_path = pdf_file.relative_to(project_dir)
                        self.log_appendix(f"  Added: {rel_path}")
                    except ValueError:
                        self.log_appendix(f"  Added: {pdf_file.name}")
                    added_count += 1
                except Exception as e:
                    self.log_appendix(f"  WARNING: Could not add {pdf_file.name}: {str(e)}")

            # Write output
            self.log_appendix("")
            self.log_appendix("Writing combined PDF...")
            merger.write(output_path)
            merger.close()

            self.log_appendix("")
            self.log_appendix(f"SUCCESS! Combined PDF created: {output_path}")
            self.status_var.set(f"Appendix generated: {added_count} PDFs combined")

            messagebox.showinfo("Success", f"Appendix PDF generated successfully!\n\nCombined {added_count} evaluation reports into:\n{output_path}")

        except Exception as e:
            self.log_appendix("")
            self.log_appendix(f"ERROR: {str(e)}")
            messagebox.showerror("Error", f"Error generating appendix: {str(e)}")
            self.status_var.set("Error generating appendix")
            import traceback
            traceback.print_exc()

    # =========================================================================
    # VISUAL VIEW FUNCTIONS
    # =========================================================================

    def refresh_visual_alignments(self):
        """Refresh the list of available highway alignments"""
        project_path = self.project_path.get()

        if self.placeholder_active or not project_path or not os.path.isdir(project_path):
            messagebox.showerror("Error", "Please select a valid project directory")
            return

        try:
            project_dir = Path(project_path)
            highway_dirs = []

            # Find direct highway directories (h*)
            direct_highways = [d for d in project_dir.iterdir()
                              if d.is_dir() and d.name[0].lower() == 'h']
            highway_dirs.extend(direct_highways)

            # Find highways inside interchanges (c*/h*)
            interchange_dirs = [d for d in project_dir.iterdir()
                               if d.is_dir() and d.name[0].lower() == 'c']

            for interchange_dir in interchange_dirs:
                nested_highways = [d for d in interchange_dir.iterdir()
                                  if d.is_dir() and d.name[0].lower() == 'h']
                highway_dirs.extend(nested_highways)

            if not highway_dirs:
                self.visual_alignment_var.set("No highway alignments found")
                self.visual_alignment_combo['values'] = []
                messagebox.showinfo("No Alignments", "No highway alignments found in project directory.")
                return

            # Get alignment names
            alignment_options = []
            for highway_dir in sorted(highway_dirs, key=lambda x: x.name):
                alignment_name = self.get_alignment_name(highway_dir)
                alignment_options.append(f"{highway_dir.name} - {alignment_name}")

            self.visual_alignment_combo['values'] = alignment_options
            if alignment_options:
                self.visual_alignment_var.set(alignment_options[0])

            self.status_var.set(f"Found {len(alignment_options)} highway alignments")

        except Exception as e:
            messagebox.showerror("Error", f"Error finding alignments: {str(e)}")

    def display_alignment(self):
        """Display the selected highway alignment visualization with rich data"""
        project_path = self.project_path.get()
        selected = self.visual_alignment_var.get()

        if self.placeholder_active or not project_path or not os.path.isdir(project_path):
            messagebox.showerror("Error", "Please select a valid project directory")
            return

        if not selected or selected == "No alignments found":
            messagebox.showwarning("No Selection", "Please select a highway alignment first")
            return

        # Extract alignment ID from selection (format: "h1 - Alignment Name")
        alignment_id = selected.split(' - ')[0]

        try:
            # Find the highway.1.xml file
            project_dir = Path(project_path)

            # Check direct highway directories
            highway_dir = project_dir / alignment_id
            if not highway_dir.exists():
                # Check inside interchanges
                for interchange_dir in project_dir.glob('c*'):
                    potential_dir = interchange_dir / alignment_id
                    if potential_dir.exists():
                        highway_dir = potential_dir
                        break

            highway_xml = highway_dir / "highway.1.xml"

            if not highway_xml.exists():
                messagebox.showerror("Error", f"Could not find highway.1.xml in {highway_dir}")
                return

            # Parse highway.1.xml
            tree = ET.parse(highway_xml)
            root = tree.getroot()

            # Find the Roadway element (actual data container)
            roadway = root.find('.//{http://www.ihsdm.org/schema/Highway-1.0}Roadway')
            if roadway is None:
                # Try without namespace
                roadway = root.find('.//Roadway')
            if roadway is None:
                messagebox.showerror("Error", "Could not find Roadway element in highway.1.xml")
                return

            # Extract alignment data from Roadway element
            min_station = roadway.get('minStation', '0.0')
            max_station = roadway.get('maxStation', '0.0')
            heading_sta = roadway.get('headingSta', '0.0')
            heading_angle = roadway.get('headingAngle', '0.0')

            # Clear canvas
            for widget in self.visual_canvas_frame.winfo_children():
                widget.destroy()

            # Create scrollable canvas
            canvas_container = tk.Frame(self.visual_canvas_frame, bg='lightblue')
            canvas_container.pack(fill=tk.BOTH, expand=True)

            scrollbar = ttk.Scrollbar(canvas_container, orient=tk.VERTICAL)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

            canvas = tk.Canvas(canvas_container, bg='white', yscrollcommand=scrollbar.set, highlightthickness=0)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.config(command=canvas.yview)

            # Draw a test element to verify canvas is working
            print(f"DEBUG: Canvas created, drawing test elements...")

            # Parse alignment data
            min_sta_float = float(min_station)
            max_sta_float = float(max_station)
            length = max_sta_float - min_sta_float

            # Get canvas dimensions (wait for canvas to render)
            canvas.update()
            canvas_width = max(canvas.winfo_width(), 800)
            canvas_height = max(canvas.winfo_height(), 600)

            # Parse all data from XML (use roadway element, not root)
            lanes_data = self.parse_lanes(roadway, min_sta_float, max_sta_float)
            shoulders_data = self.parse_shoulders(roadway, min_sta_float, max_sta_float)
            ramps_data = self.parse_ramps(roadway)
            curves_data = self.parse_horizontal_curves(roadway, min_sta_float, max_sta_float)
            traffic_data = self.parse_traffic(roadway, min_sta_float, max_sta_float)
            median_data = self.parse_median(roadway, min_sta_float, max_sta_float)
            speed_data = self.parse_speed(roadway, min_sta_float, max_sta_float)
            func_class_data = self.parse_functional_class(roadway, min_sta_float, max_sta_float)

            # Debug: Print what we found
            print(f"DEBUG: Parsed data counts:")
            print(f"  Lanes: {len(lanes_data)}")
            print(f"  Shoulders: {len(shoulders_data)}")
            print(f"  Ramps: {len(ramps_data)}")
            print(f"  Curves: {len(curves_data)}")
            print(f"  Traffic: {len(traffic_data)}")
            print(f"  Median: {len(median_data)}")
            print(f"  Speed: {len(speed_data)}")
            print(f"  Functional Class: {len(func_class_data)}")

            # Calculate total height needed
            y_offset = 0
            margin_x = 40
            panel_width = canvas_width - 2 * margin_x

            # ===== HEADER SECTION =====
            header_height = 100
            print(f"DEBUG: Drawing header at canvas size: {canvas_width} x {canvas_height}")
            canvas.create_rectangle(0, y_offset, canvas_width, y_offset + header_height,
                                  fill='#f0f0f0', outline='#cccccc', width=2)
            canvas.create_text(20, y_offset + 15, anchor='nw', text=f"Highway Alignment: {selected}",
                             font=('Segoe UI', 14, 'bold'), fill='#1e3a8a')
            canvas.create_text(20, y_offset + 45, anchor='nw',
                             text=f"Station Range: {self.format_station(min_station)} to {self.format_station(max_station)} ({length:.2f} ft)",
                             font=('Segoe UI', 10), fill='black')
            canvas.create_text(20, y_offset + 65, anchor='nw',
                             text=f"Heading: {heading_angle}° at Sta {self.format_station(heading_sta)}",
                             font=('Segoe UI', 10), fill='black')
            y_offset += header_height + 20
            print(f"DEBUG: Header drawn, y_offset now: {y_offset}")

            # ===== LANE CONFIGURATION PANEL =====
            if lanes_data or shoulders_data:
                y_offset = self.draw_lane_panel(canvas, lanes_data, shoulders_data, margin_x, panel_width, y_offset,
                                              min_sta_float, max_sta_float)
                y_offset += 30

            # ===== RAMP LOCATIONS PANEL =====
            if ramps_data:
                y_offset = self.draw_ramp_panel(canvas, ramps_data, margin_x, panel_width, y_offset,
                                              min_sta_float, max_sta_float)
                y_offset += 30

            # ===== HORIZONTAL ALIGNMENT PANEL =====
            if curves_data:
                y_offset = self.draw_curves_panel(canvas, curves_data, margin_x, panel_width, y_offset,
                                                min_sta_float, max_sta_float)
                y_offset += 30

            # ===== TRAFFIC VOLUME PANEL =====
            if traffic_data:
                y_offset = self.draw_traffic_panel(canvas, traffic_data, margin_x, panel_width, y_offset,
                                                  min_sta_float, max_sta_float)
                y_offset += 30

            # ===== MEDIAN PANEL =====
            if median_data:
                y_offset = self.draw_median_panel(canvas, median_data, margin_x, panel_width, y_offset,
                                                min_sta_float, max_sta_float)
                y_offset += 30

            # ===== POSTED SPEED PANEL =====
            if speed_data:
                y_offset = self.draw_speed_panel(canvas, speed_data, margin_x, panel_width, y_offset,
                                               min_sta_float, max_sta_float)
                y_offset += 30

            # ===== FUNCTIONAL CLASS PANEL =====
            if func_class_data:
                y_offset = self.draw_func_class_panel(canvas, func_class_data, margin_x, panel_width, y_offset,
                                                    min_sta_float, max_sta_float)
                y_offset += 20

            # If no data panels were shown, display a message
            if y_offset <= header_height + 20:
                canvas.create_text(canvas_width // 2, header_height + 100,
                                 text="No detailed alignment data found in highway.1.xml",
                                 font=('Segoe UI', 12), fill='#999999', anchor='center')
                canvas.create_text(canvas_width // 2, header_height + 130,
                                 text="(Looking for LaneNS, RampConnector, HorizontalAlignment, Traffic, etc.)",
                                 font=('Segoe UI', 10), fill='#cccccc', anchor='center')
                y_offset = header_height + 200

            # Configure scroll region
            canvas.config(scrollregion=(0, 0, canvas_width, y_offset))

            # Bind mousewheel to scroll only when mouse is over the canvas
            def _on_mousewheel(event):
                canvas.yview_scroll(int(-1*(event.delta/120)), "units")

            def _bind_mousewheel(event):
                canvas.bind_all("<MouseWheel>", _on_mousewheel)

            def _unbind_mousewheel(event):
                canvas.unbind_all("<MouseWheel>")

            canvas.bind("<Enter>", _bind_mousewheel)
            canvas.bind("<Leave>", _unbind_mousewheel)

            self.status_var.set(f"Displaying alignment: {selected}")

        except Exception as e:
            messagebox.showerror("Error", f"Error displaying alignment: {str(e)}")
            import traceback
            traceback.print_exc()

    # ===== DATA PARSING FUNCTIONS =====

    def find_elements(self, parent, tag_name):
        """Find elements by tag name, handling XML namespaces"""
        # Try with IHSDM namespace
        ns = '{http://www.ihsdm.org/schema/Highway-1.0}'
        elements = parent.findall(f'.//{ns}{tag_name}')
        if elements:
            return elements

        # Try without namespace
        elements = parent.findall(f'.//{tag_name}')
        if elements:
            return elements

        # Try as direct children with namespace in tag
        return [child for child in parent if child.tag.endswith(tag_name)]

    def parse_lanes(self, root, min_sta, max_sta):
        """Parse lane configuration data from LaneNS elements"""
        lanes = []
        for lane in self.find_elements(root, 'LaneNS'):
            try:
                # IHSDM uses startStation/endStation, startWidth/endWidth
                start_sta = float(lane.get('startStation', min_sta))
                end_sta = float(lane.get('endStation', max_sta))
                start_width = float(lane.get('startWidth', '12.0'))
                end_width = float(lane.get('endWidth', start_width))
                avg_width = (start_width + end_width) / 2
                priority = int(lane.get('priority', '10'))
                side = lane.get('sideOfRoad', 'both')

                lanes.append({
                    'begin': start_sta,
                    'end': end_sta,
                    'side': side,
                    'priority': priority,
                    'lane_type': lane.get('laneType', 'thru'),
                    'width': avg_width
                })
            except (ValueError, AttributeError):
                continue
        # Sort by station first, then by side (left < both < right), then by priority
        return sorted(lanes, key=lambda x: (x['begin'], x['side'], x['priority']))

    def parse_shoulders(self, root, min_sta, max_sta):
        """Parse shoulder data from ShoulderSection elements"""
        shoulders = []
        for shoulder in self.find_elements(root, 'ShoulderSection'):
            try:
                start_sta = float(shoulder.get('startStation', min_sta))
                end_sta = float(shoulder.get('endStation', max_sta))
                start_width = float(shoulder.get('startWidth', '0.0'))
                end_width = float(shoulder.get('endWidth', start_width))
                avg_width = (start_width + end_width) / 2
                priority = int(shoulder.get('priority', '100'))
                side = shoulder.get('sideOfRoad', 'right')
                inside_outside = shoulder.get('insideOutsideOfRoadNB', 'outside')
                material = shoulder.get('material', 'paved')

                shoulders.append({
                    'begin': start_sta,
                    'end': end_sta,
                    'side': side,
                    'priority': priority,
                    'width': avg_width,
                    'position': inside_outside,  # inside or outside
                    'material': material
                })
            except (ValueError, AttributeError):
                continue
        return sorted(shoulders, key=lambda x: (x['begin'], x['side'], x['priority']))

    def parse_ramps(self, root):
        """Parse ramp connector data"""
        ramps = []
        # Look for RampConnector elements (for interchanges)
        for ramp in self.find_elements(root, 'RampConnector'):
            try:
                ramps.append({
                    'station': float(ramp.get('station', '0.0')),
                    'name': ramp.get('name', 'Ramp'),
                    'ramp_type': ramp.get('type', 'entrance')
                })
            except (ValueError, AttributeError):
                continue
        return sorted(ramps, key=lambda x: x['station'])

    def parse_horizontal_curves(self, root, min_sta, max_sta):
        """Parse horizontal alignment curves and tangents"""
        curves = []
        # IHSDM uses HorizontalElements (not HorizontalAlignment)
        h_elements_list = self.find_elements(root, 'HorizontalElements')
        if h_elements_list:
            h_elements = h_elements_list[0]

            # Look for HTangent elements
            for tangent in self.find_elements(h_elements, 'HTangent'):
                try:
                    curves.append({
                        'type': 'tangent',
                        'begin': float(tangent.get('startStation', '0.0')),
                        'end': float(tangent.get('endStation', '0.0'))
                    })
                except (ValueError, AttributeError):
                    continue

            # Look for HSimpleCurve elements
            for curve in self.find_elements(h_elements, 'HSimpleCurve'):
                try:
                    curves.append({
                        'type': 'curve',
                        'begin': float(curve.get('startStation', '0.0')),
                        'end': float(curve.get('endStation', '0.0')),
                        'radius': float(curve.get('radius', '0.0')),
                        'direction': curve.get('curveDirection', 'left')
                    })
                except (ValueError, AttributeError):
                    continue

            # Look for HSpiralCurve elements
            for spiral in self.find_elements(h_elements, 'HSpiralCurve'):
                try:
                    curves.append({
                        'type': 'spiral',
                        'begin': float(spiral.get('startStation', '0.0')),
                        'end': float(spiral.get('endStation', '0.0')),
                        'radius': float(spiral.get('radius', '0.0'))
                    })
                except (ValueError, AttributeError):
                    continue

        return sorted(curves, key=lambda x: x['begin'])

    def parse_traffic(self, root, min_sta, max_sta):
        """Parse traffic volume data"""
        traffic = []
        for aadt in self.find_elements(root, 'AnnualAveDailyTraffic'):
            try:
                # IHSDM uses startStation/endStation and adtRate
                traffic.append({
                    'begin': float(aadt.get('startStation', min_sta)),
                    'end': float(aadt.get('endStation', max_sta)),
                    'volume': int(float(aadt.get('adtRate', '0')))
                })
            except (ValueError, AttributeError):
                continue
        return sorted(traffic, key=lambda x: x['begin'])

    def parse_median(self, root, min_sta, max_sta):
        """Parse median data"""
        median = []
        for med in self.find_elements(root, 'Median'):
            try:
                # IHSDM uses startStation/endStation
                median.append({
                    'begin': float(med.get('startStation', min_sta)),
                    'end': float(med.get('endStation', max_sta)),
                    'width': float(med.get('width', '0.0')),
                    'median_type': med.get('medianType', 'none')
                })
            except (ValueError, AttributeError):
                continue
        return sorted(median, key=lambda x: x['begin'])

    def parse_speed(self, root, min_sta, max_sta):
        """Parse posted speed data"""
        speeds = []
        for speed in self.find_elements(root, 'PostedSpeed'):
            try:
                # IHSDM uses startStation/endStation and speedLimit
                speeds.append({
                    'begin': float(speed.get('startStation', min_sta)),
                    'end': float(speed.get('endStation', max_sta)),
                    'speed': int(float(speed.get('speedLimit', '0')))
                })
            except (ValueError, AttributeError):
                continue
        return sorted(speeds, key=lambda x: x['begin'])

    def parse_functional_class(self, root, min_sta, max_sta):
        """Parse functional class data"""
        func_classes = []
        for fc in self.find_elements(root, 'FunctionalClass'):
            try:
                # IHSDM uses startStation/endStation and funcClass
                func_classes.append({
                    'begin': float(fc.get('startStation', min_sta)),
                    'end': float(fc.get('endStation', max_sta)),
                    'class_type': fc.get('funcClass', 'unknown')
                })
            except (ValueError, AttributeError):
                continue
        return sorted(func_classes, key=lambda x: x['begin'])

    # ===== DRAWING FUNCTIONS =====

    def draw_lane_panel(self, canvas, lanes_data, shoulders_data, margin_x, panel_width, y_start, min_sta, max_sta):
        """Draw lane configuration as roadway plan view with lanes building from centerline

        Layout (based on user sketch):
        - Centerline (red) runs horizontally
        - LEFT side lanes appear ABOVE centerline (opposite direction)
        - RIGHT side lanes appear BELOW centerline (forward direction, increasing stationing)
        - BOTH side lanes appear mirrored on BOTH sides of centerline
        - Inside shoulders = closest to centerline/median
        - Outside shoulders = on outer edge of road
        - Priority determines stacking: P10 closest to CL, P20 next, etc.
        """
        canvas.create_text(margin_x, y_start, anchor='nw', text="Lane Configuration (Plan View)",
                         font=('Segoe UI', 11, 'bold'), fill='#1e3a8a')
        y_start += 25

        # Count max priority levels on each side for lanes
        right_priorities = set()
        left_priorities = set()

        for lane in lanes_data:
            priority_level = lane['priority'] // 10
            if lane['side'] == 'right':
                right_priorities.add(priority_level)
            elif lane['side'] == 'left':
                left_priorities.add(priority_level)
            else:  # both - appears on BOTH sides
                right_priorities.add(priority_level)
                left_priorities.add(priority_level)

        # Count inside and outside shoulders per side
        right_inside_shoulders = 0
        right_outside_shoulders = 0
        left_inside_shoulders = 0
        left_outside_shoulders = 0

        for shoulder in shoulders_data:
            position = shoulder.get('position', 'outside')
            side = shoulder['side']
            if side == 'right' or side == 'both':
                if position == 'inside':
                    right_inside_shoulders = max(right_inside_shoulders, 1)
                else:
                    right_outside_shoulders = max(right_outside_shoulders, 1)
            if side == 'left' or side == 'both':
                if position == 'inside':
                    left_inside_shoulders = max(left_inside_shoulders, 1)
                else:
                    left_outside_shoulders = max(left_outside_shoulders, 1)

        max_right_lanes = max(right_priorities) + 1 if right_priorities else 0
        max_left_lanes = max(left_priorities) + 1 if left_priorities else 0

        lane_height = 18
        shoulder_height = 12

        # Panel height calculation:
        # Above CL: left outside shoulder + left lanes + left inside shoulder
        # Below CL: right inside shoulder + right lanes + right outside shoulder
        total_above = (left_outside_shoulders * shoulder_height +
                      max_left_lanes * lane_height +
                      left_inside_shoulders * shoulder_height)
        total_below = (right_inside_shoulders * shoulder_height +
                      max_right_lanes * lane_height +
                      right_outside_shoulders * shoulder_height)
        panel_height = max(120, total_above + total_below + 50)

        canvas.create_rectangle(margin_x, y_start, margin_x + panel_width, y_start + panel_height,
                              fill='#e5e7eb', outline='#666666', width=2)

        # Centerline position - account for inside shoulders
        centerline_y = y_start + left_outside_shoulders * shoulder_height + max_left_lanes * lane_height + left_inside_shoulders * shoulder_height + 25

        # Draw centerline (red dashed)
        canvas.create_line(margin_x + 10, centerline_y, margin_x + panel_width - 10, centerline_y,
                         fill='#dc2626', width=3, dash=(10, 5))
        canvas.create_text(margin_x + 15, centerline_y, text="CL", font=('Consolas', 8, 'bold'),
                         fill='#dc2626', anchor='w')

        # Helper to draw a bar (lane or shoulder)
        def draw_bar(item, is_shoulder=False, shoulder_position='outside'):
            x_start = margin_x + 10 + (item['begin'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)
            x_end = margin_x + 10 + (item['end'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)

            priority_level = item['priority'] // 10
            bar_height = shoulder_height if is_shoulder else lane_height

            if is_shoulder:
                color = '#22c55e'  # Green for shoulders
                pos_label = "In" if shoulder_position == 'inside' else "Out"
                label = f"{pos_label} Shldr"
            else:
                color = self.get_lane_color(item.get('lane_type', 'thru'))
                label = f"P{item['priority']} {item.get('lane_type', '')}"

            side = item['side']

            # Determine which sides to draw on
            sides_to_draw = []
            if side == 'right':
                sides_to_draw = ['right']
            elif side == 'left':
                sides_to_draw = ['left']
            else:  # both
                sides_to_draw = ['left', 'right']

            for draw_side in sides_to_draw:
                if draw_side == 'right':
                    # Right side: BELOW centerline
                    if is_shoulder:
                        if shoulder_position == 'inside':
                            # Inside shoulder: between centerline and lanes
                            y_top = centerline_y
                            y_bottom = y_top + shoulder_height
                        else:
                            # Outside shoulder: beyond all lanes
                            y_top = centerline_y + right_inside_shoulders * shoulder_height + max_right_lanes * lane_height
                            y_bottom = y_top + shoulder_height
                    else:
                        # Lanes stack from inside shoulder outward
                        y_top = centerline_y + right_inside_shoulders * shoulder_height + priority_level * lane_height
                        y_bottom = y_top + lane_height
                else:  # left
                    # Left side: ABOVE centerline
                    if is_shoulder:
                        if shoulder_position == 'inside':
                            # Inside shoulder: between centerline and lanes
                            y_bottom = centerline_y
                            y_top = y_bottom - shoulder_height
                        else:
                            # Outside shoulder: beyond all lanes
                            y_bottom = centerline_y - left_inside_shoulders * shoulder_height - max_left_lanes * lane_height
                            y_top = y_bottom - shoulder_height
                    else:
                        # Lanes stack from inside shoulder outward
                        y_bottom = centerline_y - left_inside_shoulders * shoulder_height - priority_level * lane_height
                        y_top = y_bottom - lane_height

                # Draw the bar
                canvas.create_rectangle(x_start, y_top, x_end, y_bottom,
                                      fill=color, outline='#1e3a8a', width=1)

                # Add label on the bar
                bar_width = x_end - x_start
                bar_mid_y = (y_top + y_bottom) / 2

                if bar_width > 100:
                    canvas.create_text((x_start + x_end) / 2, bar_mid_y,
                                     text=label, font=('Consolas', 7, 'bold'), fill='white')
                elif bar_width > 50:
                    short_label = f"P{item['priority']}" if not is_shoulder else pos_label
                    canvas.create_text((x_start + x_end) / 2, bar_mid_y,
                                     text=short_label, font=('Consolas', 7), fill='white')

        # Draw inside shoulders first (closest to centerline)
        for shoulder in shoulders_data:
            if shoulder.get('position', 'outside') == 'inside':
                draw_bar(shoulder, is_shoulder=True, shoulder_position='inside')

        # Draw lanes
        for lane in lanes_data:
            draw_bar(lane, is_shoulder=False)

        # Draw outside shoulders last (furthest from centerline)
        for shoulder in shoulders_data:
            if shoulder.get('position', 'outside') == 'outside':
                draw_bar(shoulder, is_shoulder=True, shoulder_position='outside')

        # Add station markers at bottom
        marker_y = y_start + panel_height - 15
        num_markers = 5
        for i in range(num_markers):
            x = margin_x + 10 + (panel_width - 20) * i / (num_markers - 1)
            sta = min_sta + (max_sta - min_sta) * i / (num_markers - 1)
            canvas.create_line(x, marker_y - 5, x, marker_y + 5, fill='#666666', width=1)
            canvas.create_text(x, marker_y + 8, text=self.format_station(str(sta)),
                             font=('Consolas', 7), fill='black', anchor='n')

        # Direction labels
        canvas.create_text(margin_x + panel_width - 10, centerline_y + 8,
                         text="Right side (forward →)", font=('Consolas', 7),
                         fill='#666666', anchor='ne')
        canvas.create_text(margin_x + panel_width - 10, centerline_y - 8,
                         text="Left side (← opposite)", font=('Consolas', 7),
                         fill='#666666', anchor='se')

        return y_start + panel_height

    def get_lane_color(self, lane_type):
        """Get color for lane type based on IHSDM lane types"""
        colors = {
            'thru': '#3b82f6',       # Blue for through lanes
            'pass': '#8b5cf6',       # Purple for passing lanes
            'climb': '#7c3aed',      # Violet for climb lanes
            'left_turn': '#f59e0b',  # Orange for left turn lanes
            'right_turn': '#eab308', # Yellow for right turn lanes
            'taper': '#f97316',      # Dark orange for tapers
            'park': '#6b7280',       # Gray for parking lanes
            'bike': '#84cc16',       # Lime for bike lanes
            'accel': '#06b6d4',      # Cyan for accel lanes
            'decel': '#ec4899',      # Pink for decel lanes
            'other': '#9ca3af',      # Light gray for other
            'unspec': '#d1d5db'      # Lighter gray for unspecified
        }
        return colors.get(lane_type, '#3b82f6')  # Default to blue

    def draw_ramp_panel(self, canvas, ramps_data, margin_x, panel_width, y_start, min_sta, max_sta):
        """Draw ramp locations panel"""
        canvas.create_text(margin_x, y_start, anchor='nw', text="Ramp Locations",
                         font=('Segoe UI', 11, 'bold'), fill='#1e3a8a')
        y_start += 25

        panel_height = 80
        canvas.create_rectangle(margin_x, y_start, margin_x + panel_width, y_start + panel_height,
                              fill='white', outline='#666666', width=2)

        # Draw baseline
        baseline_y = y_start + panel_height // 2
        canvas.create_line(margin_x + 10, baseline_y, margin_x + panel_width - 10, baseline_y,
                         fill='#666666', width=2)

        # Draw ramp markers
        for ramp in ramps_data:
            x = margin_x + 10 + (ramp['station'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)
            color = '#10b981' if ramp['ramp_type'] == 'entrance' else '#ef4444'

            # Draw ramp marker (triangle)
            if ramp['ramp_type'] == 'entrance':
                canvas.create_polygon(x, baseline_y - 15, x - 6, baseline_y - 3, x + 6, baseline_y - 3,
                                    fill=color, outline='#1e3a8a', width=1)
            else:
                canvas.create_polygon(x, baseline_y + 15, x - 6, baseline_y + 3, x + 6, baseline_y + 3,
                                    fill=color, outline='#1e3a8a', width=1)

            # Label
            canvas.create_text(x, baseline_y + (30 if ramp['ramp_type'] == 'exit' else -30),
                             text=ramp['name'], font=('Consolas', 7), fill='black', anchor='center')

        return y_start + panel_height

    def draw_curves_panel(self, canvas, curves_data, margin_x, panel_width, y_start, min_sta, max_sta):
        """Draw horizontal curves panel"""
        canvas.create_text(margin_x, y_start, anchor='nw', text="Horizontal Alignment (Curves & Tangents)",
                         font=('Segoe UI', 11, 'bold'), fill='#1e3a8a')
        y_start += 25

        panel_height = 100
        canvas.create_rectangle(margin_x, y_start, margin_x + panel_width, y_start + panel_height,
                              fill='white', outline='#666666', width=2)

        # Draw alignment elements
        baseline_y = y_start + panel_height // 2
        for curve in curves_data:
            x_start = margin_x + 10 + (curve['begin'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)
            x_end = margin_x + 10 + (curve['end'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)

            if curve['type'] == 'tangent':
                # Draw straight line for tangent
                canvas.create_line(x_start, baseline_y, x_end, baseline_y, fill='#3b82f6', width=4)
                canvas.create_text((x_start + x_end) / 2, baseline_y - 15, text="Tangent",
                                 font=('Consolas', 7), fill='#666666')
            else:
                # Draw arc for curve
                direction = curve.get('direction', 'left')
                radius = curve.get('radius', 0)

                # Simplified curve visualization
                y_offset = -20 if direction == 'left' else 20
                mid_x = (x_start + x_end) / 2

                # Draw curve as smooth arc (approximate with line segments)
                points = []
                for i in range(11):
                    t = i / 10
                    x = x_start + t * (x_end - x_start)
                    # Parabolic approximation of curve
                    y = baseline_y + y_offset * (4 * t * (1 - t))
                    points.extend([x, y])

                if len(points) >= 4:
                    canvas.create_line(points, fill='#f59e0b', width=4, smooth=True)

                canvas.create_text(mid_x, baseline_y + y_offset + (15 if direction == 'right' else -15),
                                 text=f"Curve R={radius:.0f}'", font=('Consolas', 7), fill='#f59e0b')

        return y_start + panel_height

    def draw_traffic_panel(self, canvas, traffic_data, margin_x, panel_width, y_start, min_sta, max_sta):
        """Draw traffic volume panel"""
        canvas.create_text(margin_x, y_start, anchor='nw', text="Annual Average Daily Traffic (AADT)",
                         font=('Segoe UI', 11, 'bold'), fill='#1e3a8a')
        y_start += 25

        panel_height = 100
        canvas.create_rectangle(margin_x, y_start, margin_x + panel_width, y_start + panel_height,
                              fill='white', outline='#666666', width=2)

        if not traffic_data:
            return y_start + panel_height

        # Find max volume for scaling
        max_volume = max(t['volume'] for t in traffic_data)

        # Draw traffic bars
        for traffic in traffic_data:
            x_start = margin_x + 10 + (traffic['begin'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)
            x_end = margin_x + 10 + (traffic['end'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)

            bar_height = (traffic['volume'] / max_volume) * (panel_height - 40)
            y_bar = y_start + panel_height - 20 - bar_height

            canvas.create_rectangle(x_start, y_bar, x_end, y_start + panel_height - 20,
                                  fill='#8b5cf6', outline='#5b21b6', width=1)
            canvas.create_text((x_start + x_end) / 2, y_bar - 5, text=f"{traffic['volume']:,}",
                             font=('Consolas', 8), fill='#5b21b6', anchor='s')

        return y_start + panel_height

    def draw_median_panel(self, canvas, median_data, margin_x, panel_width, y_start, min_sta, max_sta):
        """Draw median width panel"""
        canvas.create_text(margin_x, y_start, anchor='nw', text="Median Width",
                         font=('Segoe UI', 11, 'bold'), fill='#1e3a8a')
        y_start += 25

        panel_height = 70
        canvas.create_rectangle(margin_x, y_start, margin_x + panel_width, y_start + panel_height,
                              fill='white', outline='#666666', width=2)

        baseline_y = y_start + panel_height - 20
        for median in median_data:
            x_start = margin_x + 10 + (median['begin'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)
            x_end = margin_x + 10 + (median['end'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)

            canvas.create_rectangle(x_start, y_start + 15, x_end, baseline_y,
                                  fill='#a8a29e', outline='#78716c', width=1)
            canvas.create_text((x_start + x_end) / 2, y_start + 25,
                             text=f"{median['width']:.0f}' {median['median_type']}",
                             font=('Consolas', 8), fill='white')

        return y_start + panel_height

    def draw_speed_panel(self, canvas, speed_data, margin_x, panel_width, y_start, min_sta, max_sta):
        """Draw posted speed panel"""
        canvas.create_text(margin_x, y_start, anchor='nw', text="Posted Speed Limit",
                         font=('Segoe UI', 11, 'bold'), fill='#1e3a8a')
        y_start += 25

        panel_height = 60
        canvas.create_rectangle(margin_x, y_start, margin_x + panel_width, y_start + panel_height,
                              fill='white', outline='#666666', width=2)

        for speed in speed_data:
            x_start = margin_x + 10 + (speed['begin'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)
            x_end = margin_x + 10 + (speed['end'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)

            canvas.create_rectangle(x_start, y_start + 10, x_end, y_start + panel_height - 10,
                                  fill='#fef3c7', outline='#f59e0b', width=2)
            canvas.create_text((x_start + x_end) / 2, y_start + panel_height // 2,
                             text=f"{speed['speed']} mph",
                             font=('Segoe UI', 10, 'bold'), fill='#f59e0b')

        return y_start + panel_height

    def draw_func_class_panel(self, canvas, func_class_data, margin_x, panel_width, y_start, min_sta, max_sta):
        """Draw functional class panel"""
        canvas.create_text(margin_x, y_start, anchor='nw', text="Functional Classification",
                         font=('Segoe UI', 11, 'bold'), fill='#1e3a8a')
        y_start += 25

        panel_height = 60
        canvas.create_rectangle(margin_x, y_start, margin_x + panel_width, y_start + panel_height,
                              fill='white', outline='#666666', width=2)

        for fc in func_class_data:
            x_start = margin_x + 10 + (fc['begin'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)
            x_end = margin_x + 10 + (fc['end'] - min_sta) / (max_sta - min_sta) * (panel_width - 20)

            color = '#dcfce7' if 'freeway' in fc['class_type'].lower() else '#e0e7ff'
            border = '#10b981' if 'freeway' in fc['class_type'].lower() else '#6366f1'

            canvas.create_rectangle(x_start, y_start + 10, x_end, y_start + panel_height - 10,
                                  fill=color, outline=border, width=2)
            canvas.create_text((x_start + x_end) / 2, y_start + panel_height // 2,
                             text=fc['class_type'],
                             font=('Consolas', 9), fill=border)

        return y_start + panel_height

    # =========================================================================
    # CMF SCANNER TAB
    # =========================================================================

    def setup_cmf_tab(self):
        """Setup the CMF calibration factor scanner tab"""
        tab_frame = ttk.Frame(self.cmf_tab, padding="10")
        tab_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.cmf_tab.columnconfigure(0, weight=1)
        self.cmf_tab.rowconfigure(0, weight=1)

        # Scan button
        scan_frame = ttk.Frame(tab_frame)
        scan_frame.grid(row=0, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))

        scan_btn = ttk.Button(scan_frame,
                             text="Scan for CMF Calibration Factors",
                             command=self.scan_cmf_files)
        scan_btn.pack(side=tk.LEFT, padx=5)

        # Export button
        export_cmf_btn = ttk.Button(scan_frame,
                                   text="Export to Excel",
                                   command=self.export_cmf_to_excel)
        export_cmf_btn.pack(side=tk.LEFT, padx=5)

        # Summary label
        self.cmf_summary_var = tk.StringVar(value="No scan performed yet")
        summary_label = ttk.Label(scan_frame,
                                 textvariable=self.cmf_summary_var,
                                 foreground='blue')
        summary_label.pack(side=tk.LEFT, padx=20)

        # Results tree
        tree_frame = ttk.Frame(tab_frame)
        tree_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        tab_frame.rowconfigure(1, weight=1)

        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")

        self.cmf_tree = ttk.Treeview(tree_frame,
                                    columns=('Type', 'ID', 'Name', 'Evaluation', 'Calibration'),
                                    show='tree headings',
                                    yscrollcommand=vsb.set,
                                    xscrollcommand=hsb.set)

        vsb.config(command=self.cmf_tree.yview)
        hsb.config(command=self.cmf_tree.xview)

        # Column configuration
        self.cmf_tree.heading('#0', text='Item')
        self.cmf_tree.heading('Type', text='Type')
        self.cmf_tree.heading('ID', text='ID')
        self.cmf_tree.heading('Name', text='Name')
        self.cmf_tree.heading('Evaluation', text='Evaluation')
        self.cmf_tree.heading('Calibration', text='Calibration Factor')

        self.cmf_tree.column('#0', width=200, anchor='w')
        self.cmf_tree.column('Type', width=100, anchor='center')
        self.cmf_tree.column('ID', width=80, anchor='center')
        self.cmf_tree.column('Name', width=300, anchor='w')
        self.cmf_tree.column('Evaluation', width=120, anchor='w')
        self.cmf_tree.column('Calibration', width=200, anchor='w')

        # Grid layout
        self.cmf_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))

        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        # Store CMF data for export
        self.cmf_data = []

    def scan_cmf_files(self):
        """Scan project for CMF calibration files"""
        print("DEBUG: scan_cmf_files called")
        project_path = self.project_path.get()
        print(f"DEBUG: project_path = {project_path}")
        print(f"DEBUG: placeholder_active = {self.placeholder_active}")

        # Check if placeholder is active
        if self.placeholder_active or not project_path or not os.path.isdir(project_path):
            print("DEBUG: Invalid project path")
            messagebox.showerror("Error", "Please select a valid project directory")
            return

        print("DEBUG: Starting scan...")

        self.status_var.set("Scanning for CMF files...")
        self.cmf_tree.delete(*self.cmf_tree.get_children())
        self.cmf_data = []

        try:
            # Find all evaluation.1.cpm.cmf.csv files
            cmf_files = []
            for root, dirs, files in os.walk(project_path):
                for file in files:
                    if file == 'evaluation.1.cpm.cmf.csv':
                        cmf_files.append(os.path.join(root, file))

            if not cmf_files:
                self.cmf_summary_var.set("No CMF files found")
                self.status_var.set("No CMF files found in project")
                return

            # Parse each CMF file
            highway_count = 0
            intersection_count = 0
            ramp_count = 0

            for cmf_path in cmf_files:
                # Extract alignment info from path
                # Path format: .../h1/e6/evaluation.1.cpm.cmf.csv or .../i100/e1/evaluation.1.cpm.cmf.csv
                path_parts = cmf_path.replace('\\', '/').split('/')

                alignment_id = None
                alignment_type = None
                eval_folder = None

                for i, part in enumerate(path_parts):
                    if part.startswith('h') and part[1:].isdigit():
                        alignment_id = part
                        alignment_type = 'Highway'
                        if i + 1 < len(path_parts):
                            eval_folder = path_parts[i + 1]
                        break
                    elif part.startswith('i') and part[1:].isdigit():
                        alignment_id = part
                        alignment_type = 'Intersection'
                        if i + 1 < len(path_parts):
                            eval_folder = path_parts[i + 1]
                        break
                    elif part.startswith('r') and part[1:].isdigit():
                        alignment_id = part
                        alignment_type = 'Ramp Terminal'
                        if i + 1 < len(path_parts):
                            eval_folder = path_parts[i + 1]
                        break

                if not alignment_id:
                    continue

                # Read the CMF file and extract line 27, column B
                calibration = None
                alignment_name = "Unknown"

                try:
                    with open(cmf_path, 'r', encoding='utf-8') as f:
                        lines = f.readlines()

                        # Line 27 (index 26) contains calibration
                        if len(lines) >= 27:
                            line27 = lines[26].strip()
                            parts = line27.split(',')
                            if len(parts) >= 2:
                                # Remove quotes from calibration value
                                calibration = parts[1].strip('"')

                        # Line 17 contains highway/alignment name
                        if len(lines) >= 17:
                            line17 = lines[16].strip()
                            parts = line17.split(',')
                            if len(parts) >= 2:
                                alignment_name = parts[1].strip('"')

                except Exception as e:
                    calibration = f"Error: {str(e)}"

                # Store data
                cmf_entry = {
                    'type': alignment_type,
                    'id': alignment_id,
                    'name': alignment_name,
                    'evaluation': eval_folder if eval_folder else 'Unknown',
                    'calibration': calibration if calibration else 'Not found',
                    'path': cmf_path
                }
                self.cmf_data.append(cmf_entry)

                # Count by type
                if alignment_type == 'Highway':
                    highway_count += 1
                elif alignment_type == 'Intersection':
                    intersection_count += 1
                elif alignment_type == 'Ramp Terminal':
                    ramp_count += 1

            # Populate tree
            self.populate_cmf_tree()

            # Update summary
            summary = f"Total: {len(self.cmf_data)} | "
            summary += f"Highways: {highway_count} | "
            summary += f"Intersections: {intersection_count} | "
            summary += f"Ramp Terminals: {ramp_count}"
            self.cmf_summary_var.set(summary)
            self.status_var.set(f"Found {len(self.cmf_data)} CMF files")

        except Exception as e:
            messagebox.showerror("Scan Error", f"Error scanning CMF files:\n{str(e)}")
            self.status_var.set("CMF scan failed")

    def populate_cmf_tree(self):
        """Populate the CMF tree with grouped data"""
        # Group by type
        highways = [d for d in self.cmf_data if d['type'] == 'Highway']
        intersections = [d for d in self.cmf_data if d['type'] == 'Intersection']
        ramps = [d for d in self.cmf_data if d['type'] == 'Ramp Terminal']

        # Add highways
        if highways:
            hw_parent = self.cmf_tree.insert('', 'end', text=f'HIGHWAYS ({len(highways)})',
                                            values=('', '', '', '', ''))
            for hw in sorted(highways, key=lambda x: x['id']):
                self.cmf_tree.insert(hw_parent, 'end',
                                   text=hw['id'],
                                   values=(hw['type'], hw['id'], hw['name'],
                                         hw['evaluation'], hw['calibration']))

        # Add intersections
        if intersections:
            int_parent = self.cmf_tree.insert('', 'end', text=f'INTERSECTIONS ({len(intersections)})',
                                             values=('', '', '', '', ''))
            for i in sorted(intersections, key=lambda x: x['id']):
                self.cmf_tree.insert(int_parent, 'end',
                                   text=i['id'],
                                   values=(i['type'], i['id'], i['name'],
                                         i['evaluation'], i['calibration']))

        # Add ramp terminals
        if ramps:
            ramp_parent = self.cmf_tree.insert('', 'end', text=f'RAMP TERMINALS ({len(ramps)})',
                                              values=('', '', '', '', ''))
            for r in sorted(ramps, key=lambda x: x['id']):
                self.cmf_tree.insert(ramp_parent, 'end',
                                   text=r['id'],
                                   values=(r['type'], r['id'], r['name'],
                                         r['evaluation'], r['calibration']))

    def export_cmf_to_excel(self):
        """Export CMF data to Excel"""
        if not self.cmf_data:
            messagebox.showwarning("No Data", "No CMF data to export. Please scan first.")
            return

        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Missing Dependency",
                               "openpyxl is required for Excel export.\n"
                               "Install it with: pip install openpyxl")
            return

        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="CMF_Calibration_Factors.xlsx"
        )

        if not filename:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = Workbook()
            ws = wb.active
            ws.title = "CMF Calibration Factors"

            # Headers
            headers = ['Type', 'ID', 'Name', 'Evaluation', 'Calibration Factor', 'File Path']
            ws.append(headers)

            # Style headers
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')

            # Add data
            for entry in self.cmf_data:
                ws.append([
                    entry['type'],
                    entry['id'],
                    entry['name'],
                    entry['evaluation'],
                    entry['calibration'],
                    entry['path']
                ])

            # Adjust column widths
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 50
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 25
            ws.column_dimensions['F'].width = 60

            wb.save(filename)
            messagebox.showinfo("Export Successful",
                              f"CMF data exported to:\n{filename}")
            self.status_var.set(f"Exported CMF data to {os.path.basename(filename)}")

        except Exception as e:
            messagebox.showerror("Export Error",
                               f"Failed to export to Excel:\n{str(e)}")

    # =========================================================================
    # UPDATE CHECKING
    # =========================================================================

    def check_for_updates(self, show_current=False):
        """Check GitHub for new releases"""
        if not GITHUB_API_URL:
            if show_current:
                messagebox.showinfo("Version",
                    f"Current version: {__version__}\n\n"
                    "GitHub repository not configured for auto-updates.")
            return

        def check_thread():
            try:
                # Make request to GitHub API with longer timeout for corporate networks
                req = Request(GITHUB_API_URL)
                req.add_header('User-Agent', f'{__app_name__}/{__version__}')

                # Increased timeout to 15 seconds for slower corporate networks
                with urlopen(req, timeout=15) as response:
                    data = json.loads(response.read().decode('utf-8'))

                    latest_version = data['tag_name'].lstrip('v')
                    download_url = data['html_url']
                    release_notes = data['body']

                    # Compare versions
                    if self._compare_versions(latest_version, __version__) > 0:
                        # New version available
                        self.root.after(0, lambda: self._show_update_dialog(
                            latest_version, download_url, release_notes))
                    elif show_current:
                        self.root.after(0, lambda: messagebox.showinfo("Up to Date",
                            f"You have the latest version ({__version__})"))

            except URLError as e:
                if show_current:
                    self.root.after(0, lambda: self._show_update_failed_dialog())
            except Exception as e:
                if show_current:
                    self.root.after(0, lambda: self._show_update_failed_dialog())

        # Run in background thread
        thread = threading.Thread(target=check_thread, daemon=True)
        thread.start()

    def _compare_versions(self, v1, v2):
        """Compare version strings (returns: 1 if v1>v2, -1 if v1<v2, 0 if equal)"""
        parts1 = [int(x) for x in v1.split('.')]
        parts2 = [int(x) for x in v2.split('.')]

        for p1, p2 in zip(parts1, parts2):
            if p1 > p2:
                return 1
            elif p1 < p2:
                return -1

        if len(parts1) > len(parts2):
            return 1
        elif len(parts1) < len(parts2):
            return -1

        return 0

    def _show_update_failed_dialog(self):
        """Show dialog when update check fails with link to releases page"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Update Check Failed")
        dialog.geometry("450x280")
        dialog.transient(self.root)
        dialog.grab_set()

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (450 // 2)
        y = (dialog.winfo_screenheight() // 2) - (280 // 2)
        dialog.geometry(f"450x280+{x}+{y}")

        # Content
        content = tk.Frame(dialog, padx=25, pady=25)
        content.pack(fill=tk.BOTH, expand=True)

        # Title
        tk.Label(content, text="Cannot Check for Updates",
                font=('Segoe UI', 12, 'bold')).pack(pady=(0, 15))

        # Message
        message = (
            "Could not connect to GitHub to check for updates.\n\n"
            "This may be due to:\n"
            "  • Firewall blocking GitHub access\n"
            "  • No internet connection\n"
            "  • Corporate proxy settings\n\n"
            "You can still check for updates manually:"
        )
        tk.Label(content, text=message, justify=tk.LEFT,
                font=('Segoe UI', 9)).pack(pady=(0, 15))

        # Buttons
        button_frame = tk.Frame(content)
        button_frame.pack(pady=10)

        def open_releases():
            if GITHUB_RELEASES_URL:
                webbrowser.open(GITHUB_RELEASES_URL)
            dialog.destroy()

        tk.Button(button_frame, text="Open Releases Page",
                 command=open_releases,
                 bg=self.colors['primary'],
                 fg='white',
                 font=('Segoe UI', 10, 'bold'),
                 padx=20, pady=8,
                 cursor='hand2').pack(side=tk.LEFT, padx=5)

        tk.Button(button_frame, text="Close",
                 command=dialog.destroy,
                 font=('Segoe UI', 10),
                 padx=20, pady=8).pack(side=tk.LEFT, padx=5)

        # Note
        tk.Label(content, text="The application works normally without update checks.",
                font=('Segoe UI', 8, 'italic'),
                fg='gray').pack(pady=(15, 0))

    def _show_update_dialog(self, new_version, download_url, release_notes):
        """Show dialog when update is available"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Update Available")
        dialog.geometry("500x400")
        dialog.transient(self.root)
        dialog.grab_set()

        # Center dialog
        dialog.update_idletasks()
        x = (dialog.winfo_screenwidth() // 2) - (500 // 2)
        y = (dialog.winfo_screenheight() // 2) - (400 // 2)
        dialog.geometry(f"500x400+{x}+{y}")

        # Header
        header = tk.Frame(dialog, bg=self.colors['primary'], height=60)
        header.pack(fill=tk.X)
        header.pack_propagate(False)

        tk.Label(header, text="🎉 New Version Available!",
                font=('Segoe UI', 14, 'bold'),
                bg=self.colors['primary'],
                fg='white').pack(pady=15)

        # Content
        content = tk.Frame(dialog, padx=20, pady=20)
        content.pack(fill=tk.BOTH, expand=True)

        # Version info
        info_frame = tk.Frame(content)
        info_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(info_frame, text=f"Current version: {__version__}",
                font=('Segoe UI', 10)).pack(anchor=tk.W)
        tk.Label(info_frame, text=f"New version: {new_version}",
                font=('Segoe UI', 10, 'bold'),
                fg=self.colors['primary']).pack(anchor=tk.W)

        # Release notes
        tk.Label(content, text="What's New:",
                font=('Segoe UI', 10, 'bold')).pack(anchor=tk.W, pady=(10, 5))

        notes_text = scrolledtext.ScrolledText(content, wrap=tk.WORD, height=10,
                                               font=('Segoe UI', 9))
        notes_text.pack(fill=tk.BOTH, expand=True)
        notes_text.insert('1.0', release_notes if release_notes else "No release notes available.")
        notes_text.config(state=tk.DISABLED)

        # Buttons
        btn_frame = tk.Frame(content)
        btn_frame.pack(fill=tk.X, pady=(15, 0))

        def open_download():
            webbrowser.open(download_url)
            dialog.destroy()

        ttk.Button(btn_frame, text="Download Update", command=open_download,
                  style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Remind Me Later", command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    # =========================================================================
    # AADT INPUT TAB
    # =========================================================================

    def setup_aadt_tab(self):
        """Setup the AADT Input tab with step-by-step wizard"""
        tab_frame = ttk.Frame(self.aadt_tab, padding="10")
        tab_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.aadt_tab.columnconfigure(0, weight=1)
        self.aadt_tab.rowconfigure(0, weight=1)

        # Initialize AADT-specific state
        self.aadt_forecast_path = tk.StringVar()
        self.aadt_forecast_columns = tk.StringVar(value="DQ:DR,DS:DT")  # Default column ranges
        self.aadt_year = tk.StringVar(value="2028")
        self.aadt_sections = []  # Will hold parsed AADT section data
        self.aadt_forecast_ids = {}  # Will hold mapping of forecast IDs to values
        self.aadt_section_mappings = {}  # Will hold section -> forecast IDs mappings

        # Compatibility notice at top
        notice_frame = tk.Frame(tab_frame, bg='#fff3cd', padx=10, pady=8)
        notice_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        notice_text = ("NOTE: This tool is designed for use with HNTB Wisconsin's forecasting spreadsheet format. "
                      "Currently supports setting AADT for one evaluation year at a time.")
        tk.Label(notice_frame, text=notice_text, font=('Segoe UI', 9, 'italic'),
                bg='#fff3cd', fg='#856404', wraplength=900, justify=tk.LEFT).pack(anchor='w')

        # Create wizard-style interface with numbered steps
        # Step indicator
        step_indicator = ttk.Frame(tab_frame)
        step_indicator.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=(0, 10))

        self.aadt_step_labels = []
        steps = ["1. Setup IHSDM", "2. Link Forecast", "3. Scan Project", "4. Map Sections", "5. Apply AADT"]
        for i, step in enumerate(steps):
            lbl = ttk.Label(step_indicator, text=step, font=('Segoe UI', 9))
            lbl.grid(row=0, column=i, padx=10)
            self.aadt_step_labels.append(lbl)

        # Main content area with scrollable frame
        content_canvas = tk.Canvas(tab_frame, highlightthickness=0)
        content_scrollbar = ttk.Scrollbar(tab_frame, orient="vertical", command=content_canvas.yview)
        self.aadt_content_frame = ttk.Frame(content_canvas)

        self.aadt_content_frame.bind(
            "<Configure>",
            lambda e: content_canvas.configure(scrollregion=content_canvas.bbox("all"))
        )

        content_canvas.create_window((0, 0), window=self.aadt_content_frame, anchor="nw")
        content_canvas.configure(yscrollcommand=content_scrollbar.set)

        content_canvas.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        content_scrollbar.grid(row=2, column=1, sticky=(tk.N, tk.S))

        # Bind mousewheel
        def _on_mousewheel(event):
            content_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        content_canvas.bind("<Enter>", lambda e: content_canvas.bind_all("<MouseWheel>", _on_mousewheel))
        content_canvas.bind("<Leave>", lambda e: content_canvas.unbind_all("<MouseWheel>"))

        tab_frame.columnconfigure(0, weight=1)
        tab_frame.rowconfigure(2, weight=1)

        # ===== STEP 1: Setup Instructions =====
        step1_frame = ttk.LabelFrame(self.aadt_content_frame, text="Step 1: Initial IHSDM Setup (Do This First!)", padding="15")
        step1_frame.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=5, padx=5)

        step1_text = """BEFORE using this tool, you must set up AADT station ranges in IHSDM:

1. Open your IHSDM project in the IHSDM application
2. For each highway alignment, go to Traffic Data → Annual Average Daily Traffic
3. Add AADT ranges for each segment where traffic volumes change:
   • Set the Start Station and End Station for each range
   • Enter a DEFAULT VALUE of 1 for the AADT (this tool will update the actual values)
   • Enter the Year you are evaluating (e.g., 2028)

Example: If you have a mainline with 3 different volume sections:
   Range 1: Sta 1000 to 2500, Year 2028, AADT = 1
   Range 2: Sta 2500 to 4000, Year 2028, AADT = 1
   Range 3: Sta 4000 to 5500, Year 2028, AADT = 1

Once you've set up all your AADT ranges in IHSDM, proceed to Step 2."""

        step1_label = tk.Label(step1_frame, text=step1_text, font=('Segoe UI', 9),
                              justify=tk.LEFT, anchor='w', bg=self.colors['bg_light'])
        step1_label.pack(fill=tk.X, anchor='w')

        # ===== STEP 2: Link Forecast Workbook =====
        step2_frame = ttk.LabelFrame(self.aadt_content_frame, text="Step 2: Link Project Forecast Workbook", padding="15")
        step2_frame.grid(row=1, column=0, sticky=(tk.W, tk.E), pady=5, padx=5)

        ttk.Label(step2_frame, text="Forecast Workbook (with BalancedOutput sheet):").grid(row=0, column=0, sticky=tk.W, pady=5)
        forecast_entry = ttk.Entry(step2_frame, textvariable=self.aadt_forecast_path, width=60)
        forecast_entry.grid(row=0, column=1, padx=5, pady=5)
        ttk.Button(step2_frame, text="Browse...", command=self.browse_aadt_forecast).grid(row=0, column=2, padx=5, pady=5)

        ttk.Label(step2_frame, text="BalancedOutput Column Ranges (daily forecast for eval year):").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(step2_frame, textvariable=self.aadt_forecast_columns, width=30).grid(row=1, column=1, sticky=tk.W, padx=5, pady=5)
        ttk.Label(step2_frame, text="Format: MainlineID:MainlineVal,TurnID:TurnVal (e.g., DQ:DR,DS:DT)", font=('Segoe UI', 8, 'italic')).grid(row=1, column=2, sticky=tk.W)

        ttk.Label(step2_frame, text="Evaluation Year:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(step2_frame, textvariable=self.aadt_year, width=10).grid(row=2, column=1, sticky=tk.W, padx=5, pady=5)

        btn_frame = ttk.Frame(step2_frame)
        btn_frame.grid(row=3, column=0, columnspan=3, pady=10)
        ttk.Button(btn_frame, text="Load Forecast Data", command=self.load_aadt_forecast,
                  style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Show Loaded IDs", command=self.show_loaded_forecast_ids).pack(side=tk.LEFT, padx=5)

        self.aadt_forecast_status = tk.StringVar(value="No forecast loaded")
        ttk.Label(step2_frame, textvariable=self.aadt_forecast_status, foreground='gray').grid(row=4, column=0, columnspan=3, sticky=tk.W)

        step2_frame.columnconfigure(1, weight=1)

        # ===== STEP 3: Scan Project =====
        step3_frame = ttk.LabelFrame(self.aadt_content_frame, text="Step 3: Scan Project for AADT Sections", padding="15")
        step3_frame.grid(row=2, column=0, sticky=(tk.W, tk.E), pady=5, padx=5)

        step3_desc = "Scan the project to find all highway alignments and their AADT station ranges."
        ttk.Label(step3_frame, text=step3_desc).grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=5)

        ttk.Button(step3_frame, text="Scan for AADT Sections", command=self.scan_aadt_sections,
                  style='Accent.TButton').grid(row=1, column=0, pady=10)

        self.aadt_scan_status = tk.StringVar(value="No scan performed")
        ttk.Label(step3_frame, textvariable=self.aadt_scan_status, foreground='gray').grid(row=1, column=1, sticky=tk.W, padx=10)

        # ===== STEP 4: Map Sections to Forecast IDs =====
        step4_frame = ttk.LabelFrame(self.aadt_content_frame, text="Step 4: Map AADT Sections to Forecast IDs", padding="15")
        step4_frame.grid(row=3, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5, padx=5)

        step4_desc = """For each AADT section, enter the Forecast IDs that contribute to that section's volume.
Click on an alignment to expand and see its sections. Enter IDs, then click "Mark Reviewed" when done with each alignment."""
        ttk.Label(step4_frame, text=step4_desc, wraplength=800).grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=5)

        # Left side: Treeview for sections (collapsible by alignment)
        tree_container = ttk.Frame(step4_frame)
        tree_container.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)

        columns = ('Section', 'Start Sta', 'End Sta', 'Year', 'Current', 'Forecast IDs', 'Calculated')
        self.aadt_tree = ttk.Treeview(tree_container, columns=columns, show='tree headings', height=12)

        # Configure tree column (for alignment names)
        self.aadt_tree.heading('#0', text='Alignment', command=lambda: self.sort_aadt_tree('#0'))
        self.aadt_tree.column('#0', width=180, anchor='w')

        # Configure data columns with sort capability
        col_widths = {'Section': 55, 'Start Sta': 85, 'End Sta': 85, 'Year': 50,
                     'Current': 60, 'Forecast IDs': 200, 'Calculated': 75}
        for col in columns:
            self.aadt_tree.heading(col, text=col, command=lambda c=col: self.sort_aadt_tree(c))
            width = col_widths.get(col, 70)
            self.aadt_tree.column(col, width=width, anchor='center')

        vsb = ttk.Scrollbar(tree_container, orient="vertical", command=self.aadt_tree.yview)
        hsb = ttk.Scrollbar(tree_container, orient="horizontal", command=self.aadt_tree.xview)
        self.aadt_tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.aadt_tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        vsb.grid(row=0, column=1, sticky=(tk.N, tk.S))
        hsb.grid(row=1, column=0, sticky=(tk.W, tk.E))

        tree_container.columnconfigure(0, weight=1)
        tree_container.rowconfigure(0, weight=1)

        # Configure tree tags for styling
        self.aadt_tree.tag_configure('reviewed', background='#d4edda')  # Light green
        self.aadt_tree.tag_configure('pending', background='#fff3cd')   # Light yellow
        self.aadt_tree.tag_configure('alignment', font=('Segoe UI', 9, 'bold'))

        # Right side: Checklist panel
        checklist_frame = ttk.LabelFrame(step4_frame, text="Review Checklist", padding="10")
        checklist_frame.grid(row=1, column=1, sticky=(tk.N, tk.S, tk.E), pady=5, padx=(10, 0))

        # Checklist canvas with scrollbar
        checklist_canvas = tk.Canvas(checklist_frame, width=200, height=250, highlightthickness=0)
        checklist_scrollbar = ttk.Scrollbar(checklist_frame, orient="vertical", command=checklist_canvas.yview)
        self.aadt_checklist_frame = ttk.Frame(checklist_canvas)

        self.aadt_checklist_frame.bind(
            "<Configure>",
            lambda e: checklist_canvas.configure(scrollregion=checklist_canvas.bbox("all"))
        )

        checklist_canvas.create_window((0, 0), window=self.aadt_checklist_frame, anchor="nw")
        checklist_canvas.configure(yscrollcommand=checklist_scrollbar.set)

        checklist_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        checklist_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Track reviewed alignments
        self.aadt_reviewed_alignments = set()

        # Edit controls - Row 1: IDs 1-3 with signs
        edit_frame = ttk.Frame(step4_frame)
        edit_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 2))

        ttk.Label(edit_frame, text="Forecast IDs (+ to add, - to subtract):").pack(side=tk.LEFT, padx=5)

        # Initialize ID and sign variables
        self.aadt_id1_var = tk.StringVar()
        self.aadt_id2_var = tk.StringVar()
        self.aadt_id3_var = tk.StringVar()
        self.aadt_id4_var = tk.StringVar()
        self.aadt_id5_var = tk.StringVar()
        self.aadt_id6_var = tk.StringVar()
        self.aadt_sign1_var = tk.StringVar(value='+')
        self.aadt_sign2_var = tk.StringVar(value='+')
        self.aadt_sign3_var = tk.StringVar(value='+')
        self.aadt_sign4_var = tk.StringVar(value='+')
        self.aadt_sign5_var = tk.StringVar(value='+')
        self.aadt_sign6_var = tk.StringVar(value='+')

        # ID 1
        ttk.Combobox(edit_frame, textvariable=self.aadt_sign1_var, values=['+', '-'], width=2, state='readonly').pack(side=tk.LEFT, padx=(10, 0))
        ttk.Entry(edit_frame, textvariable=self.aadt_id1_var, width=8).pack(side=tk.LEFT, padx=(0, 5))
        # ID 2
        ttk.Combobox(edit_frame, textvariable=self.aadt_sign2_var, values=['+', '-'], width=2, state='readonly').pack(side=tk.LEFT, padx=(5, 0))
        ttk.Entry(edit_frame, textvariable=self.aadt_id2_var, width=8).pack(side=tk.LEFT, padx=(0, 5))
        # ID 3
        ttk.Combobox(edit_frame, textvariable=self.aadt_sign3_var, values=['+', '-'], width=2, state='readonly').pack(side=tk.LEFT, padx=(5, 0))
        ttk.Entry(edit_frame, textvariable=self.aadt_id3_var, width=8).pack(side=tk.LEFT, padx=(0, 5))

        # Edit controls - Row 2: IDs 4-6 with signs
        edit_frame1b = ttk.Frame(step4_frame)
        edit_frame1b.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=2)

        ttk.Label(edit_frame1b, text="                                                        ").pack(side=tk.LEFT)  # Spacer

        # ID 4
        ttk.Combobox(edit_frame1b, textvariable=self.aadt_sign4_var, values=['+', '-'], width=2, state='readonly').pack(side=tk.LEFT, padx=(10, 0))
        ttk.Entry(edit_frame1b, textvariable=self.aadt_id4_var, width=8).pack(side=tk.LEFT, padx=(0, 5))
        # ID 5
        ttk.Combobox(edit_frame1b, textvariable=self.aadt_sign5_var, values=['+', '-'], width=2, state='readonly').pack(side=tk.LEFT, padx=(5, 0))
        ttk.Entry(edit_frame1b, textvariable=self.aadt_id5_var, width=8).pack(side=tk.LEFT, padx=(0, 5))
        # ID 6
        ttk.Combobox(edit_frame1b, textvariable=self.aadt_sign6_var, values=['+', '-'], width=2, state='readonly').pack(side=tk.LEFT, padx=(5, 0))
        ttk.Entry(edit_frame1b, textvariable=self.aadt_id6_var, width=8).pack(side=tk.LEFT, padx=(0, 5))

        ttk.Button(edit_frame1b, text="Apply to Selected", command=self.apply_aadt_ids_to_selected).pack(side=tk.LEFT, padx=15)

        # Third row of controls - action buttons
        edit_frame2 = ttk.Frame(step4_frame)
        edit_frame2.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        ttk.Button(edit_frame2, text="Mark Alignment Reviewed", command=self.mark_alignment_reviewed,
                  style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(edit_frame2, text="Calculate All", command=self.calculate_all_aadt).pack(side=tk.LEFT, padx=5)
        ttk.Button(edit_frame2, text="Expand All", command=self.expand_all_aadt_tree).pack(side=tk.LEFT, padx=5)
        ttk.Button(edit_frame2, text="Collapse All", command=self.collapse_all_aadt_tree).pack(side=tk.LEFT, padx=5)

        # Fourth row - Save/Load buttons
        edit_frame3 = ttk.Frame(step4_frame)
        edit_frame3.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)

        ttk.Label(edit_frame3, text="Session:", font=('Segoe UI', 9, 'bold')).pack(side=tk.LEFT, padx=5)
        ttk.Button(edit_frame3, text="Save Mapping to CSV", command=self.save_aadt_mapping_csv).pack(side=tk.LEFT, padx=5)
        ttk.Button(edit_frame3, text="Load Mapping from CSV", command=self.load_aadt_mapping_csv).pack(side=tk.LEFT, padx=5)

        # Bind selection event
        self.aadt_tree.bind('<<TreeviewSelect>>', self.on_aadt_tree_select)

        step4_frame.columnconfigure(0, weight=1)
        step4_frame.rowconfigure(1, weight=1)

        # ===== STEP 5: Apply AADT Values =====
        step5_frame = ttk.LabelFrame(self.aadt_content_frame, text="Step 5: Apply AADT Values to IHSDM Project", padding="15")
        step5_frame.grid(row=4, column=0, sticky=(tk.W, tk.E), pady=5, padx=5)

        step5_desc = """Once all sections have calculated AADT values, click below to update the highway XML files.
This will modify the adtRate attribute in each AnnualAveDailyTraffic element."""
        ttk.Label(step5_frame, text=step5_desc, wraplength=800).grid(row=0, column=0, columnspan=2, sticky=tk.W, pady=5)

        btn_frame = ttk.Frame(step5_frame)
        btn_frame.grid(row=1, column=0, columnspan=2, pady=10)

        ttk.Button(btn_frame, text="Preview Changes", command=self.preview_aadt_changes).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Apply AADT to XML Files", command=self.apply_aadt_to_xml,
                  style='Accent.TButton').pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_frame, text="Export Mapping to Excel", command=self.export_aadt_mapping).pack(side=tk.LEFT, padx=5)

        self.aadt_apply_status = tk.StringVar(value="")
        ttk.Label(step5_frame, textvariable=self.aadt_apply_status, foreground='green').grid(row=2, column=0, columnspan=2, sticky=tk.W)

        # Configure main content frame
        self.aadt_content_frame.columnconfigure(0, weight=1)

    def setup_aadt_disabled_tab(self):
        """Show message when openpyxl is not available"""
        message_frame = ttk.Frame(self.aadt_tab, padding="50")
        message_frame.pack(expand=True)

        ttk.Label(message_frame, text="AADT Input Unavailable",
                 font=('Arial', 14, 'bold')).pack(pady=10)
        ttk.Label(message_frame, text="The openpyxl library is required for the AADT Input feature.",
                 font=('Arial', 10)).pack(pady=5)
        ttk.Label(message_frame, text="Install it using:", font=('Arial', 10)).pack(pady=5)
        ttk.Label(message_frame, text="pip install openpyxl",
                 font=('Consolas', 11, 'bold'), foreground='blue').pack(pady=10)
        ttk.Label(message_frame, text="Then restart this application.", font=('Arial', 10)).pack(pady=5)

    def browse_aadt_forecast(self):
        """Browse for forecast workbook"""
        file_path = filedialog.askopenfilename(
            title="Select Forecast Workbook",
            filetypes=[("Excel files", "*.xlsx *.xlsb"), ("Excel Workbook", "*.xlsx"), ("Excel Binary", "*.xlsb"), ("All files", "*.*")]
        )
        if file_path:
            self.aadt_forecast_path.set(file_path)

    def load_aadt_forecast(self):
        """Load forecast data from workbook (supports .xlsx and .xlsb)"""
        forecast_path = self.aadt_forecast_path.get()

        if not forecast_path or not os.path.isfile(forecast_path):
            messagebox.showerror("Error", "Please select a valid forecast workbook")
            return

        try:
            # Check file extension
            is_xlsb = forecast_path.lower().endswith('.xlsb')

            if is_xlsb:
                # Use pyxlsb for binary Excel files
                try:
                    from pyxlsb import open_workbook as open_xlsb
                except ImportError:
                    messagebox.showerror("Error",
                        "The pyxlsb library is required to read .xlsb files.\n\n"
                        "Install with: pip install pyxlsb")
                    return

                self.aadt_forecast_ids = {}
                col_ranges = self.aadt_forecast_columns.get().split(',')

                with open_xlsb(forecast_path) as wb:
                    # Find BalancedOutput sheet
                    sheet_name = None
                    for name in wb.sheets:
                        if 'balancedoutput' in name.lower():
                            sheet_name = name
                            break

                    if not sheet_name:
                        messagebox.showerror("Error", "Could not find 'BalancedOutput' sheet in workbook")
                        return

                    with wb.get_sheet(sheet_name) as sheet:
                        # Read all rows into memory
                        rows = list(sheet.rows())

                        for col_range in col_ranges:
                            col_range = col_range.strip()
                            if ':' not in col_range:
                                continue

                            id_col, val_col = col_range.split(':')
                            # Convert column letters to 0-based index
                            from openpyxl.utils import column_index_from_string
                            id_col_num = column_index_from_string(id_col.strip()) - 1
                            val_col_num = column_index_from_string(val_col.strip()) - 1

                            for row in rows:
                                if len(row) > max(id_col_num, val_col_num):
                                    id_cell = row[id_col_num].v if row[id_col_num] else None
                                    val_cell = row[val_col_num].v if row[val_col_num] else None

                                    if id_cell is not None:
                                        # Handle numeric IDs - convert float to int string if possible
                                        if isinstance(id_cell, float) and id_cell == int(id_cell):
                                            id_str = str(int(id_cell))
                                        else:
                                            id_str = str(id_cell).strip()
                                        if id_str:
                                            try:
                                                val = float(val_cell) if val_cell is not None else 0
                                                self.aadt_forecast_ids[id_str] = val
                                            except (ValueError, TypeError):
                                                self.aadt_forecast_ids[id_str] = 0
            else:
                # Use openpyxl for .xlsx files
                wb = load_workbook(forecast_path, data_only=True)

                # Look for BalancedOutput sheet
                sheet_name = None
                for name in wb.sheetnames:
                    if 'balancedoutput' in name.lower():
                        sheet_name = name
                        break

                if not sheet_name:
                    messagebox.showerror("Error", "Could not find 'BalancedOutput' sheet in workbook")
                    return

                ws = wb[sheet_name]

                # Parse column ranges
                col_ranges = self.aadt_forecast_columns.get().split(',')
                self.aadt_forecast_ids = {}

                for col_range in col_ranges:
                    col_range = col_range.strip()
                    if ':' not in col_range:
                        continue

                    id_col, val_col = col_range.split(':')
                    id_col = id_col.strip()
                    val_col = val_col.strip()

                    # Convert column letters to numbers
                    from openpyxl.utils import column_index_from_string
                    id_col_num = column_index_from_string(id_col)
                    val_col_num = column_index_from_string(val_col)

                    # Read all rows in this column pair
                    for row in range(1, ws.max_row + 1):
                        id_cell = ws.cell(row=row, column=id_col_num).value
                        val_cell = ws.cell(row=row, column=val_col_num).value

                        if id_cell is not None:
                            # Handle numeric IDs - convert float to int string if possible
                            if isinstance(id_cell, float) and id_cell == int(id_cell):
                                id_str = str(int(id_cell))
                            else:
                                id_str = str(id_cell).strip()
                            if id_str:
                                try:
                                    val = float(val_cell) if val_cell is not None else 0
                                    self.aadt_forecast_ids[id_str] = val
                                except (ValueError, TypeError):
                                    self.aadt_forecast_ids[id_str] = 0

                wb.close()

            # Show sample of loaded IDs for debugging
            sample_ids = list(self.aadt_forecast_ids.keys())[:10]
            print(f"Loaded {len(self.aadt_forecast_ids)} forecast IDs. Sample: {sample_ids}")

            self.aadt_forecast_status.set(f"Loaded {len(self.aadt_forecast_ids)} forecast IDs from {sheet_name}")
            self.status_var.set(f"Forecast data loaded: {len(self.aadt_forecast_ids)} IDs")

        except Exception as e:
            messagebox.showerror("Error", f"Error loading forecast: {str(e)}")
            self.aadt_forecast_status.set(f"Error: {str(e)}")

    def show_loaded_forecast_ids(self):
        """Show a popup with all loaded forecast IDs for debugging"""
        if not self.aadt_forecast_ids:
            messagebox.showinfo("No Forecast Data", "No forecast data has been loaded yet.\n\nClick 'Load Forecast Data' first.")
            return

        # Create a popup window to show all IDs
        popup = tk.Toplevel(self.root)
        popup.title("Loaded Forecast IDs")
        popup.geometry("500x600")
        popup.transient(self.root)

        ttk.Label(popup, text=f"Loaded {len(self.aadt_forecast_ids)} Forecast IDs:",
                 font=('Segoe UI', 11, 'bold')).pack(pady=10)

        # Add a search box
        search_frame = ttk.Frame(popup)
        search_frame.pack(fill=tk.X, padx=10)
        ttk.Label(search_frame, text="Search:").pack(side=tk.LEFT)
        search_var = tk.StringVar()
        search_entry = ttk.Entry(search_frame, textvariable=search_var, width=30)
        search_entry.pack(side=tk.LEFT, padx=5)

        # Listbox with scrollbar
        list_frame = ttk.Frame(popup)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        scrollbar = ttk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        listbox = tk.Listbox(list_frame, yscrollcommand=scrollbar.set, font=('Consolas', 10))
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)

        # Populate listbox sorted by ID
        def update_list(*args):
            listbox.delete(0, tk.END)
            search_text = search_var.get().lower()
            sorted_ids = sorted(self.aadt_forecast_ids.keys(), key=lambda x: (not x.isdigit(), x))
            for id_str in sorted_ids:
                if search_text in id_str.lower():
                    value = self.aadt_forecast_ids[id_str]
                    listbox.insert(tk.END, f"{id_str}: {int(value) if value == int(value) else value}")

        search_var.trace('w', update_list)
        update_list()

        ttk.Button(popup, text="Close", command=popup.destroy).pack(pady=10)

    def scan_aadt_sections(self):
        """Scan project for all AADT sections in highway XMLs"""
        project_path = self.project_path.get()

        if self.placeholder_active or not project_path or not os.path.isdir(project_path):
            messagebox.showerror("Error", "Please select a valid project directory")
            return

        self.status_var.set("Scanning for AADT sections...")
        self.aadt_sections = []

        try:
            project_dir = Path(project_path)

            # Find all highway directories
            highway_dirs = []

            # Direct highway directories
            direct_highways = [d for d in project_dir.iterdir()
                              if d.is_dir() and d.name[0].lower() == 'h']
            highway_dirs.extend(direct_highways)

            # Highways inside interchanges (c* folders)
            interchange_dirs = [d for d in project_dir.iterdir()
                               if d.is_dir() and d.name[0].lower() == 'c']
            for interchange_dir in interchange_dirs:
                nested_highways = [d for d in interchange_dir.iterdir()
                                  if d.is_dir() and d.name[0].lower() == 'h']
                highway_dirs.extend(nested_highways)

            # Parse each highway XML for AADT data
            for highway_dir in highway_dirs:
                # Find the most recent highway.*.xml file
                highway_xmls = list(highway_dir.glob("highway.*.xml"))
                if not highway_xmls:
                    highway_xmls = list(highway_dir.glob("highway.xml"))
                if not highway_xmls:
                    continue

                # Use the most recent version (highest number)
                highway_xml = sorted(highway_xmls)[-1]

                try:
                    tree = ET.parse(highway_xml)
                    root = tree.getroot()

                    # Find Roadway element
                    ns = '{http://www.ihsdm.org/schema/Highway-1.0}'
                    roadway = root.find(f'.//{ns}Roadway')
                    if roadway is None:
                        roadway = root.find('.//Roadway')
                    if roadway is None:
                        continue

                    roadway_title = roadway.get('title', highway_dir.name)

                    # Find all AnnualAveDailyTraffic elements
                    aadt_elements = self.find_elements(roadway, 'AnnualAveDailyTraffic')

                    section_num = 1
                    for aadt in aadt_elements:
                        start_sta = aadt.get('startStation', '')
                        end_sta = aadt.get('endStation', '')
                        year = aadt.get('adtYear', '')
                        rate = aadt.get('adtRate', '1')

                        section = {
                            'roadway_title': roadway_title,
                            'highway_dir': str(highway_dir),
                            'xml_file': str(highway_xml),
                            'section_num': section_num,
                            'start_station': start_sta,
                            'end_station': end_sta,
                            'year': year,
                            'current_aadt': rate,
                            'id1': '',
                            'id2': '',
                            'id3': '',
                            'id4': '',
                            'id5': '',
                            'id6': '',
                            'sign1': '+',
                            'sign2': '+',
                            'sign3': '+',
                            'sign4': '+',
                            'sign5': '+',
                            'sign6': '+',
                            'calculated_aadt': ''
                        }
                        self.aadt_sections.append(section)
                        section_num += 1

                except Exception as e:
                    print(f"Error parsing {highway_xml}: {e}")
                    continue

            # Populate treeview - grouped by alignment (collapsible)
            self.aadt_tree.delete(*self.aadt_tree.get_children())
            self.aadt_reviewed_alignments = set()

            # Group sections by roadway_title
            alignments = {}
            for i, section in enumerate(self.aadt_sections):
                title = section['roadway_title']
                if title not in alignments:
                    alignments[title] = []
                alignments[title].append((i, section))

            # Create parent nodes for each alignment and child nodes for sections
            self.aadt_alignment_nodes = {}  # Map alignment title to tree node id
            for title in sorted(alignments.keys()):
                sections_list = alignments[title]
                # Create parent node for alignment
                parent_id = f"align_{title}"
                section_count = len(sections_list)
                self.aadt_tree.insert('', 'end', iid=parent_id, text=f"{title} ({section_count} sections)",
                                     values=('', '', '', '', '', '', ''),
                                     tags=('alignment', 'pending'), open=False)
                self.aadt_alignment_nodes[title] = parent_id

                # Create child nodes for each section
                for idx, section in sections_list:
                    # Initialize section with all 6 ID slots and signs
                    for i in range(1, 7):
                        if f'id{i}' not in section:
                            section[f'id{i}'] = ''
                        if f'sign{i}' not in section:
                            section[f'sign{i}'] = '+'

                    id_display = self.get_ids_display(section)
                    values = (
                        section['section_num'],
                        self.format_station(section['start_station']),
                        self.format_station(section['end_station']),
                        section['year'],
                        section['current_aadt'],
                        id_display,
                        section['calculated_aadt']
                    )
                    self.aadt_tree.insert(parent_id, 'end', iid=str(idx), text='',
                                         values=values, tags=('pending',))

            # Update checklist
            self.update_aadt_checklist()

            self.aadt_scan_status.set(f"Found {len(self.aadt_sections)} AADT sections across {len(alignments)} alignments")
            self.status_var.set(f"Scan complete: {len(self.aadt_sections)} AADT sections in {len(alignments)} alignments")

        except Exception as e:
            messagebox.showerror("Error", f"Error scanning project: {str(e)}")
            self.aadt_scan_status.set(f"Error: {str(e)}")

    def on_aadt_tree_select(self, event):
        """Handle tree selection - populate ID entry fields"""
        selection = self.aadt_tree.selection()
        if not selection:
            return

        item = selection[0]
        # Skip if alignment parent node selected
        if item.startswith('align_'):
            return

        try:
            idx = int(item)
            if idx < len(self.aadt_sections):
                section = self.aadt_sections[idx]
                self.aadt_id1_var.set(section.get('id1', ''))
                self.aadt_id2_var.set(section.get('id2', ''))
                self.aadt_id3_var.set(section.get('id3', ''))
                self.aadt_id4_var.set(section.get('id4', ''))
                self.aadt_id5_var.set(section.get('id5', ''))
                self.aadt_id6_var.set(section.get('id6', ''))
                # Set signs
                self.aadt_sign1_var.set(section.get('sign1', '+'))
                self.aadt_sign2_var.set(section.get('sign2', '+'))
                self.aadt_sign3_var.set(section.get('sign3', '+'))
                self.aadt_sign4_var.set(section.get('sign4', '+'))
                self.aadt_sign5_var.set(section.get('sign5', '+'))
                self.aadt_sign6_var.set(section.get('sign6', '+'))
        except ValueError:
            pass  # Not a section node

    def apply_aadt_ids_to_selected(self):
        """Apply entered forecast IDs to selected row"""
        selection = self.aadt_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select a section row first")
            return

        for item in selection:
            # Skip alignment parent nodes
            if item.startswith('align_'):
                continue

            try:
                idx = int(item)
                if idx < len(self.aadt_sections):
                    section = self.aadt_sections[idx]
                    section['id1'] = self.aadt_id1_var.get().strip()
                    section['id2'] = self.aadt_id2_var.get().strip()
                    section['id3'] = self.aadt_id3_var.get().strip()
                    section['id4'] = self.aadt_id4_var.get().strip()
                    section['id5'] = self.aadt_id5_var.get().strip()
                    section['id6'] = self.aadt_id6_var.get().strip()
                    section['sign1'] = self.aadt_sign1_var.get()
                    section['sign2'] = self.aadt_sign2_var.get()
                    section['sign3'] = self.aadt_sign3_var.get()
                    section['sign4'] = self.aadt_sign4_var.get()
                    section['sign5'] = self.aadt_sign5_var.get()
                    section['sign6'] = self.aadt_sign6_var.get()

                    # Calculate AADT for this section
                    calculated = self.calculate_section_aadt(section)
                    section['calculated_aadt'] = str(int(calculated)) if calculated > 0 else ''

                    # Update tree - get combined ID display
                    id_display = self.get_ids_display(section)

                    # Update tree
                    values = (
                        section['section_num'],
                        self.format_station(section['start_station']),
                        self.format_station(section['end_station']),
                        section['year'],
                        section['current_aadt'],
                        id_display,
                        section['calculated_aadt']
                    )
                    self.aadt_tree.item(item, values=values)
            except ValueError:
                continue

        self.status_var.set("Forecast IDs applied and AADT calculated")

    def get_ids_display(self, section):
        """Get a compact display of IDs with signs"""
        parts = []
        for i in range(1, 7):
            id_val = section.get(f'id{i}', '').strip()
            if id_val:
                sign = section.get(f'sign{i}', '+')
                if sign == '-':
                    parts.append(f'-{id_val}')
                else:
                    parts.append(f'+{id_val}' if parts else id_val)
        return ', '.join(parts) if parts else ''

    def calculate_section_aadt(self, section):
        """Calculate AADT for a section by summing/subtracting forecast ID values based on sign"""
        total = 0
        for i in range(1, 7):
            id_key = f'id{i}'
            sign_key = f'sign{i}'
            forecast_id = section.get(id_key, '').strip()
            sign = section.get(sign_key, '+')

            if forecast_id and forecast_id in self.aadt_forecast_ids:
                value = self.aadt_forecast_ids[forecast_id]
                if sign == '-':
                    total -= value
                else:
                    total += value
        return max(0, total)  # Don't allow negative AADT

    def calculate_all_aadt(self):
        """Calculate AADT for all sections that have forecast IDs assigned"""
        if not self.aadt_forecast_ids:
            messagebox.showwarning("No Forecast", "Please load forecast data first (Step 2)")
            return

        calculated_count = 0
        missing_ids = set()

        for i, section in enumerate(self.aadt_sections):
            # Check if section has any IDs assigned
            has_ids = False
            for j in range(1, 7):
                if section.get(f'id{j}', '').strip():
                    has_ids = True
                    # Track missing IDs for debugging
                    id_val = section.get(f'id{j}', '').strip()
                    if id_val and id_val not in self.aadt_forecast_ids:
                        missing_ids.add(id_val)

            if has_ids:
                calculated = self.calculate_section_aadt(section)
                section['calculated_aadt'] = str(int(calculated))
                calculated_count += 1
            else:
                section['calculated_aadt'] = ''

        # Update tree - iterate through all alignment parent nodes and their children
        for align_node in self.aadt_tree.get_children(''):
            for child_item in self.aadt_tree.get_children(align_node):
                try:
                    idx = int(child_item)
                    if idx < len(self.aadt_sections):
                        section = self.aadt_sections[idx]
                        id_display = self.get_ids_display(section)
                        values = (
                            section['section_num'],
                            self.format_station(section['start_station']),
                            self.format_station(section['end_station']),
                            section['year'],
                            section['current_aadt'],
                            id_display,
                            section['calculated_aadt']
                        )
                        self.aadt_tree.item(child_item, values=values)
                except (ValueError, tk.TclError) as e:
                    print(f"Error updating tree item {child_item}: {e}")

        status_msg = f"Calculated AADT for {calculated_count} of {len(self.aadt_sections)} sections"
        if missing_ids:
            status_msg += f" (Warning: {len(missing_ids)} IDs not found in forecast)"
            print(f"Missing forecast IDs: {missing_ids}")
            # Show popup with missing IDs to help user debug
            if len(missing_ids) <= 10:
                missing_list = ", ".join(sorted(missing_ids))
            else:
                missing_list = ", ".join(sorted(list(missing_ids)[:10])) + f"... and {len(missing_ids) - 10} more"
            messagebox.showwarning("Missing Forecast IDs",
                f"The following IDs were not found in the forecast workbook:\n\n{missing_list}\n\n"
                f"Make sure the ID values match exactly (case-sensitive).\n"
                f"Loaded forecast contains {len(self.aadt_forecast_ids)} IDs.")
        elif calculated_count == 0:
            messagebox.showinfo("No IDs Assigned",
                "No forecast IDs have been assigned to any sections yet.\n\n"
                "Select a section, enter Forecast IDs in the 'Assign IDs' section, "
                "then click 'Apply to Selected' before using Calculate All.")
        self.status_var.set(status_msg)

    def update_aadt_checklist(self):
        """Update the review checklist panel"""
        # Clear existing checklist items
        for widget in self.aadt_checklist_frame.winfo_children():
            widget.destroy()

        if not hasattr(self, 'aadt_alignment_nodes'):
            return

        # Add header
        ttk.Label(self.aadt_checklist_frame, text="Alignments:",
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w', pady=(0, 5))

        # Add each alignment with status
        for title in sorted(self.aadt_alignment_nodes.keys()):
            is_reviewed = title in self.aadt_reviewed_alignments
            status_icon = "✓" if is_reviewed else "○"
            status_color = 'green' if is_reviewed else 'gray'

            frame = ttk.Frame(self.aadt_checklist_frame)
            frame.pack(fill=tk.X, pady=1)

            lbl = tk.Label(frame, text=f"{status_icon} {title}",
                          font=('Segoe UI', 8),
                          fg=status_color,
                          anchor='w')
            lbl.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # Add summary at bottom
        total = len(self.aadt_alignment_nodes)
        reviewed = len(self.aadt_reviewed_alignments)
        ttk.Separator(self.aadt_checklist_frame, orient='horizontal').pack(fill=tk.X, pady=5)
        ttk.Label(self.aadt_checklist_frame,
                 text=f"Progress: {reviewed}/{total}",
                 font=('Segoe UI', 9, 'bold')).pack(anchor='w')

    def mark_alignment_reviewed(self):
        """Mark the currently selected alignment as reviewed"""
        selection = self.aadt_tree.selection()
        if not selection:
            messagebox.showwarning("No Selection", "Please select an alignment or section first")
            return

        item = selection[0]

        # Find the alignment title
        alignment_title = None
        if item.startswith('align_'):
            # Direct alignment selection
            alignment_title = item[6:]  # Remove 'align_' prefix
        else:
            # Section selected - get parent alignment
            try:
                idx = int(item)
                if idx < len(self.aadt_sections):
                    alignment_title = self.aadt_sections[idx]['roadway_title']
            except ValueError:
                pass

        if alignment_title:
            self.aadt_reviewed_alignments.add(alignment_title)

            # Update tree tags to show reviewed status
            if alignment_title in self.aadt_alignment_nodes:
                node_id = self.aadt_alignment_nodes[alignment_title]
                self.aadt_tree.item(node_id, tags=('alignment', 'reviewed'))

                # Also update all child sections
                for child in self.aadt_tree.get_children(node_id):
                    self.aadt_tree.item(child, tags=('reviewed',))

            # Update checklist
            self.update_aadt_checklist()
            self.status_var.set(f"Marked '{alignment_title}' as reviewed")

    def expand_all_aadt_tree(self):
        """Expand all alignment nodes"""
        for item in self.aadt_tree.get_children():
            self.aadt_tree.item(item, open=True)

    def collapse_all_aadt_tree(self):
        """Collapse all alignment nodes"""
        for item in self.aadt_tree.get_children():
            self.aadt_tree.item(item, open=False)

    def sort_aadt_tree(self, col):
        """Sort treeview - sorts alignments by name, and sections within each alignment by the clicked column"""
        # Determine sort order (toggle between ascending/descending)
        if not hasattr(self, '_aadt_sort_reverse'):
            self._aadt_sort_reverse = {}
        reverse = self._aadt_sort_reverse.get(col, False)

        # Column index mapping for section values
        col_index = {
            'Section': 0, 'Start Sta': 1, 'End Sta': 2, 'Year': 3,
            'Current': 4, 'Forecast IDs': 5, 'Calculated': 6
        }

        if col == '#0':
            # Sort alignment parent nodes by name
            items = [(self.aadt_tree.item(item)['text'], item)
                    for item in self.aadt_tree.get_children('')]
            items.sort(key=lambda x: x[0].lower(), reverse=reverse)
            for index, (_, item) in enumerate(items):
                self.aadt_tree.move(item, '', index)
        else:
            # Sort sections within each expanded alignment
            idx = col_index.get(col, 0)

            for align_node in self.aadt_tree.get_children(''):
                # Get children (sections) of this alignment
                children = self.aadt_tree.get_children(align_node)
                if not children:
                    continue

                # Get values and sort
                child_data = []
                for child in children:
                    values = self.aadt_tree.item(child)['values']
                    if values and len(values) > idx:
                        sort_val = values[idx]
                        # Try to convert to float for numeric sorting
                        try:
                            # Handle station format like "71+300.00"
                            if isinstance(sort_val, str) and '+' in sort_val:
                                sort_val = float(sort_val.replace('+', ''))
                            else:
                                sort_val = float(sort_val) if sort_val else 0
                        except (ValueError, TypeError):
                            sort_val = str(sort_val).lower() if sort_val else ''
                        child_data.append((sort_val, child))
                    else:
                        child_data.append(('', child))

                # Sort children
                child_data.sort(key=lambda x: x[0] if isinstance(x[0], (int, float)) else str(x[0]), reverse=reverse)

                # Rearrange children in sorted order
                for index, (_, child) in enumerate(child_data):
                    self.aadt_tree.move(child, align_node, index)

        # Toggle sort direction for next click
        self._aadt_sort_reverse[col] = not reverse
        self.status_var.set(f"Sorted by {col} ({'descending' if reverse else 'ascending'})")

    def preview_aadt_changes(self):
        """Preview changes that will be made to XML files"""
        changes = []
        for section in self.aadt_sections:
            if section['calculated_aadt']:
                changes.append(f"{section['roadway_title']} Section {section['section_num']}: "
                             f"Sta {self.format_station(section['start_station'])} to "
                             f"{self.format_station(section['end_station'])} → "
                             f"AADT {section['calculated_aadt']}")

        if not changes:
            messagebox.showinfo("No Changes", "No AADT values calculated yet. "
                              "Please enter forecast IDs and calculate first.")
            return

        # Show preview dialog
        preview_dialog = tk.Toplevel(self.root)
        preview_dialog.title("Preview AADT Changes")
        preview_dialog.geometry("700x500")
        preview_dialog.transient(self.root)

        ttk.Label(preview_dialog, text=f"The following {len(changes)} changes will be made:",
                 font=('Segoe UI', 10, 'bold')).pack(pady=10, padx=10, anchor='w')

        text_widget = scrolledtext.ScrolledText(preview_dialog, wrap=tk.WORD, font=('Consolas', 9))
        text_widget.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        text_widget.insert('1.0', '\n'.join(changes))
        text_widget.config(state=tk.DISABLED)

        ttk.Button(preview_dialog, text="Close", command=preview_dialog.destroy).pack(pady=10)

    def apply_aadt_to_xml(self):
        """Apply calculated AADT values to highway XML files"""
        # Group sections by XML file
        file_sections = {}
        for section in self.aadt_sections:
            if section['calculated_aadt']:
                xml_file = section['xml_file']
                if xml_file not in file_sections:
                    file_sections[xml_file] = []
                file_sections[xml_file].append(section)

        if not file_sections:
            messagebox.showwarning("No Changes", "No AADT values to apply. "
                                  "Please calculate AADT values first.")
            return

        # Confirm
        total_changes = sum(len(sections) for sections in file_sections.values())
        if not messagebox.askyesno("Confirm Changes",
                                   f"This will modify {total_changes} AADT values in {len(file_sections)} XML files.\n\n"
                                   "It is recommended to backup your project before proceeding.\n\n"
                                   "Continue?"):
            return

        self.status_var.set("Applying AADT values to XML files...")
        self.root.update()

        updated_count = 0
        error_count = 0

        for xml_file, sections in file_sections.items():
            try:
                # Read file as text for regex replacement (like R script)
                with open(xml_file, 'r', encoding='utf-8') as f:
                    content = f.read()

                original_content = content

                for section in sections:
                    new_aadt = section['calculated_aadt']
                    start_sta = section['start_station']
                    end_sta = section['end_station']
                    year = section['year']

                    # Build pattern to find the specific AnnualAveDailyTraffic element
                    # Pattern matches: <AnnualAveDailyTraffic startStation="X" endStation="Y" adtYear="Z" adtRate="W" />
                    import re

                    # Flexible pattern that matches the element with these station values
                    pattern = (
                        r'(<AnnualAveDailyTraffic\s+'
                        rf'startStation="{re.escape(start_sta)}"\s+'
                        rf'endStation="{re.escape(end_sta)}"\s*'
                        r'[^>]*?adtRate=")(\d+)(")'
                    )

                    def replace_aadt(match):
                        return match.group(1) + new_aadt + match.group(3)

                    content, count = re.subn(pattern, replace_aadt, content)

                    if count == 0:
                        # Try alternative attribute order
                        pattern2 = (
                            r'(<AnnualAveDailyTraffic\s+[^>]*?'
                            rf'startStation="{re.escape(start_sta)}"[^>]*?'
                            rf'endStation="{re.escape(end_sta)}"[^>]*?'
                            r'adtRate=")(\d+)(")'
                        )
                        content, count = re.subn(pattern2, replace_aadt, content)

                    if count > 0:
                        updated_count += 1

                # Write back if changes were made
                if content != original_content:
                    with open(xml_file, 'w', encoding='utf-8') as f:
                        f.write(content)

            except Exception as e:
                print(f"Error updating {xml_file}: {e}")
                error_count += 1

        self.aadt_apply_status.set(f"Updated {updated_count} AADT values. Errors: {error_count}")
        self.status_var.set(f"AADT update complete: {updated_count} values updated")

        if error_count > 0:
            messagebox.showwarning("Partial Success",
                                  f"Updated {updated_count} values with {error_count} errors.\n"
                                  "Check the console for error details.")
        else:
            messagebox.showinfo("Success", f"Successfully updated {updated_count} AADT values!")

    def save_aadt_mapping_csv(self):
        """Save AADT mapping progress to CSV for later resumption"""
        if not self.aadt_sections:
            messagebox.showwarning("No Data", "No AADT sections to save. Please scan the project first.")
            return

        file_path = filedialog.asksaveasfilename(
            title="Save AADT Mapping Progress",
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)

                # Headers - include all data needed to restore state
                headers = ['Roadway_Title', 'Section_Num', 'Start_Station', 'End_Station', 'Year',
                          'Current_AADT', 'ID1', 'Sign1', 'ID2', 'Sign2', 'ID3', 'Sign3',
                          'ID4', 'Sign4', 'ID5', 'Sign5', 'ID6', 'Sign6',
                          'Calculated_AADT', 'Reviewed', 'Highway_Dir', 'XML_File']
                writer.writerow(headers)

                # Data
                for section in self.aadt_sections:
                    is_reviewed = section['roadway_title'] in self.aadt_reviewed_alignments
                    row = [
                        section['roadway_title'],
                        section['section_num'],
                        section['start_station'],
                        section['end_station'],
                        section['year'],
                        section['current_aadt'],
                        section.get('id1', ''),
                        section.get('sign1', '+'),
                        section.get('id2', ''),
                        section.get('sign2', '+'),
                        section.get('id3', ''),
                        section.get('sign3', '+'),
                        section.get('id4', ''),
                        section.get('sign4', '+'),
                        section.get('id5', ''),
                        section.get('sign5', '+'),
                        section.get('id6', ''),
                        section.get('sign6', '+'),
                        section['calculated_aadt'],
                        'Yes' if is_reviewed else 'No',
                        section['highway_dir'],
                        section['xml_file']
                    ]
                    writer.writerow(row)

            messagebox.showinfo("Save Complete", f"AADT mapping saved to:\n{file_path}\n\nYou can load this file later to resume your work.")
            self.status_var.set(f"Mapping saved to {os.path.basename(file_path)}")

        except Exception as e:
            messagebox.showerror("Error", f"Error saving mapping: {str(e)}")

    def load_aadt_mapping_csv(self):
        """Load AADT mapping progress from a previously saved CSV"""
        file_path = filedialog.askopenfilename(
            title="Load AADT Mapping Progress",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)

                loaded_sections = []
                reviewed_alignments = set()

                for row in reader:
                    section = {
                        'roadway_title': row.get('Roadway_Title', ''),
                        'section_num': int(row.get('Section_Num', 0)),
                        'start_station': row.get('Start_Station', ''),
                        'end_station': row.get('End_Station', ''),
                        'year': row.get('Year', ''),
                        'current_aadt': row.get('Current_AADT', ''),
                        'id1': row.get('ID1', ''),
                        'sign1': row.get('Sign1', '+'),
                        'id2': row.get('ID2', ''),
                        'sign2': row.get('Sign2', '+'),
                        'id3': row.get('ID3', ''),
                        'sign3': row.get('Sign3', '+'),
                        'id4': row.get('ID4', ''),
                        'sign4': row.get('Sign4', '+'),
                        'id5': row.get('ID5', ''),
                        'sign5': row.get('Sign5', '+'),
                        'id6': row.get('ID6', ''),
                        'sign6': row.get('Sign6', '+'),
                        'calculated_aadt': row.get('Calculated_AADT', ''),
                        'highway_dir': row.get('Highway_Dir', ''),
                        'xml_file': row.get('XML_File', '')
                    }
                    loaded_sections.append(section)

                    if row.get('Reviewed', 'No') == 'Yes':
                        reviewed_alignments.add(section['roadway_title'])

            if not loaded_sections:
                messagebox.showwarning("No Data", "No sections found in the CSV file.")
                return

            # Update state with loaded data
            self.aadt_sections = loaded_sections
            self.aadt_reviewed_alignments = reviewed_alignments

            # Rebuild tree view with loaded data
            self.aadt_tree.delete(*self.aadt_tree.get_children())

            # Group sections by roadway_title
            alignments = {}
            for i, section in enumerate(self.aadt_sections):
                title = section['roadway_title']
                if title not in alignments:
                    alignments[title] = []
                alignments[title].append((i, section))

            # Create parent nodes for each alignment and child nodes for sections
            self.aadt_alignment_nodes = {}
            for title in sorted(alignments.keys()):
                sections_list = alignments[title]
                parent_id = f"align_{title}"
                section_count = len(sections_list)

                # Tag as reviewed if in reviewed set
                tags = ('alignment', 'reviewed') if title in reviewed_alignments else ('alignment', 'pending')
                self.aadt_tree.insert('', 'end', iid=parent_id, text=f"{title} ({section_count} sections)",
                                     values=('', '', '', '', '', '', ''),
                                     tags=tags, open=False)
                self.aadt_alignment_nodes[title] = parent_id

                # Create child nodes for each section
                for idx, section in sections_list:
                    id_display = self.get_ids_display(section)
                    values = (
                        section['section_num'],
                        self.format_station(section['start_station']),
                        self.format_station(section['end_station']),
                        section['year'],
                        section['current_aadt'],
                        id_display,
                        section['calculated_aadt']
                    )
                    self.aadt_tree.insert(parent_id, 'end', iid=str(idx), text='',
                                         values=values, tags=('pending',))

            # Update checklist
            self.update_aadt_checklist()

            section_count = len(self.aadt_sections)
            align_count = len(alignments)
            reviewed_count = len(reviewed_alignments)

            self.aadt_scan_status.set(f"Loaded {section_count} sections across {align_count} alignments ({reviewed_count} reviewed)")
            self.status_var.set(f"Loaded mapping from {os.path.basename(file_path)}")
            messagebox.showinfo("Load Complete",
                              f"Loaded {section_count} sections across {align_count} alignments.\n"
                              f"{reviewed_count} alignments marked as reviewed.\n\n"
                              f"Note: You still need to load the forecast workbook (Step 2) to calculate AADT values.")

        except Exception as e:
            messagebox.showerror("Error", f"Error loading mapping: {str(e)}")

    def export_aadt_mapping(self):
        """Export AADT mapping to Excel"""
        if not self.aadt_sections:
            messagebox.showwarning("No Data", "No AADT sections to export. Please scan the project first.")
            return

        file_path = filedialog.asksaveasfilename(
            title="Export AADT Mapping",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )

        if not file_path:
            return

        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "AADT Mapping"

            # Headers
            headers = ['Roadway_Title', 'Section', 'Start_Station', 'End_Station', 'Year',
                      'Current_AADT', 'Sign1', 'ID1', 'Sign2', 'ID2', 'Sign3', 'ID3',
                      'Sign4', 'ID4', 'Sign5', 'ID5', 'Sign6', 'ID6',
                      'Calculated_AADT', 'Reviewed', 'XML_File']
            ws.append(headers)

            # Data
            for section in self.aadt_sections:
                is_reviewed = section['roadway_title'] in self.aadt_reviewed_alignments
                row = [
                    section['roadway_title'],
                    section['section_num'],
                    section['start_station'],
                    section['end_station'],
                    section['year'],
                    section['current_aadt'],
                    section.get('sign1', '+'),
                    section.get('id1', ''),
                    section.get('sign2', '+'),
                    section.get('id2', ''),
                    section.get('sign3', '+'),
                    section.get('id3', ''),
                    section.get('sign4', '+'),
                    section.get('id4', ''),
                    section.get('sign5', '+'),
                    section.get('id5', ''),
                    section.get('sign6', '+'),
                    section.get('id6', ''),
                    section['calculated_aadt'],
                    'Yes' if is_reviewed else 'No',
                    section['xml_file']
                ]
                ws.append(row)

            wb.save(file_path)
            messagebox.showinfo("Export Complete", f"AADT mapping exported to:\n{file_path}")

        except Exception as e:
            messagebox.showerror("Error", f"Error exporting: {str(e)}")

    # =========================================================================
    # APPLICATION CONTROL
    # =========================================================================

    def run(self):
        """Start the application"""
        # Check for updates on startup (disabled due to firewall issues)
        # Users can manually check by clicking the version number in the top-right
        # self.root.after(1000, lambda: self.check_for_updates(show_current=False))
        self.root.mainloop()


def main():
    """Entry point"""
    app = IHSDMWisconsinHelper()
    app.run()


if __name__ == "__main__":
    main()
